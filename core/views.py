from datetime import datetime, timedelta
from decimal import Decimal
from fileinput import filename
import io
import re
import shutil
import subprocess
import tempfile
from urllib import response
import zipfile
from pathlib import Path

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import redirect
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.generic import TemplateView

from openpyxl import Workbook
from openpyxl.styles import Font

from employees.access import get_workspace_profile_url
from employees.models import (
    Employee,
    EmployeeActionRecord,
    EmployeeAttendanceEvent,
    EmployeeAttendanceLedger,
    EmployeeDocumentRequest,
    EmployeeLeave,
    EmployeeRequiredSubmission,
)
from hr.models import HRAnnouncement, HRPolicy
from organization.models import (
    Branch,
    BranchDocument,
    BranchDocumentRequirement,
    Company,
    Department,
    JobTitle,
    Section,
)
from payroll.models import PayrollBonus, PayrollLine, PayrollObligation, PayrollPeriod, PayrollProfile

class DashboardHomeView(LoginRequiredMixin, TemplateView):
    template_name = "core/dashboard_home.html"

    def get_employee_profile(self):
        user = self.request.user
        if not user or not user.is_authenticated:
            return None
        return (
            Employee.objects.filter(user=user)
            .select_related("company", "department", "branch", "section", "job_title")
            .first()
        )

    def is_admin_compatible(self, user):
        if not user or not user.is_authenticated:
            return False
        if getattr(user, "is_superuser", False):
            return True
        role_value = (getattr(user, "role", "") or "").strip().lower()
        if role_value in {"hr", "finance_manager", "supervisor", "operations_manager", "employee"}:
            return False
        return bool(getattr(user, "is_staff", False))

    def is_hr_user(self, user):
        return bool(user and user.is_authenticated and getattr(user, "is_hr", False))

    def is_finance_manager_user(self, user):
        return bool(user and user.is_authenticated and getattr(user, "is_finance_manager", False))

    def is_supervisor_user(self, user):
        return bool(user and user.is_authenticated and getattr(user, "is_supervisor", False))

    def is_operations_manager_user(self, user):
        return bool(user and user.is_authenticated and getattr(user, "is_operations_manager", False))

    def is_employee_role_user(self, user):
        return bool(user and user.is_authenticated and getattr(user, "is_employee_role", False))

    def is_management_user(self, user):
        return bool(
            user
            and user.is_authenticated
            and (
                self.is_admin_compatible(user)
                or self.is_hr_user(user)
                or self.is_finance_manager_user(user)
                or self.is_supervisor_user(user)
                or self.is_operations_manager_user(user)
            )
        )

    def can_access_dashboard(self, user):
        return bool(self.is_management_user(user) or self.is_employee_role_user(user))

    def handle_no_permission(self):
        user = getattr(self.request, "user", None)

        if not user or not user.is_authenticated:
            return super().handle_no_permission()

        messages.error(self.request, "You do not have permission to access the dashboard.")

        linked_employee = self.get_employee_profile()
        if linked_employee:
            return redirect(get_workspace_profile_url(user, linked_employee))

        raise PermissionDenied("You do not have permission to access the dashboard.")

    def dispatch(self, request, *args, **kwargs):
        if not self.can_access_dashboard(request.user):
            return self.handle_no_permission()
        return super().dispatch(request, *args, **kwargs)

    def get_scoped_branch(self, user, employee_profile):
        if (
            self.is_supervisor_user(user)
            and not self.is_admin_compatible(user)
            and not self.is_hr_user(user)
            and not self.is_operations_manager_user(user)
        ):
            return getattr(employee_profile, "branch", None)
        return None

    def is_branch_scoped_supervisor(self, user, scoped_branch):
        return bool(
            self.is_supervisor_user(user)
            and not self.is_admin_compatible(user)
            and not self.is_hr_user(user)
            and not self.is_operations_manager_user(user)
            and scoped_branch is not None
        )

    def build_branch_team_groups(self, employee):
        if not employee or not employee.branch_id:
            return []
        branch_team_members = list(
            Employee.objects.select_related("job_title", "section", "branch")
            .filter(branch_id=employee.branch_id, is_active=True)
            .order_by("full_name")
        )
        groups = {"Supervisor": [], "Team Leaders": [], "Team Members": []}
        for member in branch_team_members:
            title = (member.job_title.name if member.job_title else "").lower()
            if "supervisor" in title:
                groups["Supervisor"].append(member)
            elif "team leader" in title or title == "leader" or title.endswith(" leader"):
                groups["Team Leaders"].append(member)
            else:
                groups["Team Members"].append(member)
        return [{"label": label, "members": members} for label, members in groups.items() if members]

    def get_request_state_label(self, leave_record):
        if leave_record.status == EmployeeLeave.STATUS_APPROVED:
            return "Final Approved"
        if leave_record.status == EmployeeLeave.STATUS_REJECTED:
            return "Final Rejected"
        if leave_record.status == EmployeeLeave.STATUS_CANCELLED:
            return "Cancelled / Recalled"
        if leave_record.current_stage == EmployeeLeave.STAGE_SUPERVISOR_REVIEW:
            return "Waiting for Supervisor"
        if leave_record.current_stage == EmployeeLeave.STAGE_OPERATIONS_REVIEW:
            return "Approved by Supervisor • Waiting for Operations"
        if leave_record.current_stage == EmployeeLeave.STAGE_HR_REVIEW:
            return "Approved by Operations • Waiting for HR"
        return leave_record.get_status_display()

    def get_request_state_badge_class(self, leave_record):
        if leave_record.status == EmployeeLeave.STATUS_APPROVED:
            return "badge-success"
        if leave_record.status in {EmployeeLeave.STATUS_REJECTED, EmployeeLeave.STATUS_CANCELLED}:
            return "badge-danger"
        if leave_record.current_stage == EmployeeLeave.STAGE_SUPERVISOR_REVIEW:
            return "badge-warning"
        return "badge-primary"

    def build_metrics(self, queryset):
        total_employees = queryset.count()
        active_employees = queryset.filter(is_active=True).count()
        inactive_employees = queryset.filter(is_active=False).count()
        recent_hires_30_days = queryset.filter(
            hire_date__gte=timezone.localdate() - timedelta(days=30)
        ).count()

        def ratio(value):
            if not total_employees:
                return Decimal("0.0")
            return (Decimal(value) * Decimal("100") / Decimal(total_employees)).quantize(
                Decimal("0.1")
            )

        return {
            "total_employees": total_employees,
            "active_employees": active_employees,
            "inactive_employees": inactive_employees,
            "recent_hires_30_days": recent_hires_30_days,
            "active_ratio": ratio(active_employees),
            "inactive_ratio": ratio(inactive_employees),
            "total_companies": Company.objects.filter(is_active=True).count(),
            "total_departments": Department.objects.filter(is_active=True).count(),
            "total_branches": Branch.objects.filter(is_active=True).count(),
            "total_sections": Section.objects.filter(is_active=True).count(),
            "total_job_titles": JobTitle.objects.filter(is_active=True).count(),
        }

    def get_requirement_status_payload(self, selected_document):
        if not selected_document:
            return {"state_key": "missing", "status_label": "Missing"}
        if selected_document.is_expired:
            return {"state_key": "expired", "status_label": "Expired"}
        if selected_document.is_expiring_soon:
            return {"state_key": "expiring_soon", "status_label": "Expiring Soon"}
        if selected_document.expiry_date:
            return {"state_key": "valid", "status_label": "Valid"}
        return {"state_key": "recorded", "status_label": "Recorded"}

    def get_branch_compliance_status_payload(self, summary):
        requirement_total = summary.get("requirement_total", 0)
        missing_total = summary.get("missing_total", 0)
        expired_total = summary.get("expired_total", 0)
        expiring_soon_total = summary.get("expiring_soon_total", 0)
        compliant_total = summary.get("compliant_total", 0)

        if requirement_total == 0:
            return {
                "label": "No Checklist",
                "badge_class": "badge-light",
                "help_text": "No active required checklist items configured yet.",
            }
        if missing_total or expired_total:
            return {
                "label": "Critical",
                "badge_class": "badge-danger",
                "help_text": "At least one required document is missing or expired.",
            }
        if expiring_soon_total:
            return {
                "label": "Needs Attention",
                "badge_class": "badge-warning",
                "help_text": "Required documents exist, but one or more will expire soon.",
            }
        if compliant_total >= requirement_total:
            return {
                "label": "Compliant",
                "badge_class": "badge-success",
                "help_text": "All required checklist items are currently covered.",
            }
        return {
            "label": "In Review",
            "badge_class": "badge-primary",
            "help_text": "Checklist is partially covered and should be reviewed.",
        }

    def build_branch_compliance_dashboard(self):
        branches = list(
            Branch.objects.filter(is_active=True)
            .select_related("company")
            .annotate(
                employee_total=Count("employees", distinct=True),
                document_total=Count("documents", distinct=True),
            )
            .order_by("company__name", "name")
        )
        branch_ids = [branch.pk for branch in branches]

        requirements = list(
            BranchDocumentRequirement.objects.filter(branch_id__in=branch_ids, is_active=True)
            .select_related("branch", "branch__company")
            .order_by("branch__company__name", "branch__name", "document_type", "title")
        )
        documents = list(
            BranchDocument.objects.filter(branch_id__in=branch_ids)
            .select_related("branch", "branch__company")
            .order_by("branch_id", "document_type", "-issue_date", "-pk")
        )

        requirements_by_branch = {}
        for requirement in requirements:
            requirements_by_branch.setdefault(requirement.branch_id, []).append(requirement)

        documents_by_branch = {}
        for document in documents:
            documents_by_branch.setdefault(document.branch_id, []).append(document)

        rows = []
        summary = {
            "branch_total": len(branches),
            "requirement_total": 0,
            "missing_total": 0,
            "expired_total": 0,
            "expiring_soon_total": 0,
            "compliant_total": 0,
            "critical_total": 0,
            "needs_attention_total": 0,
            "compliant_branch_total": 0,
            "no_checklist_total": 0,
        }

        for branch in branches:
            latest_documents_by_type = {}
            for document in documents_by_branch.get(branch.pk, []):
                latest_documents_by_type.setdefault(document.document_type, document)

            requirement_total = 0
            missing_total = 0
            expired_total = 0
            expiring_soon_total = 0
            compliant_total = 0

            for requirement in requirements_by_branch.get(branch.pk, []):
                if not requirement.is_mandatory:
                    continue
                requirement_total += 1
                selected_document = latest_documents_by_type.get(requirement.document_type)
                status_payload = self.get_requirement_status_payload(selected_document)
                if status_payload["state_key"] == "missing":
                    missing_total += 1
                elif status_payload["state_key"] == "expired":
                    expired_total += 1
                elif status_payload["state_key"] == "expiring_soon":
                    expiring_soon_total += 1
                elif status_payload["state_key"] in {"valid", "recorded"}:
                    compliant_total += 1

            compliance_percentage = (
                int(round((compliant_total / requirement_total) * 100)) if requirement_total else 0
            )
            status_payload = self.get_branch_compliance_status_payload(
                {
                    "requirement_total": requirement_total,
                    "missing_total": missing_total,
                    "expired_total": expired_total,
                    "expiring_soon_total": expiring_soon_total,
                    "compliant_total": compliant_total,
                }
            )
            row = {
                "branch_id": branch.pk,
                "branch_name": branch.name,
                "company_name": branch.company.name if getattr(branch, "company_id", None) else "—",
                "employee_total": getattr(branch, "employee_total", 0),
                "document_total": getattr(branch, "document_total", 0),
                "requirement_total": requirement_total,
                "missing_total": missing_total,
                "expired_total": expired_total,
                "expiring_soon_total": expiring_soon_total,
                "compliant_total": compliant_total,
                "compliance_percentage": compliance_percentage,
                "status_label": status_payload["label"],
                "status_badge_class": status_payload["badge_class"],
                "status_help_text": status_payload["help_text"],
                "detail_url": f"/organization/branches/{branch.pk}/",
                "document_center_url": f"/organization/branch-documents/?branch={branch.pk}",
            }
            rows.append(row)
            summary["requirement_total"] += requirement_total
            summary["missing_total"] += missing_total
            summary["expired_total"] += expired_total
            summary["expiring_soon_total"] += expiring_soon_total
            summary["compliant_total"] += compliant_total
            if row["status_label"] == "Critical":
                summary["critical_total"] += 1
            elif row["status_label"] == "Needs Attention":
                summary["needs_attention_total"] += 1
            elif row["status_label"] == "Compliant":
                summary["compliant_branch_total"] += 1
            elif row["status_label"] == "No Checklist":
                summary["no_checklist_total"] += 1

        metric_cards = [
            {
                "label": "Branches",
                "value": summary["branch_total"],
                "help_text": "Active branch records in compliance monitoring.",
                "card_class": "dashboard-core-metric-card",
            },
            {
                "label": "Compliant",
                "value": summary["compliant_branch_total"],
                "help_text": "Branches with all required checklist items covered.",
                "card_class": "dashboard-core-metric-card dashboard-core-metric-success",
            },
            {
                "label": "Needs Attention",
                "value": summary["needs_attention_total"],
                "help_text": "Branches with required documents expiring soon.",
                "card_class": "dashboard-core-metric-card dashboard-core-metric-warning",
            },
            {
                "label": "Critical",
                "value": summary["critical_total"],
                "help_text": "Branches with missing or expired required documents.",
                "card_class": "dashboard-core-metric-card dashboard-core-metric-danger",
            },
        ]

        quick_stats = [
            {"label": "Missing Items", "value": summary["missing_total"]},
            {"label": "Expired Items", "value": summary["expired_total"]},
            {"label": "Expiring Soon", "value": summary["expiring_soon_total"]},
            {"label": "No Checklist", "value": summary["no_checklist_total"]},
            {"label": "Covered Items", "value": summary["compliant_total"]},
        ]

        quick_links = [
            {"label": "Open Branch Compliance Overview", "url": "/organization/branches/"},
            {"label": "Open Branch Documents Center", "url": "/organization/branch-documents/"},
        ]

        critical_rows = [row for row in rows if row["status_label"] == "Critical"][:5]
        expiring_rows = [row for row in rows if row["expiring_soon_total"] > 0][:5]

        return {
            "summary": summary,
            "metric_cards": metric_cards,
            "quick_stats": quick_stats,
            "critical_rows": critical_rows,
            "expiring_rows": expiring_rows,
            "quick_links": quick_links,
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        role_value = (getattr(user, "role", "") or "").strip().lower()
        today = timezone.localdate()
        employee_profile = self.get_employee_profile()
        scoped_branch = self.get_scoped_branch(user, employee_profile)
        supervisor_setup_required = bool(
            self.is_supervisor_user(user) and employee_profile and scoped_branch is None
        )
        supervisor_scope_missing = supervisor_setup_required

        is_employee_self_service_dashboard = bool(
            self.is_employee_role_user(user)
            and employee_profile
            and not (
                self.is_admin_compatible(user)
                or self.is_hr_user(user)
                or self.is_operations_manager_user(user)
                or scoped_branch
            )
        )
        is_executive_dashboard = bool(
            (self.is_admin_compatible(user) or getattr(user, "is_superuser", False))
            and scoped_branch is None
            and not is_employee_self_service_dashboard
        )
        is_finance_dashboard = bool(
            role_value == "finance_manager"
            and not is_employee_self_service_dashboard
            and not is_executive_dashboard
        )

        if is_employee_self_service_dashboard:
            leave_qs = employee_profile.leave_records.all().order_by("-created_at", "-id")
            request_state_records = []
            for leave_record in leave_qs[:10]:
                request_state_records.append(
                    {
                        "record": leave_record,
                        "state_label": self.get_request_state_label(leave_record),
                        "state_badge_class": self.get_request_state_badge_class(leave_record),
                    }
                )

            context.update(
                {
                    "is_employee_self_service_dashboard": True,
                    "employee": employee_profile,
                    "pending_request_count": leave_qs.filter(status=EmployeeLeave.STATUS_PENDING).count(),
                    "approved_request_count": leave_qs.filter(status=EmployeeLeave.STATUS_APPROVED).count(),
                    "rejected_request_count": leave_qs.filter(status=EmployeeLeave.STATUS_REJECTED).count(),
                    "cancelled_request_count": leave_qs.filter(status=EmployeeLeave.STATUS_CANCELLED).count(),
                    "branch_team_groups": self.build_branch_team_groups(employee_profile),
                    "request_state_records": request_state_records,
                    "attendance_event_today": employee_profile.attendance_events.filter(attendance_date=timezone.localdate()).first(),
                    "self_service_quick_links": [
                        {"label": "My Workspace", "url": get_workspace_profile_url(user, employee_profile)},
                        {"label": "My Leave", "url": reverse("employees:self_service_leave")},
                        {"label": "My Documents", "url": reverse("employees:self_service_documents")},
                        {"label": "My Attendance", "url": reverse("employees:self_service_attendance")},
                        {"label": "Working Time", "url": reverse("employees:self_service_working_time")},
                    ],
                }
            )
            return context

        employee_queryset = Employee.objects.select_related(
            "company", "department", "branch", "section", "job_title"
        ).all()
        if scoped_branch:
            employee_queryset = employee_queryset.filter(branch_id=scoped_branch.id)

        recent_employees = list(employee_queryset.order_by("-created_at", "-id")[:8])
        recent_hires = list(
            employee_queryset.filter(hire_date__isnull=False).order_by("-hire_date", "-id")[:8]
        )
        recent_employees = list(employee_queryset.order_by("-created_at", "-id")[:8])
        employees_by_company = list(
            employee_queryset.exclude(company__isnull=True)
            .values("company__name")
            .annotate(total=Count("id"))
            .order_by("-total", "company__name")[:8]
        )
        employees_by_branch = list(
            employee_queryset.exclude(branch__isnull=True)
            .values("branch__name")
            .annotate(total=Count("id"))
            .order_by("-total", "branch__name")[:8]
        )
        employees_by_department = list(
            employee_queryset.exclude(department__isnull=True)
            .values("department__name")
            .annotate(total=Count("id"))
            .order_by("-total", "department__name")[:8]
        )
        payroll_profile_queryset = PayrollProfile.objects.select_related("employee", "company").order_by("employee__full_name")
        finance_review_periods = list(
            PayrollPeriod.objects.select_related("company")
            .filter(status=PayrollPeriod.STATUS_REVIEW)
            .order_by("-period_start", "-id")[:6]
        )
        finance_ready_periods = list(
            PayrollPeriod.objects.select_related("company")
            .filter(status=PayrollPeriod.STATUS_APPROVED)
            .order_by("-period_start", "-id")[:6]
        )
        finance_recent_paid_periods = list(
            PayrollPeriod.objects.select_related("company")
            .filter(status=PayrollPeriod.STATUS_PAID)
            .order_by("-pay_date", "-period_end", "-id")[:6]
        )
        finance_active_obligations = list(
            PayrollObligation.objects.select_related("employee", "company")
            .filter(status=PayrollObligation.STATUS_ACTIVE)
            .order_by("-updated_at", "-id")[:6]
        )
        finance_active_bonuses = list(
            PayrollBonus.objects.select_related("employee", "company")
            .filter(status=PayrollBonus.STATUS_ACTIVE)
            .order_by("-updated_at", "-id")[:6]
        )
        finance_recent_lines = list(
            PayrollLine.objects.select_related("employee", "employee__company", "payroll_period", "payroll_period__company")
            .filter(payroll_period__status__in=[PayrollPeriod.STATUS_REVIEW, PayrollPeriod.STATUS_APPROVED, PayrollPeriod.STATUS_PAID])
            .order_by("-payroll_period__period_start", "employee__full_name")[:8]
        )
        finance_hold_profiles = list(
            payroll_profile_queryset.filter(status=PayrollProfile.STATUS_HOLD)[:6]
        )
        finance_profile_total = payroll_profile_queryset.count()
        finance_hold_profile_total = payroll_profile_queryset.filter(status=PayrollProfile.STATUS_HOLD).count()
        finance_active_obligation_total = PayrollObligation.objects.filter(
            status=PayrollObligation.STATUS_ACTIVE
        ).count()
        finance_active_bonus_total = PayrollBonus.objects.filter(
            status=PayrollBonus.STATUS_ACTIVE
        ).count()
        finance_outstanding_obligation_balance = sum(
            obligation.remaining_balance
            for obligation in PayrollObligation.objects.filter(status=PayrollObligation.STATUS_ACTIVE)
        )
        finance_outstanding_bonus_balance = sum(
            bonus.remaining_balance
            for bonus in PayrollBonus.objects.filter(status=PayrollBonus.STATUS_ACTIVE)
        )
        finance_status_cards = [
            {
                "label": "In Review",
                "value": PayrollPeriod.objects.filter(status=PayrollPeriod.STATUS_REVIEW).count(),
                "tone": "warning",
                "help_text": "Payroll periods waiting for finance approval.",
            },
            {
                "label": "Ready For Payment",
                "value": PayrollPeriod.objects.filter(status=PayrollPeriod.STATUS_APPROVED).count(),
                "tone": "primary",
                "help_text": "Approved payroll periods still waiting to be marked paid.",
            },
            {
                "label": "Active Bonuses",
                "value": finance_active_bonus_total,
                "tone": "primary",
                "help_text": "Bonus balances still available to be applied.",
            },
            {
                "label": "Active Obligations",
                "value": finance_active_obligation_total,
                "tone": "warning",
                "help_text": "Loans and advances still affecting payroll lines.",
            },
        ]
        finance_quick_links = [
            {"label": "Open Payroll Workspace", "url": reverse("payroll:home")},
            {"label": "Notifications", "url": reverse("notifications:home")},
            {"label": "Approval Queue", "url": "#finance-approval-queue"},
            {"label": "Payslip Output", "url": "#finance-payslip-output"},
            {"label": "Payment Status", "url": "#finance-payment-status"},
        ]

        can_view_organization_setup = bool(
            self.is_admin_compatible(user)
            or self.is_hr_user(user)
            or self.is_operations_manager_user(user)
        )
        leave_queryset = EmployeeLeave.objects.select_related("employee", "employee__branch")
        document_request_queryset = EmployeeDocumentRequest.objects.select_related("employee", "employee__branch")
        attendance_event_queryset = EmployeeAttendanceEvent.objects.select_related("employee", "employee__branch")
        attendance_ledger_queryset = EmployeeAttendanceLedger.objects.select_related("employee", "employee__branch")

        if scoped_branch:
            leave_queryset = leave_queryset.filter(employee__branch_id=scoped_branch.id)
            document_request_queryset = document_request_queryset.filter(employee__branch_id=scoped_branch.id)
            attendance_event_queryset = attendance_event_queryset.filter(employee__branch_id=scoped_branch.id)
            attendance_ledger_queryset = attendance_ledger_queryset.filter(employee__branch_id=scoped_branch.id)

        pending_leave_queue = list(
            leave_queryset.filter(status=EmployeeLeave.STATUS_PENDING).order_by("-created_at", "-id")[:6]
        )
        open_document_request_queue = list(
            document_request_queryset.filter(status=EmployeeDocumentRequest.STATUS_REQUESTED).order_by("-created_at", "-id")[:6]
        )
        open_attendance_sessions = list(
            attendance_event_queryset.filter(status=EmployeeAttendanceEvent.STATUS_OPEN).order_by("-check_in_at", "-id")[:6]
        )
        attendance_exceptions = list(
            attendance_ledger_queryset.filter(
                attendance_date=today,
                day_status__in=[
                    EmployeeAttendanceLedger.DAY_STATUS_ABSENT,
                    EmployeeAttendanceLedger.DAY_STATUS_UNPAID_LEAVE,
                    EmployeeAttendanceLedger.DAY_STATUS_OTHER,
                ],
            ).order_by("employee__full_name")[:6]
        )
        recent_leave_activity = list(
            leave_queryset.order_by("-updated_at", "-created_at", "-id")[:6]
        )
        operations_quick_links = []
        if scoped_branch:
            operations_quick_links.append({"label": "My Branch", "url": reverse("employees:self_service_branch")})
            operations_quick_links.append({"label": "Weekly Schedule", "url": reverse("employees:self_service_weekly_schedule")})
        if (
            self.is_admin_compatible(user)
            or self.is_hr_user(user)
            or self.is_operations_manager_user(user)
            or scoped_branch
        ):
            operations_quick_links.append({"label": "Attendance Management", "url": reverse("employees:attendance_management")})
            if self.is_admin_compatible(user) or self.is_hr_user(user) or self.is_operations_manager_user(user):
                operations_quick_links.append({"label": "Branch Schedules", "url": reverse("employees:branch_schedule_overview")})
            operations_quick_links.append({"label": "Action Center", "url": reverse("employees:employee_admin_action_center")})
            operations_quick_links.append({"label": "Employee Requests", "url": reverse("employees:employee_requests_overview")})
        operational_status_cards = [
            {
                "label": "Open Attendance",
                "value": len(open_attendance_sessions),
                "tone": "neutral",
                "help_text": "Employees still checked in and not checked out yet.",
            },
            {
                "label": "Attendance Exceptions",
                "value": len(attendance_exceptions),
                "tone": "warning",
                "help_text": "Absence or unpaid leave flags from today's attendance ledger.",
            },
            {
                "label": "Pending Leave",
                "value": len(pending_leave_queue),
                "tone": "warning",
                "help_text": "Leave approvals still waiting for review.",
            },
            {
                "label": "Requested Documents",
                "value": len(open_document_request_queue),
                "tone": "primary",
                "help_text": "Employee document requests currently open.",
            },
        ]
        branch_compliance_dashboard = (
            self.build_branch_compliance_dashboard() if can_view_organization_setup else None
        )
        executive_quick_links = [
            {"label": "HR Control Center", "url": reverse("hr:home")},
            {"label": "Payroll Workspace", "url": reverse("payroll:home")},
            {"label": "Notifications", "url": reverse("notifications:home")},
            {"label": "Branch Schedules", "url": reverse("employees:branch_schedule_overview")},
            {"label": "Attendance Management", "url": reverse("employees:attendance_management")},
            {"label": "Action Center", "url": reverse("employees:employee_admin_action_center")},
            {"label": "Organization Setup", "url": reverse("organization:company_list")},
        ]
        pending_submission_queue = list(
            EmployeeRequiredSubmission.objects.exclude(
                status=EmployeeRequiredSubmission.STATUS_COMPLETED
            ).select_related("employee").order_by("-updated_at", "-created_at")[:6]
        )
        executive_document_queue = list(
            EmployeeDocumentRequest.objects.filter(
                status=EmployeeDocumentRequest.STATUS_REQUESTED
            ).select_related("employee").order_by("-created_at")[:6]
        )
        executive_recent_actions = list(
            EmployeeActionRecord.objects.select_related("employee").order_by("-action_date", "-created_at")[:6]
        )
        executive_payroll_periods = list(
            PayrollPeriod.objects.select_related("company").order_by("-period_start", "-id")[:6]
        )
        executive_active_obligations = list(
            PayrollObligation.objects.select_related("employee", "company").filter(
                status=PayrollObligation.STATUS_ACTIVE
            ).order_by("-updated_at", "-id")[:6]
        )
        executive_policies = list(
            HRPolicy.objects.select_related("company").filter(is_active=True).order_by("title")[:6]
        )
        executive_announcements = list(
            HRAnnouncement.objects.filter(is_active=True).order_by("-published_at", "-id")[:5]
        )
        executive_status_cards = [
            {"label": "Total Workforce", "value": employee_queryset.count(), "tone": "neutral"},
            {"label": "Pending Leave", "value": EmployeeLeave.objects.filter(status=EmployeeLeave.STATUS_PENDING).count(), "tone": "warning"},
            {"label": "Attendance Exceptions", "value": EmployeeAttendanceLedger.objects.filter(
                attendance_date=today,
                day_status__in=[
                    EmployeeAttendanceLedger.DAY_STATUS_ABSENT,
                    EmployeeAttendanceLedger.DAY_STATUS_UNPAID_LEAVE,
                    EmployeeAttendanceLedger.DAY_STATUS_OTHER,
                ],
            ).count(), "tone": "warning"},
            {"label": "Open Document Requests", "value": EmployeeDocumentRequest.objects.filter(
                status=EmployeeDocumentRequest.STATUS_REQUESTED
            ).count(), "tone": "primary"},
            {"label": "Open Submissions", "value": EmployeeRequiredSubmission.objects.exclude(
                status=EmployeeRequiredSubmission.STATUS_COMPLETED
            ).count(), "tone": "primary"},
            {"label": "Live Payroll Cycles", "value": PayrollPeriod.objects.exclude(status=PayrollPeriod.STATUS_PAID).count(), "tone": "neutral"},
            {"label": "Active Obligations", "value": PayrollObligation.objects.filter(status=PayrollObligation.STATUS_ACTIVE).count(), "tone": "warning"},
            {"label": "Policies", "value": HRPolicy.objects.filter(is_active=True).count(), "tone": "neutral"},
        ]

        context.update(
            {
                "is_employee_self_service_dashboard": False,
                "is_finance_dashboard": is_finance_dashboard,
                "is_executive_dashboard": is_executive_dashboard,
                "is_branch_scoped_supervisor": scoped_branch is not None,
                "scoped_branch": scoped_branch,
                "metrics": self.build_metrics(employee_queryset),
                "recent_employees": recent_employees,
                "recent_hires": recent_hires,
                "employees_by_company": employees_by_company,
                "employees_by_branch": employees_by_branch,
                "employees_by_department": employees_by_department,
                "operations_today": today,
                "pending_leave_queue": pending_leave_queue,
                "open_document_request_queue": open_document_request_queue,
                "open_attendance_sessions": open_attendance_sessions,
                "attendance_exceptions": attendance_exceptions,
                "recent_leave_activity": recent_leave_activity,
                "operations_quick_links": operations_quick_links,
                "operational_status_cards": operational_status_cards,
                "executive_quick_links": executive_quick_links,
                "executive_status_cards": executive_status_cards,
                "pending_submission_queue": pending_submission_queue,
                "executive_document_queue": executive_document_queue,
                "executive_recent_actions": executive_recent_actions,
                "executive_payroll_periods": executive_payroll_periods,
                "executive_active_obligations": executive_active_obligations,
                "executive_policies": executive_policies,
                "executive_announcements": executive_announcements,
                "finance_quick_links": finance_quick_links,
                "finance_status_cards": finance_status_cards,
                "finance_review_periods": finance_review_periods,
                "finance_ready_periods": finance_ready_periods,
                "finance_recent_paid_periods": finance_recent_paid_periods,
                "finance_recent_lines": finance_recent_lines,
                "finance_active_obligations": finance_active_obligations,
                "finance_active_bonuses": finance_active_bonuses,
                "finance_hold_profiles": finance_hold_profiles,
                "finance_profile_total": finance_profile_total,
                "finance_hold_profile_total": finance_hold_profile_total,
                "finance_active_obligation_total": finance_active_obligation_total,
                "finance_active_bonus_total": finance_active_bonus_total,
                "finance_outstanding_obligation_balance": finance_outstanding_obligation_balance,
                "finance_outstanding_bonus_balance": finance_outstanding_bonus_balance,
                "can_view_employee_directory": bool(
                    self.is_admin_compatible(user)
                    or self.is_hr_user(user)
                    or self.is_operations_manager_user(user)
                    or scoped_branch
                ),
                "can_view_organization_setup": can_view_organization_setup,
                "show_branch_compliance_dashboard": bool(
                    can_view_organization_setup and not scoped_branch
                ),
                "branch_compliance_dashboard": branch_compliance_dashboard,
                "supervisor_setup_required": supervisor_setup_required,
                "supervisor_scope_missing": supervisor_scope_missing,
                "management_open_attendance_events": attendance_event_queryset.filter(status=EmployeeAttendanceEvent.STATUS_OPEN).count(),
            }
        )
        return context


class BackupCenterView(LoginRequiredMixin, TemplateView):
    template_name = "core/backup_center.html"
    backup_table_limit = 20

    def dispatch(self, request, *args, **kwargs):
        if not self.can_access_backup_center(request.user):
            messages.error(request, "Only the top admin can access the Backup & Export Center.")
            if request.user.is_authenticated:
                return redirect("home")
            raise PermissionDenied("You do not have permission to access the Backup & Export Center.")
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        action = (request.POST.get("action") or "create_backup_server").strip()

        if action in {"create_backup", "create_backup_server"}:
            note = (request.POST.get("backup_note") or "").strip()
            try:
                backup_file, database_backup_warning = self.create_backup(note=note)
            except Exception as exc:
                messages.error(request, f"Backup creation failed: {exc}")
            else:
                messages.success(request, f"Backup created successfully on server: {backup_file.name}")
                if database_backup_warning:
                    messages.warning(
                        request,
                        f"Backup was created, but the database dump was not included. {database_backup_warning}",
                    )
            return redirect("backup_center")

        if action == "download_backup_now":
            note = (request.POST.get("backup_note") or "").strip()
            try:
                return self.download_backup_response(note=note)
            except Exception as exc:
                messages.error(request, f"Backup download failed: {exc}")
                return redirect("backup_center")

        if action == "check_database_backup":
            try:
                result = self.run_database_backup_preflight()
            except Exception as exc:
                messages.error(request, f"Database backup pre-check failed: {exc}")
            else:
                if result.get("ok"):
                    messages.success(request, result["message"])
                else:
                    messages.warning(request, result["message"])
            return redirect("backup_center")

        if action == "export_employee_master":
            return self.export_employee_master_data()

        if action == "export_attendance":
            return self.export_attendance_data(
                start_date=parse_date((request.POST.get("attendance_start_date") or "").strip())
                if (request.POST.get("attendance_start_date") or "").strip()
                else None,
                end_date=parse_date((request.POST.get("attendance_end_date") or "").strip())
                if (request.POST.get("attendance_end_date") or "").strip()
                else None,
            )

        if action == "export_leave_records":
            return self.export_leave_data(
                start_date=parse_date((request.POST.get("leave_start_date") or "").strip())
                if (request.POST.get("leave_start_date") or "").strip()
                else None,
                end_date=parse_date((request.POST.get("leave_end_date") or "").strip())
                if (request.POST.get("leave_end_date") or "").strip()
                else None,
            )

        if action == "export_backup_audit":
            return self.export_backup_audit_data()

        messages.error(request, "Unknown utility action requested.")
        return redirect("backup_center")

    def can_access_backup_center(self, user):
        return bool(user and user.is_authenticated and getattr(user, "is_superuser", False))

    def get_backup_root(self):
        backup_root = Path(settings.HR_BACKUP_ROOT)
        backup_root.mkdir(parents=True, exist_ok=True)
        return backup_root

    def sanitize_note(self, note):
        cleaned = re.sub(r"[^A-Za-z0-9_-]+", "-", note.strip())
        cleaned = re.sub(r"-+", "-", cleaned).strip("-")
        return cleaned[:50]

    def get_backup_recent_update_items(self):
        return [
            {
                "title": "Employee workspace refactor",
                "summary": "The large employee view is now split into shared, directory, self-service, management, action-center, and API modules for safer maintenance.",
                "paths": [
                    "employees/views.py",
                    "employees/views_shared.py",
                    "employees/views_directory.py",
                    "employees/views_self_service.py",
                    "employees/views_management.py",
                    "employees/views_action_center.py",
                    "employees/views_api.py",
                ],
            },
            {
                "title": "Payroll hardening",
                "summary": "Payroll now includes approval snapshots, approved-paid edit locks, overtime calculation, unpaid leave deduction, calculation trace, PDF payslips, and paid-period delivery.",
                "paths": [
                    "payroll/views.py",
                    "payroll/tests.py",
                    "templates/payroll/payslip.html",
                    "templates/payroll/payslip_pdf.html",
                ],
            },
            {
                "title": "Notification platform",
                "summary": "The app now has an in-app notification center with category filters, preferences, payroll delivery rules, and request-tracking alerts across key workflows.",
                "paths": [
                    "notifications/models.py",
                    "notifications/views.py",
                    "notifications/forms.py",
                    "templates/notifications/center.html",
                ],
            },
            {
                "title": "System tracking coverage",
                "summary": "Notifications now cover employee requests, branch operations, HR publications, work calendar changes, and employee status-schedule updates.",
                "paths": [
                    "employees/views.py",
                    "operations/views.py",
                    "hr/signals.py",
                    "workcalendar/views.py",
                ],
            },
        ]

    def get_backup_next_focus_items(self):
        return [
            "Expand notifications into any still-silent workflows, especially future finance or compliance modules.",
            "Add feed pagination, bulk read actions, and richer category unread summaries inside the notification center.",
            "Start the finance foundation or Kuwait compliance layer only after the current notification and payroll flows feel stable in daily use.",
        ]

    def get_include_paths(self):
        include_items = []

        for raw_item in getattr(settings, "HR_BACKUP_INCLUDE_PATHS", []):
            item = str(raw_item).strip()
            if not item:
                continue

            if item == "media":
                candidate = Path(settings.MEDIA_ROOT)
                archive_root = "media"
                display_label = "media"
            else:
                candidate = Path(item)
                if not candidate.is_absolute():
                    candidate = settings.BASE_DIR / item
                archive_root = item.replace("\\", "/").strip("/")
                display_label = item

            if candidate.exists():
                include_items.append(
                    {
                        "path": candidate,
                        "archive_root": archive_root,
                        "display_label": display_label,
                    }
                )

        return include_items

    def build_archive_name(self, include_root, archive_root, child_path=None):
        archive_root = str(archive_root).replace("\\", "/").strip("/")

        if child_path is None:
            return archive_root

        relative_child = child_path.relative_to(include_root).as_posix()
        if archive_root:
            return f"{archive_root}/{relative_child}"
        return relative_child

    def get_database_settings(self):
        return (getattr(settings, "DATABASES", {}) or {}).get("default", {})

    def get_database_engine(self):
        return (self.get_database_settings().get("ENGINE") or "").lower()

    def is_postgres_database(self):
        return "postgresql" in self.get_database_engine() or "postgres" in self.get_database_engine()

    def is_sqlite_database(self):
        return "sqlite" in self.get_database_engine()

    def get_pg_dump_command(self):
        configured_command = str(
            getattr(settings, "HR_BACKUP_PG_DUMP_COMMAND", "pg_dump") or "pg_dump"
        ).strip()
        if not configured_command:
            configured_command = "pg_dump"
        if Path(configured_command).exists():
            return configured_command
        resolved_command = shutil.which(configured_command)
        return resolved_command or configured_command

    def build_postgres_conninfo(self):
        database_settings = self.get_database_settings()
        options = database_settings.get("OPTIONS") or {}
        conninfo_parts = []
        field_map = {
            "NAME": "dbname",
            "USER": "user",
            "PASSWORD": "password",
            "HOST": "host",
            "PORT": "port",
        }
        for field_name, conninfo_key in field_map.items():
            field_value = str(database_settings.get(field_name) or "").strip()
            if field_value:
                conninfo_parts.append(f"{conninfo_key}={field_value}")
        sslmode = str(options.get("sslmode") or "").strip()
        if sslmode:
            conninfo_parts.append(f"sslmode={sslmode}")
        if not conninfo_parts:
            raise ValueError("Database connection settings are missing for PostgreSQL backup.")
        return " ".join(conninfo_parts)

    def get_database_backup_status(self):
        if self.is_postgres_database():
            pg_dump_command = self.get_pg_dump_command()
            pg_dump_available = Path(pg_dump_command).exists() or shutil.which(pg_dump_command) is not None
            return {
                "mode": "PostgreSQL dump",
                "engine": "PostgreSQL",
                "archive_name": "database/postgresql_dump.sql",
                "available": pg_dump_available,
                "help_text": "Full backup includes a live PostgreSQL SQL dump together with project files and media.",
            }
        if self.is_sqlite_database():
            database_name = str(self.get_database_settings().get("NAME") or "").strip()
            return {
                "mode": "SQLite file",
                "engine": "SQLite",
                "archive_name": "db.sqlite3" if database_name.endswith("db.sqlite3") else "database/sqlite_backup.sqlite3",
                "available": bool(database_name),
                "help_text": "Full backup includes the SQLite database file together with project files and media.",
            }
        return {
            "mode": "Project files only",
            "engine": "Unknown",
            "archive_name": "",
            "available": False,
            "help_text": "The current database engine is not configured for automatic backup export yet.",
        }

    def get_pg_dump_version_text(self):
        if not self.is_postgres_database():
            return ""

        pg_dump_command = self.get_pg_dump_command()
        try:
            completed = subprocess.run(
                [pg_dump_command, "--version"],
                check=False,
                capture_output=True,
                text=True,
                timeout=20,
            )
        except (FileNotFoundError, subprocess.TimeoutExpired):
            return ""

        version_text = (completed.stdout or completed.stderr or "").strip()
        return version_text

    def run_database_backup_preflight(self):
        if self.is_postgres_database():
            pg_dump_command = self.get_pg_dump_command()
            version_text = self.get_pg_dump_version_text()
            with tempfile.TemporaryDirectory() as temp_dir:
                probe_output_path = Path(temp_dir) / "postgresql_preflight.sql"
                command = [
                    pg_dump_command,
                    "--schema-only",
                    "--file",
                    str(probe_output_path),
                    "--format=plain",
                    "--encoding=UTF8",
                    "--no-owner",
                    "--no-privileges",
                    "--dbname",
                    self.build_postgres_conninfo(),
                ]
                try:
                    completed = subprocess.run(
                        command,
                        check=False,
                        capture_output=True,
                        text=True,
                        timeout=120,
                    )
                except FileNotFoundError:
                    return {
                        "ok": False,
                        "message": "Database backup pre-check failed because pg_dump is not installed in this environment.",
                    }
                except subprocess.TimeoutExpired:
                    return {
                        "ok": False,
                        "message": "Database backup pre-check timed out while trying to reach PostgreSQL with pg_dump.",
                    }

            if completed.returncode != 0:
                stderr_text = (completed.stderr or completed.stdout or "").strip()
                version_suffix = f" Tool: {version_text}." if version_text else ""
                return {
                    "ok": False,
                    "message": f"Database backup pre-check failed.{version_suffix} {stderr_text or 'pg_dump returned a non-zero exit code.'}".strip(),
                }

            version_suffix = f" using {version_text}" if version_text else ""
            return {
                "ok": True,
                "message": f"Database backup pre-check passed{version_suffix}. The server can generate a PostgreSQL dump.",
            }

        if self.is_sqlite_database():
            return {
                "ok": True,
                "message": "This environment uses SQLite, so the database file can be copied directly into the backup ZIP.",
            }

        return {
            "ok": False,
            "message": "This environment is not configured with a supported automatic database backup mode.",
        }

    def build_database_backup_file(self, temp_dir):
        database_settings = self.get_database_settings()

        if self.is_postgres_database():
            pg_dump_command = self.get_pg_dump_command()
            dump_output_path = Path(temp_dir) / "postgresql_dump.sql"
            command = [
                pg_dump_command,
                "--file",
                str(dump_output_path),
                "--format=plain",
                "--encoding=UTF8",
                "--no-owner",
                "--no-privileges",
                "--dbname",
                self.build_postgres_conninfo(),
            ]
            try:
                completed = subprocess.run(
                    command,
                    check=False,
                    capture_output=True,
                    text=True,
                )
            except FileNotFoundError:
                return None, "PostgreSQL dump was skipped because pg_dump is not available in this environment."

            if completed.returncode != 0:
                stderr_text = (completed.stderr or completed.stdout or "").strip()
                return None, f"PostgreSQL dump was skipped. {stderr_text or 'pg_dump returned a non-zero exit code.'}"

            if not dump_output_path.exists():
                return None, "PostgreSQL dump was skipped because pg_dump did not create the dump file."

            return {
                "path": dump_output_path,
                "archive_name": "database/postgresql_dump.sql",
                "manifest_label": "PostgreSQL SQL dump",
            }, ""

        if self.is_sqlite_database():
            database_name = str(database_settings.get("NAME") or "").strip()
            if not database_name:
                return None, ""
            sqlite_path = Path(database_name)
            if not sqlite_path.exists():
                return None, f"SQLite database file not found: {sqlite_path}"
            try:
                sqlite_path.relative_to(settings.BASE_DIR)
                return None, ""
            except ValueError:
                return {
                    "path": sqlite_path,
                    "archive_name": "database/sqlite_backup.sqlite3",
                    "manifest_label": "SQLite database file",
                }, ""

        return None, ""

    def should_skip_dir(self, dir_name):
        excluded = set(getattr(settings, "HR_BACKUP_EXCLUDE_DIR_NAMES", set()))
        return dir_name in excluded

    def should_skip_file(self, file_path):
        excluded_suffixes = set(getattr(settings, "HR_BACKUP_EXCLUDE_FILE_SUFFIXES", set()))
        return file_path.suffix.lower() in excluded_suffixes

    def iter_backup_files(self, path):
        if path.is_file():
            if not self.should_skip_file(path):
                yield path
            return

        for child in path.iterdir():
            if child.is_dir():
                if self.should_skip_dir(child.name):
                    continue
                yield from self.iter_backup_files(child)
            elif child.is_file() and not self.should_skip_file(child):
                yield child

    def build_backup_filename(self, note=""):
        timestamp = timezone.localtime().strftime("%Y-%m-%d_%H-%M-%S")
        safe_note = self.sanitize_note(note)
        filename = f"nouraxis_backup_{timestamp}"
        if safe_note:
            filename = f"{filename}_{safe_note}"
        return f"{filename}.zip", safe_note

    def write_backup_zip(
        self,
        zip_handle,
        archive_label,
        safe_note,
        backup_root_label,
        database_backup_file=None,
        database_backup_warning="",
    ):
        include_paths = self.get_include_paths()
        if not include_paths:
            raise ValueError("No valid backup include paths were found in settings.py.")
        recent_update_items = self.get_backup_recent_update_items()
        next_focus_items = self.get_backup_next_focus_items()

        manifest_lines = [
            "NourAxis Backup Manifest",
            f"Created at: {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S %Z')}",
            f"Created by: {self.request.user.get_username()}",
            f"Backup file: {archive_label}",
            f"Backup root: {backup_root_label}",
            f"Note: {safe_note or '—'}",
            "",
            "Database backup:",
        ]
        if database_backup_file:
            manifest_lines.append(
                f"- Included: {database_backup_file['manifest_label']} -> {database_backup_file['archive_name']}"
            )
        else:
            manifest_lines.append("- Included: No database dump file was added to this ZIP.")
        if database_backup_warning:
            manifest_lines.append(f"- Warning: {database_backup_warning}")

        manifest_lines.extend([
            "",
            "Included paths:",
        ])
        for include_item in include_paths:
            manifest_lines.append(f"- {include_item['display_label']}")

        manifest_lines.extend([
            "",
            "Recent implementation snapshot:",
        ])
        for update_item in recent_update_items:
            manifest_lines.append(f"- {update_item['title']}: {update_item['summary']}")
            for path in update_item["paths"]:
                manifest_lines.append(f"  * {path}")

        manifest_lines.extend([
            "",
            "Recommended next focus:",
        ])
        for item in next_focus_items:
            manifest_lines.append(f"- {item}")

        for include_item in include_paths:
            include_path = include_item["path"]
            archive_root = include_item["archive_root"]

            if include_path.is_file():
                archive_name = self.build_archive_name(include_path, archive_root)
                zip_handle.write(include_path, arcname=archive_name)
                continue

            for child_file in self.iter_backup_files(include_path):
                archive_name = self.build_archive_name(include_path, archive_root, child_file)
                zip_handle.write(child_file, arcname=archive_name)

        if database_backup_file:
            zip_handle.write(
                database_backup_file["path"],
                arcname=database_backup_file["archive_name"],
            )

        zip_handle.writestr("backup_manifest.txt", "\n".join(manifest_lines))

    def create_backup(self, note=""):
        backup_root = self.get_backup_root()
        filename, safe_note = self.build_backup_filename(note=note)
        backup_file = backup_root / filename

        with tempfile.TemporaryDirectory() as temp_dir:
            database_backup_file, database_backup_warning = self.build_database_backup_file(temp_dir)
            with zipfile.ZipFile(backup_file, "w", compression=zipfile.ZIP_DEFLATED) as zip_handle:
                self.write_backup_zip(
                    zip_handle=zip_handle,
                    archive_label=backup_file.name,
                    safe_note=safe_note,
                    backup_root_label=str(backup_root),
                    database_backup_file=database_backup_file,
                    database_backup_warning=database_backup_warning,
                )

        return backup_file, database_backup_warning

    def download_backup_response(self, note=""):
        filename, safe_note = self.build_backup_filename(note=note)
        buffer = io.BytesIO()

        with tempfile.TemporaryDirectory() as temp_dir:
            database_backup_file, database_backup_warning = self.build_database_backup_file(temp_dir)
            with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as zip_handle:
                self.write_backup_zip(
                    zip_handle=zip_handle,
                    archive_label=filename,
                    safe_note=safe_note,
                    backup_root_label="Downloaded to browser",
                    database_backup_file=database_backup_file,
                    database_backup_warning=database_backup_warning,
                )

        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        if database_backup_warning:
            messages.warning(
                self.request,
                f"Backup downloaded, but the database dump was not included. {database_backup_warning}",
            )
        return response

    def format_size_label(self, size_bytes):
        if size_bytes in {None, ""}:
            return "—"
        size_value = float(size_bytes)
        for unit in ["B", "KB", "MB", "GB", "TB"]:
            if size_value < 1024 or unit == "TB":
                if unit == "B":
                    return f"{int(size_value)} {unit}"
                return f"{size_value:.2f} {unit}"
            size_value /= 1024
        return f"{int(size_bytes)} B"

    def get_database_archive_candidates(self):
        database_backup_status = self.get_database_backup_status()
        archive_name = str(database_backup_status.get("archive_name") or "").strip()
        candidates = []
        if archive_name:
            candidates.append(archive_name)
        if self.is_sqlite_database():
            candidates.extend(["db.sqlite3", "database/sqlite_backup.sqlite3"])
        return list(dict.fromkeys(candidates))

    def inspect_backup_database_payload(self, backup_path):
        candidates = set(self.get_database_archive_candidates())
        if not candidates:
            return {
                "database_entry_name": "",
                "database_included": False,
                "database_status_label": "Not configured",
                "database_status_class": "backup-status-muted",
                "database_size_bytes": None,
                "database_size_label": "—",
                "database_inspection_error": "",
            }

        try:
            with zipfile.ZipFile(backup_path, "r") as archive:
                for zip_info in archive.infolist():
                    normalized_name = zip_info.filename.rstrip("/")
                    if normalized_name in candidates:
                        return {
                            "database_entry_name": normalized_name,
                            "database_included": True,
                            "database_status_label": "Included",
                            "database_status_class": "backup-status-success",
                            "database_size_bytes": zip_info.file_size,
                            "database_size_label": self.format_size_label(zip_info.file_size),
                            "database_inspection_error": "",
                        }
        except (FileNotFoundError, OSError, zipfile.BadZipFile) as exc:
            return {
                "database_entry_name": "",
                "database_included": False,
                "database_status_label": "Unreadable ZIP",
                "database_status_class": "backup-status-danger",
                "database_size_bytes": None,
                "database_size_label": "—",
                "database_inspection_error": str(exc),
            }

        return {
            "database_entry_name": "",
            "database_included": False,
            "database_status_label": "Missing from ZIP",
            "database_status_class": "backup-status-warning",
            "database_size_bytes": None,
            "database_size_label": "—",
            "database_inspection_error": "",
        }

    def get_latest_backups(self):
        backup_root = self.get_backup_root()
        backups = []
        for path in backup_root.glob("*.zip"):
            stat = path.stat()
            backup_item = {
                "name": path.name,
                "path": str(path),
                "modified_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.get_current_timezone()),
                "size_bytes": stat.st_size,
                "size_mb": f"{stat.st_size / (1024 * 1024):.2f}",
            }
            backup_item.update(self.inspect_backup_database_payload(path))
            backups.append(backup_item)
        backups.sort(key=lambda item: item["modified_at"], reverse=True)
        return backups

    def get_employee_export_queryset(self):
        return Employee.objects.select_related(
            "user",
            "company",
            "department",
            "branch",
            "section",
            "job_title",
        ).order_by("employee_id", "full_name")

    def get_attendance_export_queryset(self, start_date=None, end_date=None):
        queryset = EmployeeAttendanceLedger.objects.select_related(
            "employee",
            "employee__company",
            "employee__department",
            "employee__branch",
            "employee__section",
            "employee__job_title",
            "linked_leave",
            "linked_action_record",
        ).order_by("-attendance_date", "employee__employee_id", "-id")
        if start_date:
            queryset = queryset.filter(attendance_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(attendance_date__lte=end_date)
        return queryset

    def get_leave_export_queryset(self, start_date=None, end_date=None):
        queryset = EmployeeLeave.objects.select_related(
            "employee",
            "employee__company",
            "employee__department",
            "employee__branch",
            "employee__section",
            "employee__job_title",
            "requested_by",
            "reviewed_by",
            "approved_by",
            "rejected_by",
            "cancelled_by",
            "supervisor_reviewed_by",
            "operations_reviewed_by",
            "hr_reviewed_by",
        ).order_by("-start_date", "employee__employee_id", "-id")
        if start_date:
            queryset = queryset.filter(start_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(end_date__lte=end_date)
        return queryset

    def style_sheet(self, worksheet, headers):
        header_font = Font(bold=True)
        for index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=index)
            cell.font = header_font
            column_letter = worksheet.cell(row=1, column=index).column_letter
            worksheet.column_dimensions[column_letter].width = max(16, min(len(str(header)) + 4, 30))
        worksheet.freeze_panes = "A2"

    def build_workbook_response(self, workbook, filename):
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response

    def export_employee_master_data(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Employees"

        headers = [
            "Employee ID",
            "Full Name",
            "Login Email",
            "Contact Email",
            "Phone",
            "Company",
            "Department",
            "Branch",
            "Section",
            "Job Title",
            "Hire Date",
            "Employment Status",
            "Operationally Active",
            "Passport Issue Date",
            "Passport Expiry Date",
            "Civil ID Issue Date",
            "Civil ID Expiry Date",
            "Salary",
            "Notes",
            "Created At",
            "Updated At",
        ]
        worksheet.append(headers)

        for employee in self.get_employee_export_queryset():
            worksheet.append([
                employee.employee_id,
                employee.full_name,
                getattr(employee.user, "email", "") or "",
                employee.email or "",
                employee.phone or "",
                employee.company.name if employee.company_id else "",
                employee.department.name if employee.department_id else "",
                employee.branch.name if employee.branch_id else "",
                employee.section.name if employee.section_id else "",
                employee.job_title.name if employee.job_title_id else "",
                employee.hire_date.isoformat() if employee.hire_date else "",
                employee.get_employment_status_display(),
                "Yes" if employee.is_active else "No",
                employee.passport_issue_date.isoformat() if employee.passport_issue_date else "",
                employee.passport_expiry_date.isoformat() if employee.passport_expiry_date else "",
                employee.civil_id_issue_date.isoformat() if employee.civil_id_issue_date else "",
                employee.civil_id_expiry_date.isoformat() if employee.civil_id_expiry_date else "",
                str(employee.salary) if employee.salary is not None else "",
                employee.notes or "",
                timezone.localtime(employee.created_at).strftime("%Y-%m-%d %H:%M:%S") if employee.created_at else "",
                timezone.localtime(employee.updated_at).strftime("%Y-%m-%d %H:%M:%S") if employee.updated_at else "",
            ])

        self.style_sheet(worksheet, headers)
        filename = f"employee_master_data_{timezone.localtime().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        return self.build_workbook_response(workbook, filename)

    def export_attendance_data(self, start_date=None, end_date=None):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Attendance"

        headers = [
            "Attendance Date",
            "Employee ID",
            "Employee Name",
            "Company",
            "Department",
            "Branch",
            "Section",
            "Job Title",
            "Day Status",
            "Clock In",
            "Clock Out",
            "Scheduled Hours",
            "Worked Hours",
            "Late Minutes",
            "Early Departure Minutes",
            "Overtime Minutes",
            "Paid Day",
            "Source",
            "Linked Leave",
            "Linked Action Record",
            "Notes",
            "Created By",
            "Updated By",
            "Created At",
            "Updated At",
        ]
        worksheet.append(headers)

        for entry in self.get_attendance_export_queryset(start_date=start_date, end_date=end_date):
            worksheet.append([
                entry.attendance_date.isoformat() if entry.attendance_date else "",
                entry.employee.employee_id if entry.employee_id else "",
                entry.employee.full_name if entry.employee_id else "",
                entry.employee.company.name if entry.employee and entry.employee.company_id else "",
                entry.employee.department.name if entry.employee and entry.employee.department_id else "",
                entry.employee.branch.name if entry.employee and entry.employee.branch_id else "",
                entry.employee.section.name if entry.employee and entry.employee.section_id else "",
                entry.employee.job_title.name if entry.employee and entry.employee.job_title_id else "",
                entry.get_day_status_display(),
                entry.clock_in_time.strftime("%H:%M") if entry.clock_in_time else "",
                entry.clock_out_time.strftime("%H:%M") if entry.clock_out_time else "",
                str(entry.scheduled_hours),
                str(entry.worked_hours),
                entry.late_minutes,
                entry.early_departure_minutes,
                entry.overtime_minutes,
                "Yes" if entry.is_paid_day else "No",
                entry.get_source_display(),
                entry.linked_leave.get_leave_type_display() if entry.linked_leave_id else "",
                getattr(entry.linked_action_record, "title", "") if entry.linked_action_record_id else "",
                entry.notes or "",
                entry.created_by or "",
                entry.updated_by or "",
                timezone.localtime(entry.created_at).strftime("%Y-%m-%d %H:%M:%S") if entry.created_at else "",
                timezone.localtime(entry.updated_at).strftime("%Y-%m-%d %H:%M:%S") if entry.updated_at else "",
            ])

        self.style_sheet(worksheet, headers)
        suffix_parts = []
        if start_date:
            suffix_parts.append(f"from_{start_date.isoformat()}")
        if end_date:
            suffix_parts.append(f"to_{end_date.isoformat()}")
        suffix = "_" + "_".join(suffix_parts) if suffix_parts else ""
        filename = f"attendance_export{suffix}_{timezone.localtime().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        return self.build_workbook_response(workbook, filename)

    def export_leave_data(self, start_date=None, end_date=None):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Leave Records"

        headers = [
            "Employee ID",
            "Employee Name",
            "Company",
            "Department",
            "Branch",
            "Section",
            "Job Title",
            "Leave Type",
            "Start Date",
            "End Date",
            "Total Days",
            "Status",
            "Current Stage",
            "Reason",
            "Approval Note",
            "Requested By",
            "Reviewed By",
            "Approved By",
            "Rejected By",
            "Cancelled By",
            "Supervisor Reviewed By",
            "Operations Reviewed By",
            "HR Reviewed By",
            "Supervisor Review Note",
            "Operations Review Note",
            "HR Review Note",
            "Finalized At",
            "Created At",
            "Updated At",
        ]
        worksheet.append(headers)

        for leave_record in self.get_leave_export_queryset(start_date=start_date, end_date=end_date):
            worksheet.append([
                leave_record.employee.employee_id if leave_record.employee_id else "",
                leave_record.employee.full_name if leave_record.employee_id else "",
                leave_record.employee.company.name if leave_record.employee and leave_record.employee.company_id else "",
                leave_record.employee.department.name if leave_record.employee and leave_record.employee.department_id else "",
                leave_record.employee.branch.name if leave_record.employee and leave_record.employee.branch_id else "",
                leave_record.employee.section.name if leave_record.employee and leave_record.employee.section_id else "",
                leave_record.employee.job_title.name if leave_record.employee and leave_record.employee.job_title_id else "",
                leave_record.get_leave_type_display(),
                leave_record.start_date.isoformat() if leave_record.start_date else "",
                leave_record.end_date.isoformat() if leave_record.end_date else "",
                leave_record.total_days,
                leave_record.get_status_display(),
                leave_record.get_current_stage_display(),
                leave_record.reason or "",
                leave_record.approval_note or "",
                leave_record.requested_by.get_username() if leave_record.requested_by_id else "",
                leave_record.reviewed_by.get_username() if leave_record.reviewed_by_id else "",
                leave_record.approved_by.get_username() if leave_record.approved_by_id else "",
                leave_record.rejected_by.get_username() if leave_record.rejected_by_id else "",
                leave_record.cancelled_by.get_username() if leave_record.cancelled_by_id else "",
                leave_record.supervisor_reviewed_by.get_username() if leave_record.supervisor_reviewed_by_id else "",
                leave_record.operations_reviewed_by.get_username() if leave_record.operations_reviewed_by_id else "",
                leave_record.hr_reviewed_by.get_username() if leave_record.hr_reviewed_by_id else "",
                leave_record.supervisor_review_note or "",
                leave_record.operations_review_note or "",
                leave_record.hr_review_note or "",
                timezone.localtime(leave_record.finalized_at).strftime("%Y-%m-%d %H:%M:%S") if leave_record.finalized_at else "",
                timezone.localtime(leave_record.created_at).strftime("%Y-%m-%d %H:%M:%S") if leave_record.created_at else "",
                timezone.localtime(leave_record.updated_at).strftime("%Y-%m-%d %H:%M:%S") if leave_record.updated_at else "",
            ])

        self.style_sheet(worksheet, headers)
        suffix_parts = []
        if start_date:
            suffix_parts.append(f"from_{start_date.isoformat()}")
        if end_date:
            suffix_parts.append(f"to_{end_date.isoformat()}")
        suffix = "_" + "_".join(suffix_parts) if suffix_parts else ""
        filename = f"leave_records_export{suffix}_{timezone.localtime().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        return self.build_workbook_response(workbook, filename)

    def export_backup_audit_data(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Backup Audit"
        recent_updates_label = " | ".join(
            f"{item['title']}: {item['summary']}" for item in self.get_backup_recent_update_items()
        )
        next_focus_label = " | ".join(self.get_backup_next_focus_items())

        headers = [
            "Backup File Name",
            "Created At",
            "Size (Bytes)",
            "Size (MB)",
            "Database Included",
            "Database Entry",
            "Database Size (Bytes)",
            "Database Size",
            "ZIP Inspection",
            "Recent Updates Snapshot",
            "Recommended Next Focus",
            "Save Path",
        ]
        worksheet.append(headers)

        for backup in self.get_latest_backups():
            worksheet.append([
                backup["name"],
                backup["modified_at"].strftime("%Y-%m-%d %H:%M:%S"),
                backup["size_bytes"],
                backup["size_mb"],
                "Yes" if backup["database_included"] else "No",
                backup["database_entry_name"],
                backup["database_size_bytes"] if backup["database_size_bytes"] is not None else "",
                backup["database_size_label"],
                backup["database_status_label"],
                recent_updates_label,
                next_focus_label,
                backup["path"],
            ])

        self.style_sheet(worksheet, headers)
        filename = f"backup_audit_{timezone.localtime().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        return self.build_workbook_response(workbook, filename)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        include_paths = self.get_include_paths()
        latest_backups = self.get_latest_backups()
        database_backup_status = self.get_database_backup_status()
        context.update(
            {
                "backup_root": self.get_backup_root(),
                "include_paths": [item["display_label"] for item in include_paths],
                "database_backup_status": database_backup_status,
                "backup_recent_update_items": self.get_backup_recent_update_items(),
                "backup_next_focus_items": self.get_backup_next_focus_items(),
                "latest_backups": latest_backups[: self.backup_table_limit],
                "backup_count": len(latest_backups),
                "has_backups": bool(latest_backups),
                "employee_export_count": self.get_employee_export_queryset().count(),
                "attendance_export_count": EmployeeAttendanceLedger.objects.count(),
                "leave_export_count": EmployeeLeave.objects.count(),
                "today": timezone.localdate(),
            }
        )
        return context
