from datetime import date, datetime, timedelta
import csv
from decimal import Decimal
import io
from math import asin, cos, radians, sin, sqrt
import mimetypes
import re
from pathlib import Path

from django import forms
from django.apps import apps
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.db.models import Case, IntegerField, Prefetch, Q, When
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.generic import CreateView, DeleteView, DetailView, ListView, UpdateView

from config.mixins import ProtectedDeleteMixin
from organization.models import Branch, Company, Department, JobTitle, Section
from payroll.forms import PayrollProfileForm
from operations.forms import BranchPostForm
from operations.services import build_branch_workspace_context, build_employee_schedule_snapshot
from openpyxl import Workbook, load_workbook

from .access import (
    get_user_scope_branch as get_user_scope_branch_for_role,
    get_workspace_home_label,
    get_workspace_profile_url,
    is_admin_compatible as is_admin_compatible_role,
    is_branch_scoped_supervisor as is_branch_scoped_supervisor_for_role,
    is_employee_role_user as is_employee_role_user_role,
    is_hr_user as is_hr_user_role,
    is_operations_manager_user as is_operations_manager_user_role,
    is_supervisor_user as is_supervisor_user_role,
    should_use_management_own_profile,
)
from .forms import (
    AttendanceFilterForm,
    AttendanceManagementFilterForm,
    BranchWeeklyScheduleThemeForm,
    BranchWeeklyDutyOptionStyleForm,
    BranchWeeklyDutyOptionTimingForm,
    BranchWeeklyDutyOptionForm,
    BranchWeeklyScheduleImportForm,
    BranchWeeklyPendingOffForm,
    BranchWeeklyScheduleEntryForm,
    EmployeeActionRecordForm,
    EmployeeAttendanceCorrectionForm,
    EmployeeSelfServiceAttendanceForm,
    EmployeeAttendanceLedgerForm,
    EmployeeDocumentForm,
    EmployeeForm,
    EmployeeHistoryForm,
    EmployeeLeaveForm,
    EmployeeRequiredSubmissionCreateForm,
    EmployeeRequiredSubmissionResponseForm,
    EmployeeRequiredSubmissionReviewForm,
    EmployeeDocumentRequestCreateForm,
    EmployeeDocumentRequestReviewForm,
    EmployeeSelfServiceLeaveRequestForm,
    EmployeeTransferForm,
)
from .models import (
    BranchScheduleGridCell,
    BranchScheduleGridHeader,
    BranchScheduleGridRow,
    BranchWeeklyScheduleTheme,
    BranchWeeklyDutyOption,
    BranchWeeklyPendingOff,
    BranchWeeklyScheduleEntry,
    Employee,
    EmployeeActionRecord,
    EmployeeAttendanceCorrection,
    EmployeeAttendanceEvent,
    EmployeeAttendanceLedger,
    EmployeeDocument,
    EmployeeHistory,
    EmployeeLeave,
    EmployeeRequiredSubmission,
    EmployeeDocumentRequest,
    WORKING_HOURS_PER_DAY,
    build_employee_working_time_summary,
    count_policy_working_days,
    get_schedule_week_start,
    is_policy_holiday,
    is_policy_weekly_off_day,
)

FREE_SCHEDULE_GRID_DEFAULT_HEADERS = {
    0: "Employee",
    1: "Job Title",
    2: "Sunday",
    3: "Monday",
    4: "Tuesday",
    5: "Wednesday",
    6: "Thursday",
    7: "Friday",
    8: "Saturday",
    9: "Notes",
    10: "Orders",
    11: "Follow Up",
}


def create_employee_history(
    employee,
    title,
    description="",
    event_type=EmployeeHistory.EVENT_NOTE,
    created_by="",
    is_system_generated=False,
    event_date=None,
):
    EmployeeHistory.objects.create(
        employee=employee,
        title=title,
        description=description,
        event_type=event_type,
        created_by=created_by,
        is_system_generated=is_system_generated,
        event_date=event_date or timezone.localdate(),
    )


def get_actor_label(user):
    if not user or not user.is_authenticated:
        return ""

    full_name = ""
    if hasattr(user, "get_full_name"):
        full_name = user.get_full_name().strip()

    return full_name or getattr(user, "username", "") or str(user)


def get_user_employee_profile(user):
    if not user or not user.is_authenticated:
        return None

    try:
        return user.employee_profile
    except Employee.DoesNotExist:
        return None
    except AttributeError:
        return None


def get_safe_next_url(request, fallback_url):
    next_url = (request.POST.get("next") or request.GET.get("next") or "").strip()
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return next_url
    return fallback_url


def get_self_service_home_label(employee, user):
    return get_workspace_home_label(user, employee)


def build_self_service_page_context(request, employee, *, current_section):
    detail_view = EmployeeDetailView()
    detail_view.request = request
    detail_view.object = employee
    context = detail_view.get_context_data(object=employee)
    context["self_service_home_label"] = get_self_service_home_label(employee, request.user)
    context["self_service_current_section"] = current_section
    context["self_service_profile_url"] = get_workspace_profile_url(request.user, employee)
    context["self_service_leave_url"] = reverse("employees:self_service_leave")
    context["self_service_documents_url"] = reverse("employees:self_service_documents")
    context["self_service_attendance_url"] = reverse("employees:self_service_attendance")
    context["self_service_working_time_url"] = reverse("employees:self_service_working_time")
    context["self_service_branch_url"] = reverse("employees:self_service_branch")
    context["self_service_weekly_schedule_url"] = reverse("employees:self_service_weekly_schedule")
    context["self_service_my_schedule_url"] = reverse("employees:self_service_my_schedule")
    return context


def get_shift_time_defaults(shift_value):
    shift_map = EmployeeAttendanceLedger.get_shift_time_map()
    shift_config = shift_map.get(shift_value or EmployeeAttendanceLedger.SHIFT_MORNING)
    return shift_config or shift_map[EmployeeAttendanceLedger.SHIFT_MORNING]


def normalize_attendance_shift_label(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def resolve_attendance_shift_value(*, label="", start_time=None, end_time=None):
    shift_map = EmployeeAttendanceLedger.get_shift_time_map()
    normalized_label = normalize_attendance_shift_label(label)
    start_value = start_time.strftime("%H:%M") if start_time else ""
    end_value = end_time.strftime("%H:%M") if end_time else ""

    if normalized_label and start_value and end_value:
        for shift_value, shift_config in shift_map.items():
            if (
                normalize_attendance_shift_label(shift_config.get("label")) == normalized_label
                and shift_config.get("start") == start_value
                and shift_config.get("end") == end_value
            ):
                return shift_value

    if start_value and end_value:
        for shift_value, shift_config in shift_map.items():
            if shift_config.get("start") == start_value and shift_config.get("end") == end_value:
                return shift_value

    if normalized_label:
        for shift_value, shift_config in shift_map.items():
            if normalize_attendance_shift_label(shift_config.get("label")) == normalized_label:
                return shift_value

    return ""


def build_self_service_shift_choices(branch):
    if not branch:
        return list(EmployeeAttendanceLedger.SHIFT_CHOICES)

    duty_options = list(
        BranchWeeklyDutyOption.objects.filter(
            branch=branch,
            is_active=True,
            duty_type=BranchWeeklyScheduleEntry.DUTY_TYPE_SHIFT,
        ).order_by("display_order", "label", "id")
    )
    if not duty_options:
        return list(EmployeeAttendanceLedger.SHIFT_CHOICES)

    choices = []
    seen_pairs = set()
    fallback_seen_values = set()
    fallback_choices = []

    for duty_option in duty_options:
        shift_value = resolve_attendance_shift_value(
            label=duty_option.label,
            start_time=duty_option.default_start_time,
            end_time=duty_option.default_end_time,
        )
        if not shift_value:
            continue

        pair = (shift_value, duty_option.label)
        if pair not in seen_pairs:
            choices.append(pair)
            seen_pairs.add(pair)

        if shift_value not in fallback_seen_values:
            fallback_choices.append((shift_value, dict(EmployeeAttendanceLedger.SHIFT_CHOICES).get(shift_value, duty_option.label)))
            fallback_seen_values.add(shift_value)

    return choices or fallback_choices or list(EmployeeAttendanceLedger.SHIFT_CHOICES)


def build_attendance_location_label(cleaned_data):
    if not cleaned_data:
        return ""

    primary_label = (cleaned_data.get("location_label") or "").strip()
    location_address = (cleaned_data.get("location_address") or "").strip()

    if primary_label and location_address:
        return f"{primary_label} - {location_address}"
    if location_address:
        return location_address
    return primary_label


def calculate_haversine_distance_meters(latitude_a, longitude_a, latitude_b, longitude_b):
    earth_radius_meters = 6371000

    latitude_a = radians(float(latitude_a))
    longitude_a = radians(float(longitude_a))
    latitude_b = radians(float(latitude_b))
    longitude_b = radians(float(longitude_b))

    delta_latitude = latitude_b - latitude_a
    delta_longitude = longitude_b - longitude_a

    haversine_value = (
        sin(delta_latitude / 2) ** 2
        + cos(latitude_a) * cos(latitude_b) * sin(delta_longitude / 2) ** 2
    )
    central_angle = 2 * asin(sqrt(haversine_value))
    return int(round(earth_radius_meters * central_angle))


def get_branch_attendance_validation_result(employee, live_latitude, live_longitude):
    branch = getattr(employee, "branch", None) if employee else None
    if branch is None:
        return {
            "is_configured": False,
            "error_message": "This employee is not assigned to a branch attendance location.",
        }

    if not getattr(branch, "has_attendance_location_config", False):
        return {
            "is_configured": False,
            "error_message": (
                f"{branch.name} does not have a fixed attendance point configured yet. "
                "Please contact HR or Operations."
            ),
        }

    distance_meters = calculate_haversine_distance_meters(
        live_latitude,
        live_longitude,
        branch.attendance_latitude,
        branch.attendance_longitude,
    )
    allowed_radius_meters = int(branch.attendance_radius_meters or 0)
    is_inside_radius = distance_meters <= allowed_radius_meters

    return {
        "is_configured": True,
        "branch": branch,
        "branch_latitude": branch.attendance_latitude,
        "branch_longitude": branch.attendance_longitude,
        "allowed_radius_meters": allowed_radius_meters,
        "distance_meters": distance_meters,
        "is_inside_radius": is_inside_radius,
        "validation_status": (
            EmployeeAttendanceEvent.LOCATION_STATUS_INSIDE
            if is_inside_radius
            else EmployeeAttendanceEvent.LOCATION_STATUS_OUTSIDE
        ),
        "branch_location_label": f"{branch.name} attendance point",
        "validation_summary": (
            f"Validated against {branch.name} fixed attendance point. "
            f"Distance {distance_meters} m of allowed {allowed_radius_meters} m."
        ),
    }


def sync_attendance_event_to_ledger(event, actor_label="System"):
    if not event or not event.employee_id or not event.check_in_at or not event.check_out_at:
        return None

    clock_in_time = timezone.localtime(event.check_in_at).time().replace(microsecond=0)
    clock_out_time = timezone.localtime(event.check_out_at).time().replace(microsecond=0)

    ledger, _created = EmployeeAttendanceLedger.objects.get_or_create(
        employee=event.employee,
        attendance_date=event.attendance_date,
        defaults={
            "day_status": EmployeeAttendanceLedger.DAY_STATUS_PRESENT,
            "shift": event.shift or EmployeeAttendanceLedger.SHIFT_MORNING,
            "clock_in_time": clock_in_time,
            "clock_out_time": clock_out_time,
            "scheduled_hours": WORKING_HOURS_PER_DAY,
            "source": EmployeeAttendanceLedger.SOURCE_SYSTEM,
            "notes": "",
            "created_by": actor_label,
            "updated_by": actor_label,
        },
    )

    ledger.day_status = EmployeeAttendanceLedger.DAY_STATUS_PRESENT
    ledger.shift = event.shift or EmployeeAttendanceLedger.SHIFT_MORNING
    ledger.clock_in_time = clock_in_time
    ledger.clock_out_time = clock_out_time
    ledger.scheduled_hours = WORKING_HOURS_PER_DAY
    ledger.source = EmployeeAttendanceLedger.SOURCE_SYSTEM
    ledger.check_in_latitude = event.check_in_latitude
    ledger.check_in_longitude = event.check_in_longitude
    ledger.check_out_latitude = event.check_out_latitude
    ledger.check_out_longitude = event.check_out_longitude
    ledger.check_in_location_label = event.check_in_location_label or ""
    ledger.check_out_location_label = event.check_out_location_label or ""
    ledger.check_in_address = event.check_in_address or ""
    ledger.check_out_address = event.check_out_address or ""
    location_bits = []
    if event.check_in_location_label:
        location_bits.append(f"Check-in: {event.check_in_location_label}")
    if event.check_out_location_label:
        location_bits.append(f"Check-out: {event.check_out_location_label}")
    if event.check_in_address and event.check_in_address != event.check_in_location_label:
        location_bits.append(f"Check-in address: {event.check_in_address}")
    if event.check_out_address and event.check_out_address != event.check_out_location_label:
        location_bits.append(f"Check-out address: {event.check_out_address}")
    if event.check_in_latitude is not None and event.check_in_longitude is not None:
        location_bits.append(
            f"Check-in coordinates: {event.check_in_latitude}, {event.check_in_longitude}"
        )
    if event.check_in_distance_meters is not None:
        location_bits.append(
            f"Check-in validation: {event.get_check_in_location_validation_status_display() or event.check_in_location_validation_status} at {event.check_in_distance_meters} m"
        )
    if event.check_out_latitude is not None and event.check_out_longitude is not None:
        location_bits.append(
            f"Check-out coordinates: {event.check_out_latitude}, {event.check_out_longitude}"
        )
    if event.check_out_distance_meters is not None:
        location_bits.append(
            f"Check-out validation: {event.get_check_out_location_validation_status_display() or event.check_out_location_validation_status} at {event.check_out_distance_meters} m"
        )
    if event.branch_latitude_used is not None and event.branch_longitude_used is not None:
        location_bits.append(
            f"Branch point used: {event.branch_latitude_used}, {event.branch_longitude_used}"
        )
    if event.attendance_radius_meters_used is not None:
        location_bits.append(f"Allowed radius: {event.attendance_radius_meters_used} m")
    event_note = "Self-service attendance"
    if location_bits:
        event_note = f"{event_note}. {' | '.join(location_bits)}"
    ledger.notes = event_note
    ledger.updated_by = actor_label
    if not ledger.created_by:
        ledger.created_by = actor_label
    ledger.save()

    event.synced_ledger = ledger
    event.status = EmployeeAttendanceEvent.STATUS_COMPLETED
    event.save(update_fields=["synced_ledger", "status", "updated_at"])
    return ledger


def is_admin_compatible(user):
    return is_admin_compatible_role(user)


def is_hr_user(user):
    return is_hr_user_role(user)


def is_supervisor_user(user):
    return is_supervisor_user_role(user)


def is_operations_manager_user(user):
    return is_operations_manager_user_role(user)


def is_employee_role_user(user):
    return is_employee_role_user_role(user)


def is_management_user(user):
    return bool(
        user
        and user.is_authenticated
        and (
            is_admin_compatible(user)
            or is_hr_user(user)
            or is_supervisor_user(user)
            or is_operations_manager_user(user)
        )
    )


def is_self_employee(user, employee):
    linked_employee = get_user_employee_profile(user)
    return bool(linked_employee and employee and linked_employee.pk == employee.pk)


def get_user_scope_branch(user):
    linked_employee = get_user_employee_profile(user)
    return get_user_scope_branch_for_role(user, linked_employee)


def is_branch_scoped_supervisor(user):
    linked_employee = get_user_employee_profile(user)
    return is_branch_scoped_supervisor_for_role(user, linked_employee)


def can_supervisor_view_employee(user, employee):
    scoped_branch = get_user_scope_branch(user)
    if not scoped_branch or not employee:
        return False
    return bool(employee.branch_id and employee.branch_id == scoped_branch.id)


def get_employee_directory_queryset_for_user(user, queryset=None):
    base_queryset = queryset or Employee.objects.all()

    if is_branch_scoped_supervisor(user):
        scoped_branch = get_user_scope_branch(user)
        return base_queryset.filter(branch_id=scoped_branch.id)

    if is_admin_compatible(user) or is_hr_user(user) or is_operations_manager_user(user):
        return base_queryset

    linked_employee = get_user_employee_profile(user)
    if linked_employee:
        return base_queryset.filter(pk=linked_employee.pk)

    return base_queryset.none()


def get_leave_queryset_for_user(user, queryset=None):
    base_queryset = queryset or EmployeeLeave.objects.all()

    if is_branch_scoped_supervisor(user):
        scoped_branch = get_user_scope_branch(user)
        return base_queryset.filter(employee__branch_id=scoped_branch.id)

    if is_admin_compatible(user) or is_hr_user(user) or is_operations_manager_user(user):
        return base_queryset

    linked_employee = get_user_employee_profile(user)
    if linked_employee:
        return base_queryset.filter(employee=linked_employee)

    return base_queryset.none()


def can_view_employee_requests_overview(user):
    return bool(
        user
        and user.is_authenticated
        and (
            is_admin_compatible(user)
            or is_hr_user(user)
            or is_operations_manager_user(user)
            or is_branch_scoped_supervisor(user)
        )
    )


def can_view_management_employee_sections(user, employee=None):
    if not user or not user.is_authenticated:
        return False

    if is_admin_compatible(user) or is_hr_user(user) or is_operations_manager_user(user):
        return True

    return can_supervisor_view_employee(user, employee)


def can_view_employee_directory(user):
    return bool(
        user
        and user.is_authenticated
        and (
            is_admin_compatible(user)
            or is_hr_user(user)
            or is_operations_manager_user(user)
            or is_branch_scoped_supervisor(user)
        )
    )


def can_view_employee_profile(user, employee):
    return bool(
        is_self_employee(user, employee)
        or is_admin_compatible(user)
        or is_hr_user(user)
        or is_operations_manager_user(user)
        or can_supervisor_view_employee(user, employee)
    )


def can_create_or_edit_employees(user):
    return is_admin_compatible(user) or is_hr_user(user) or is_operations_manager_user(user)


def can_delete_employee(user):
    return is_admin_compatible(user) or is_hr_user(user) or is_operations_manager_user(user)


def can_transfer_employee(user):
    return is_admin_compatible(user) or is_hr_user(user) or is_operations_manager_user(user)


def can_change_employee_status(user):
    return is_admin_compatible(user) or is_hr_user(user) or is_operations_manager_user(user)


def can_manage_employee_documents(user, employee=None):
    return bool(
        is_admin_compatible(user)
        or is_hr_user(user)
        or is_operations_manager_user(user)
        or can_supervisor_view_employee(user, employee)
    )


def can_manage_employee_required_submissions(user, employee=None):
    return bool(
        is_admin_compatible(user)
        or is_hr_user(user)
        or is_operations_manager_user(user)
        or can_supervisor_view_employee(user, employee)
    )


def can_submit_employee_required_submission(user, submission_request):
    return bool(
        submission_request
        and is_self_employee(user, submission_request.employee)
        and submission_request.status in {
            EmployeeRequiredSubmission.STATUS_REQUESTED,
            EmployeeRequiredSubmission.STATUS_NEEDS_CORRECTION,
        }
    )


def can_review_employee_required_submission(user, submission_request):
    return bool(
        submission_request
        and can_manage_employee_required_submissions(user, submission_request.employee)
    )


def can_create_employee_document_request(user, employee):
    return bool(employee and is_self_employee(user, employee))


def can_review_employee_document_request(user, document_request):
    return bool(
        document_request
        and (
            is_admin_compatible(user)
            or is_hr_user(user)
            or is_operations_manager_user(user)
            or can_supervisor_view_employee(user, document_request.employee)
        )
    )


def can_cancel_employee_document_request(user, document_request):
    return bool(
        document_request
        and is_self_employee(user, document_request.employee)
        and document_request.can_employee_cancel
    )


def build_employee_document_request_summary(document_request):
    parts = [
        f"Document request created as {document_request.get_request_type_display()}.",
        f"Status: {document_request.get_status_display()}.",
    ]

    if document_request.needed_by_date:
        parts.append(f"Needed by: {document_request.needed_by_date.strftime('%B %d, %Y')}.")
    if document_request.request_note:
        parts.append(f"Employee note: {document_request.request_note}")

    return " ".join(parts)


def build_employee_document_request_review_summary(document_request, previous_status, new_status, management_note=""):
    parts = [
        f"Employee document request status changed from {previous_status} to {new_status}.",
        f"Request type: {document_request.get_request_type_display()}.",
    ]

    if management_note:
        parts.append(f"Management note: {management_note}")

    return " ".join(parts)


def can_request_leave(user, employee):
    return bool(
        is_self_employee(user, employee)
        or is_admin_compatible(user)
        or is_hr_user(user)
        or is_operations_manager_user(user)
        or can_supervisor_view_employee(user, employee)
    )


def can_review_leave(user):
    return (
        is_admin_compatible(user)
        or is_hr_user(user)
        or is_operations_manager_user(user)
        or is_branch_scoped_supervisor(user)
    )


def get_leave_current_stage_owner_label(leave_record):
    if not leave_record:
        return ""

    if leave_record.status == EmployeeLeave.STATUS_PENDING:
        if leave_record.current_stage == EmployeeLeave.STAGE_SUPERVISOR_REVIEW:
            return "Supervisor review"
        if leave_record.current_stage == EmployeeLeave.STAGE_OPERATIONS_REVIEW:
            return "Operations review"
        if leave_record.current_stage == EmployeeLeave.STAGE_HR_REVIEW:
            return "HR final review"
        return "Pending workflow review"

    if leave_record.status == EmployeeLeave.STATUS_APPROVED:
        return "Workflow completed"
    if leave_record.status == EmployeeLeave.STATUS_REJECTED:
        return "Workflow closed as rejected"
    if leave_record.status == EmployeeLeave.STATUS_CANCELLED:
        return "Workflow cancelled"

    return leave_record.get_status_display()


def can_user_review_leave_stage(user, leave_record):
    if not user or not leave_record or leave_record.status != EmployeeLeave.STATUS_PENDING:
        return False

    if is_branch_scoped_supervisor(user):
        return leave_record.current_stage == EmployeeLeave.STAGE_SUPERVISOR_REVIEW

    if is_operations_manager_user(user):
        return leave_record.current_stage == EmployeeLeave.STAGE_OPERATIONS_REVIEW

    if is_hr_user(user) or is_admin_compatible(user):
        return leave_record.current_stage == EmployeeLeave.STAGE_HR_REVIEW

    return False


def can_cancel_leave(user, leave_record):
    if can_review_leave(user):
        if is_branch_scoped_supervisor(user):
            return can_supervisor_view_employee(user, leave_record.employee)
        return True

    if is_self_employee(user, leave_record.employee):
        return leave_record.status == EmployeeLeave.STATUS_PENDING

    return False


def can_create_action_records(user):
    return (
        is_admin_compatible(user)
        or is_hr_user(user)
        or is_operations_manager_user(user)
        or is_branch_scoped_supervisor(user)
    )


def can_manage_attendance_records(user):
    return is_admin_compatible(user) or is_hr_user(user) or is_operations_manager_user(user)


def can_view_attendance_management(user):
    return is_admin_compatible(user) or is_hr_user(user) or is_operations_manager_user(user)


def can_view_branch_self_service(employee):
    return bool(employee and employee.branch_id)


def can_manage_branch_weekly_schedule(user, branch):
    if not user or not user.is_authenticated or not branch:
        return False

    if is_admin_compatible(user) or is_hr_user(user) or is_operations_manager_user(user):
        return True

    linked_employee = get_user_employee_profile(user)
    return bool(
        is_branch_scoped_supervisor(user)
        and linked_employee
        and linked_employee.branch_id == branch.id
    )


def can_add_manual_history(user):
    return is_admin_compatible(user) or is_hr_user(user) or is_operations_manager_user(user)


def can_access_employee_form_dependencies(user):
    return can_create_or_edit_employees(user) or can_transfer_employee(user)


def get_employee_access_redirect(user, employee=None):
    if employee and can_view_employee_profile(user, employee):
        return redirect("employees:employee_detail", pk=employee.pk)

    linked_employee = get_user_employee_profile(user)
    if linked_employee and can_view_employee_profile(user, linked_employee):
        return redirect("employees:employee_detail", pk=linked_employee.pk)

    if can_view_employee_directory(user):
        return redirect("employees:employee_list")

    raise PermissionDenied("You do not have permission to access this area.")


def deny_employee_access(request, message, employee=None):
    messages.error(request, message)
    return get_employee_access_redirect(request.user, employee=employee)


def deny_json_access(message="You do not have permission to access this endpoint."):
    return JsonResponse({"results": [], "error": message}, status=403)


def format_history_value(value):
    if value in [None, ""]:
        return "empty"

    if hasattr(value, "strftime"):
        return value.strftime("%B %d, %Y")

    if isinstance(value, bool):
        return "Active" if value else "Inactive"

    return str(value)


def build_employee_change_summary(old_employee, new_employee):
    changes = []

    tracked_fields = [
        ("full_name", "Full name"),
        ("email", "Email"),
        ("phone", "Phone"),
        ("birth_date", "Birth date"),
        ("marital_status", "Marital status"),
        ("nationality", "Nationality"),
        ("company", "Company"),
        ("department", "Department"),
        ("branch", "Branch"),
        ("section", "Section"),
        ("job_title", "Job title"),
        ("hire_date", "Hire date"),
        ("salary", "Salary"),
        ("is_active", "Operational status"),
        ("notes", "Notes"),
    ]

    for field_name, label in tracked_fields:
        old_value = getattr(old_employee, field_name)
        new_value = getattr(new_employee, field_name)
        if old_value != new_value:
            changes.append(
                f"{label} changed from {format_history_value(old_value)} to {format_history_value(new_value)}."
            )

    return " ".join(changes) if changes else "Employee profile details were updated."


def build_employee_transfer_summary(old_employee, new_employee, transfer_note=""):
    changes = []

    tracked_fields = [
        ("company", "Company"),
        ("department", "Department"),
        ("branch", "Branch"),
        ("section", "Section"),
        ("job_title", "Job title"),
    ]

    for field_name, label in tracked_fields:
        old_value = getattr(old_employee, field_name)
        new_value = getattr(new_employee, field_name)
        if old_value != new_value:
            changes.append(
                f"{label} changed from {format_history_value(old_value)} to {format_history_value(new_value)}."
            )

    if transfer_note:
        changes.append(f"Transfer note: {transfer_note}")

    return " ".join(changes)




PREVIEWABLE_FILE_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "webp", "gif", "bmp", "txt"}


def get_file_extension(file_field):
    if not file_field or not getattr(file_field, "name", ""):
        return ""
    return Path(file_field.name).suffix.lower().lstrip(".")


def build_browser_file_response(file_field, *, force_download=False):
    if not file_field or not getattr(file_field, "name", ""):
        raise Http404("The requested file is not available.")

    storage = getattr(file_field, "storage", None)
    if storage and not storage.exists(file_field.name):
        raise Http404("The requested file is not available on this system.")

    filename = Path(file_field.name).name
    extension = get_file_extension(file_field)
    content_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    as_attachment = force_download or extension not in PREVIEWABLE_FILE_EXTENSIONS

    file_handle = file_field.open("rb")
    response = FileResponse(
        file_handle,
        as_attachment=as_attachment,
        filename=filename,
        content_type=content_type,
    )

    if not as_attachment:
        response["Content-Disposition"] = f'inline; filename="{filename}"'

    return response

def build_document_summary(document):
    parts = [
        f"Document type: {document.get_document_type_display()}",
        f"Reference number: {document.reference_number or 'empty'}",
    ]

    if document.issue_date:
        parts.append(f"Issue date: {document.issue_date.strftime('%B %d, %Y')}")
    if document.expiry_date:
        parts.append(f"Expiry date: {document.expiry_date.strftime('%B %d, %Y')}")
    if document.description:
        parts.append(f"Description: {document.description}")

    return ". ".join(parts)


def build_leave_request_summary(leave_record):
    parts = [
        f"Leave request created as {leave_record.get_leave_type_display()}.",
        f"Date range: {leave_record.start_date.strftime('%B %d, %Y')} to {leave_record.end_date.strftime('%B %d, %Y')}.",
        f"Total days: {leave_record.total_days}.",
        f"Status: {leave_record.get_status_display()}.",
    ]

    if leave_record.reason:
        parts.append(f"Reason: {leave_record.reason}")

    return " ".join(parts)


def build_leave_status_summary(leave_record, previous_status, new_status, approval_note=""):
    parts = [
        f"{leave_record.get_leave_type_display()} status changed from {previous_status} to {new_status}.",
        f"Date range: {leave_record.start_date.strftime('%B %d, %Y')} to {leave_record.end_date.strftime('%B %d, %Y')}.",
        f"Total days: {leave_record.total_days}.",
    ]

    if approval_note:
        parts.append(f"Action note: {approval_note}")

    return " ".join(parts)


def build_action_record_summary(action_record):
    lines = [
        f"Action Type: {action_record.get_action_type_display()}",
        f"Action Date: {format_history_value(action_record.action_date)}",
        f"Status: {action_record.get_status_display()}",
        f"Severity: {action_record.get_severity_display()}",
    ]

    if action_record.description:
        lines.append(f"Note: {action_record.description}")

    return "\n".join(lines)


def build_attendance_ledger_summary(attendance_entry):
    lines = [
        f"Attendance Date: {format_history_value(attendance_entry.attendance_date)}",
        f"Day Status: {attendance_entry.get_day_status_display()}",
        f"Scheduled Hours: {attendance_entry.scheduled_hours}",
        f"Worked Hours: {attendance_entry.worked_hours}",
        f"Late Minutes: {attendance_entry.late_minutes}",
        f"Early Departure Minutes: {attendance_entry.early_departure_minutes}",
        f"Overtime Minutes: {attendance_entry.overtime_minutes}",
        f"Paid Day: {'Yes' if attendance_entry.is_paid_day else 'No'}",
        f"Source: {attendance_entry.get_source_display()}",
    ]

    if attendance_entry.notes:
        lines.append(f"Notes: {attendance_entry.notes}")

    return "\n".join(lines)


def build_attendance_correction_summary(correction):
    lines = [
        f"Attendance Date: {format_history_value(correction.linked_attendance.attendance_date)}",
        f"Requested Status: {correction.get_requested_day_status_display()}",
        f"Requested Scheduled Hours: {correction.requested_scheduled_hours}",
        f"Requested Late Minutes: {correction.requested_late_minutes}",
        f"Requested Early Departure Minutes: {correction.requested_early_departure_minutes}",
        f"Requested Overtime Minutes: {correction.requested_overtime_minutes}",
        f"Review Status: {correction.get_status_display()}",
    ]

    if correction.request_reason:
        lines.append(f"Request Reason: {correction.request_reason}")
    if correction.review_notes:
        lines.append(f"Review Notes: {correction.review_notes}")

    return "\n".join(lines)


def get_month_date_range(target_date):
    start_date = target_date.replace(day=1)
    if target_date.month == 12:
        next_month = target_date.replace(year=target_date.year + 1, month=1, day=1)
    else:
        next_month = target_date.replace(month=target_date.month + 1, day=1)
    end_date = next_month - timedelta(days=1)
    return start_date, end_date


def build_attendance_summary(attendance_entries):
    total_scheduled_hours = Decimal("0.00")
    total_worked_hours = Decimal("0.00")
    total_overtime_minutes = 0
    total_late_minutes = 0
    total_early_departure_minutes = 0

    present_count = 0
    absence_count = 0
    leave_count = 0
    weekly_off_count = 0
    holiday_count = 0
    paid_day_count = 0
    unpaid_day_count = 0

    for entry in attendance_entries:
        total_scheduled_hours += entry.scheduled_hours or Decimal("0.00")
        total_worked_hours += entry.worked_hours or Decimal("0.00")
        total_overtime_minutes += entry.overtime_minutes or 0
        total_late_minutes += entry.late_minutes or 0
        total_early_departure_minutes += entry.early_departure_minutes or 0

        if entry.is_paid_day:
            paid_day_count += 1
        else:
            unpaid_day_count += 1

        if entry.day_status == EmployeeAttendanceLedger.DAY_STATUS_PRESENT:
            present_count += 1
        elif entry.day_status == EmployeeAttendanceLedger.DAY_STATUS_ABSENT:
            absence_count += 1
        elif entry.day_status == EmployeeAttendanceLedger.DAY_STATUS_WEEKLY_OFF:
            weekly_off_count += 1
        elif entry.day_status == EmployeeAttendanceLedger.DAY_STATUS_HOLIDAY:
            holiday_count += 1
        elif entry.day_status in {
            EmployeeAttendanceLedger.DAY_STATUS_PAID_LEAVE,
            EmployeeAttendanceLedger.DAY_STATUS_UNPAID_LEAVE,
            EmployeeAttendanceLedger.DAY_STATUS_SICK_LEAVE,
        }:
            leave_count += 1

    punctuality_deduction_hours = (
        Decimal(total_late_minutes + total_early_departure_minutes) / Decimal("60")
    ).quantize(Decimal("0.01"))
    overtime_hours = (Decimal(total_overtime_minutes) / Decimal("60")).quantize(Decimal("0.01"))

    return {
        "attendance_total": len(attendance_entries),
        "present_attendance_count": present_count,
        "absence_attendance_count": absence_count,
        "leave_attendance_count": leave_count,
        "weekly_off_attendance_count": weekly_off_count,
        "holiday_attendance_count": holiday_count,
        "paid_day_count": paid_day_count,
        "unpaid_day_count": unpaid_day_count,
        "total_scheduled_hours": total_scheduled_hours.quantize(Decimal("0.01")),
        "total_worked_hours": total_worked_hours.quantize(Decimal("0.01")),
        "total_overtime_minutes": total_overtime_minutes,
        "total_overtime_hours": overtime_hours,
        "total_late_minutes": total_late_minutes,
        "total_early_departure_minutes": total_early_departure_minutes,
        "punctuality_deduction_hours": punctuality_deduction_hours,
    }


def apply_attendance_management_form_scope(form, user):
    if not form:
        return form

    scoped_branch = get_user_scope_branch(user)
    employee_queryset = get_employee_directory_queryset_for_user(
        user,
        Employee.objects.select_related(
            "company",
            "branch",
            "department",
            "section",
            "job_title",
        ),
    ).order_by("full_name", "employee_id")

    form.fields["employee"].queryset = employee_queryset

    if is_branch_scoped_supervisor(user) and scoped_branch:
        form.fields["company"].queryset = Company.objects.filter(
            id=scoped_branch.company_id,
            is_active=True,
        ).order_by("name")
        form.fields["branch"].queryset = Branch.objects.filter(
            id=scoped_branch.id,
            is_active=True,
        ).order_by("name")
        form.fields["department"].queryset = Department.objects.filter(
            company_id=scoped_branch.company_id,
            is_active=True,
        ).order_by("name")
        form.fields["section"].queryset = Section.objects.filter(
            department__company_id=scoped_branch.company_id,
            is_active=True,
        ).order_by("name")
        return form

    form.fields["company"].queryset = Company.objects.filter(is_active=True).order_by("name")
    form.fields["branch"].queryset = Branch.objects.filter(is_active=True).order_by("name")
    form.fields["department"].queryset = Department.objects.filter(is_active=True).order_by("name")
    form.fields["section"].queryset = Section.objects.filter(is_active=True).order_by("name")
    return form


def build_attendance_history_pagination(page_obj):
    if not page_obj:
        return []

    paginator = page_obj.paginator
    current_page = page_obj.number
    total_pages = paginator.num_pages
    page_numbers = {1, total_pages}

    for page_number in range(current_page - 1, current_page + 2):
        if 1 <= page_number <= total_pages:
            page_numbers.add(page_number)

    sorted_pages = sorted(page_numbers)
    pagination_items = []
    previous_page = None

    for page_number in sorted_pages:
        if previous_page is not None and page_number - previous_page > 1:
            pagination_items.append({"type": "ellipsis"})
        pagination_items.append(
            {
                "type": "page",
                "number": page_number,
                "is_current": page_number == current_page,
            }
        )
        previous_page = page_number

    return pagination_items


def build_attendance_management_filter_state(request, user=None):
    initial = AttendanceManagementFilterForm.default_initial()
    has_query = bool(request.GET)
    form = AttendanceManagementFilterForm(request.GET or None, initial=initial)
    if user is not None:
        apply_attendance_management_form_scope(form, user)

    today = timezone.localdate()
    start_date = initial["start_date"]
    end_date = initial["end_date"]
    filter_type = initial["filter_type"]
    search_value = ""
    employee = None
    company = None
    branch = None
    department = None
    section = None
    day_status = ""

    if form.is_bound and form.is_valid():
        filter_type = form.cleaned_data.get("filter_type") or AttendanceFilterForm.FILTER_THIS_MONTH
        start_date = form.cleaned_data.get("start_date")
        end_date = form.cleaned_data.get("end_date")
        search_value = form.cleaned_data.get("search") or ""
        employee = form.cleaned_data.get("employee")
        company = form.cleaned_data.get("company")
        branch = form.cleaned_data.get("branch")
        department = form.cleaned_data.get("department")
        section = form.cleaned_data.get("section")
        day_status = form.cleaned_data.get("day_status") or ""

        if filter_type == AttendanceFilterForm.FILTER_THIS_MONTH:
            start_date, end_date = get_month_date_range(today)
        elif filter_type == AttendanceFilterForm.FILTER_LAST_MONTH:
            last_month_anchor = today.replace(day=1) - timedelta(days=1)
            start_date, end_date = get_month_date_range(last_month_anchor)
        elif filter_type == AttendanceFilterForm.FILTER_ALL:
            start_date = None
            end_date = None
    elif has_query:
        start_date = None
        end_date = None
        filter_type = AttendanceFilterForm.FILTER_ALL
    else:
        form = AttendanceManagementFilterForm(initial=initial)
        if user is not None:
            apply_attendance_management_form_scope(form, user)

    if filter_type == AttendanceFilterForm.FILTER_CUSTOM and start_date and end_date:
        period_label = f"{start_date:%b %d, %Y} to {end_date:%b %d, %Y}"
    elif filter_type == AttendanceFilterForm.FILTER_LAST_MONTH and start_date and end_date:
        period_label = f"Last Month ({start_date:%b %d, %Y} to {end_date:%b %d, %Y})"
    elif filter_type == AttendanceFilterForm.FILTER_ALL:
        period_label = "All attendance records"
    else:
        period_label = f"This Month ({start_date:%b %d, %Y} to {end_date:%b %d, %Y})"

    return {
        "form": form,
        "search_value": search_value,
        "employee": employee,
        "company": company,
        "branch": branch,
        "department": department,
        "section": section,
        "day_status": day_status,
        "filter_type": filter_type,
        "start_date": start_date,
        "end_date": end_date,
        "period_label": period_label,
        "is_applied": has_query,
    }


def build_attendance_filter_state(request):
    initial = AttendanceFilterForm.default_initial()
    form = AttendanceFilterForm(request.GET or initial)

    if request.GET and form.is_valid():
        cleaned = form.cleaned_data
        filter_type = cleaned.get("filter_type") or AttendanceFilterForm.FILTER_THIS_MONTH
        start_date = cleaned.get("start_date")
        end_date = cleaned.get("end_date")
    else:
        filter_type = initial["filter_type"]
        start_date = initial["start_date"]
        end_date = initial["end_date"]

    today = timezone.localdate()

    if filter_type == AttendanceFilterForm.FILTER_THIS_MONTH:
        start_date, end_date = get_month_date_range(today)
        period_label = f"{start_date.strftime('%B %Y')}"
    elif filter_type == AttendanceFilterForm.FILTER_LAST_MONTH:
        previous_month_anchor = today.replace(day=1) - timedelta(days=1)
        start_date, end_date = get_month_date_range(previous_month_anchor)
        period_label = f"{start_date.strftime('%B %Y')}"
    elif filter_type == AttendanceFilterForm.FILTER_CUSTOM and start_date and end_date:
        period_label = f"{start_date.strftime('%b %d, %Y')} to {end_date.strftime('%b %d, %Y')}"
    else:
        start_date = None
        end_date = None
        period_label = "All attendance records"

    return {
        "form": form,
        "filter_type": filter_type,
        "start_date": start_date,
        "end_date": end_date,
        "period_label": period_label,
        "is_applied": bool(start_date or end_date or filter_type != AttendanceFilterForm.FILTER_THIS_MONTH),
    }


class EmployeeListView(LoginRequiredMixin, ListView):
    model = Employee
    template_name = "employees/employee_list.html"
    context_object_name = "employees"
    paginate_by = 5

    def dispatch(self, request, *args, **kwargs):
        if can_view_employee_directory(request.user):
            return super().dispatch(request, *args, **kwargs)

        if is_supervisor_user(request.user):
            messages.error(
                request,
                "Supervisor access requires linking this login account to an employee profile with an assigned branch.",
            )
            return redirect("dashboard_home")

        linked_employee = get_user_employee_profile(request.user)
        if linked_employee:
            return redirect("employees:employee_detail", pk=linked_employee.pk)

        raise PermissionDenied("You do not have permission to view the employee directory.")

    def get_queryset(self):
        queryset = get_employee_directory_queryset_for_user(
            self.request.user,
            Employee.objects.select_related(
                "user",
                "company",
                "department",
                "branch",
                "section",
                "job_title",
            ).order_by("employee_id", "full_name"),
        )

        search = self.request.GET.get("search", "").strip()
        company_id = self.request.GET.get("company", "").strip()
        department_id = self.request.GET.get("department", "").strip()
        branch_id = self.request.GET.get("branch", "").strip()
        section_id = self.request.GET.get("section", "").strip()
        job_title_id = self.request.GET.get("job_title", "").strip()
        status = self.request.GET.get("status", "").strip()

        if search:
            queryset = queryset.filter(
                Q(full_name__icontains=search)
                | Q(employee_id__icontains=search)
                | Q(email__icontains=search)
                | Q(phone__icontains=search)
            )

        if company_id:
            queryset = queryset.filter(company_id=company_id)

        if department_id:
            queryset = queryset.filter(department_id=department_id)

        if branch_id:
            queryset = queryset.filter(branch_id=branch_id)

        if section_id:
            queryset = queryset.filter(section_id=section_id)

        if job_title_id:
            queryset = queryset.filter(job_title_id=job_title_id)

        if status == "active":
            queryset = queryset.filter(is_active=True)
        elif status == "inactive":
            queryset = queryset.filter(is_active=False)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        filtered_queryset = self.get_queryset()
        overall_queryset = Employee.objects.all()

        context["search_value"] = self.request.GET.get("search", "").strip()
        context["selected_company"] = self.request.GET.get("company", "").strip()
        context["selected_department"] = self.request.GET.get("department", "").strip()
        context["selected_branch"] = self.request.GET.get("branch", "").strip()
        context["selected_section"] = self.request.GET.get("section", "").strip()
        context["selected_job_title"] = self.request.GET.get("job_title", "").strip()
        context["selected_status"] = self.request.GET.get("status", "").strip()

        scoped_branch = get_user_scope_branch(self.request.user)
        if is_branch_scoped_supervisor(self.request.user) and scoped_branch:
            context["companies"] = Company.objects.filter(id=scoped_branch.company_id, is_active=True).order_by("name")
            context["departments"] = (
                Department.objects.filter(company_id=scoped_branch.company_id, is_active=True)
                .select_related("company")
                .order_by("company__name", "name")
            )
            context["branches"] = (
                Branch.objects.filter(id=scoped_branch.id, is_active=True)
                .select_related("company")
                .order_by("company__name", "name")
            )
            context["sections"] = (
                Section.objects.filter(department__company_id=scoped_branch.company_id, is_active=True)
                .select_related("department", "department__company")
                .order_by("department__company__name", "department__name", "name")
            )
            context["job_titles"] = (
                JobTitle.objects.filter(department__company_id=scoped_branch.company_id, is_active=True)
                .select_related("department", "department__company", "section")
                .order_by("department__company__name", "department__name", "name")
            )
        else:
            context["companies"] = Company.objects.filter(is_active=True).order_by("name")
            context["departments"] = (
                Department.objects.filter(is_active=True)
                .select_related("company")
                .order_by("company__name", "name")
            )
            context["branches"] = (
                Branch.objects.filter(is_active=True)
                .select_related("company")
                .order_by("company__name", "name")
            )
            context["sections"] = (
                Section.objects.filter(is_active=True)
                .select_related("department", "department__company")
                .order_by("department__company__name", "department__name", "name")
            )
            context["job_titles"] = (
                JobTitle.objects.filter(is_active=True)
                .select_related("department", "department__company", "section")
                .order_by("department__company__name", "department__name", "name")
            )
        context["employment_status_choices"] = Employee.EMPLOYMENT_STATUS_CHOICES

        context["filtered_total"] = filtered_queryset.count()
        context["filtered_active"] = filtered_queryset.filter(is_active=True).count()
        context["filtered_inactive"] = filtered_queryset.filter(is_active=False).count()
        context["overall_total"] = overall_queryset.count()
        context["can_manage_employees"] = can_create_or_edit_employees(self.request.user)
        context["can_edit_employee_records"] = can_create_or_edit_employees(self.request.user)
        context["can_delete_employee_records"] = can_delete_employee(self.request.user)
        context["is_branch_scoped_supervisor"] = is_branch_scoped_supervisor(self.request.user)
        context["scoped_branch"] = get_user_scope_branch(self.request.user)

        query_params = self.request.GET.copy()
        query_params.pop("page", None)
        context["pagination_querystring"] = query_params.urlencode()

        page_obj = context.get("page_obj")
        if page_obj:
            paginator = page_obj.paginator
            current_page = page_obj.number
            total_pages = paginator.num_pages
            page_numbers = {1, total_pages}

            for page_number in range(current_page - 1, current_page + 2):
                if 1 <= page_number <= total_pages:
                    page_numbers.add(page_number)

            sorted_pages = sorted(page_numbers)
            pagination_items = []
            previous_page = None

            for page_number in sorted_pages:
                if previous_page is not None and page_number - previous_page > 1:
                    pagination_items.append({"type": "ellipsis"})
                pagination_items.append(
                    {
                        "type": "page",
                        "number": page_number,
                        "is_current": page_number == current_page,
                    }
                )
                previous_page = page_number

            context["pagination_items"] = pagination_items

        return context


def get_department_manager_employee(employee):
    if not employee or not employee.department_id:
        return None

    return (
        Employee.objects.select_related("job_title", "branch", "department", "section", "company")
        .filter(department_id=employee.department_id, is_active=True, job_title__isnull=False)
        .filter(job_title__name__iregex=r"(?i)manager")
        .exclude(pk=getattr(employee, "pk", None))
        .order_by("full_name")
        .first()
    )


def get_department_manager_display(employee):
    manager_employee = get_department_manager_employee(employee)
    if manager_employee:
        return manager_employee.full_name

    department = getattr(employee, "department", None)
    manager_name = getattr(department, "manager_name", "") or ""
    return manager_name.strip()


def get_branch_supervisor_display(employee):
    supervisor_employee = get_employee_supervisor(employee)
    if supervisor_employee:
        return supervisor_employee.full_name

    section = getattr(employee, "section", None)
    supervisor_name = getattr(section, "supervisor_name", "") or ""
    return supervisor_name.strip()


def get_team_leader_display(employee):
    team_leader_employee = get_employee_team_leader(employee)
    if team_leader_employee:
        return team_leader_employee.full_name
    return ""


def get_employee_supervisor(employee):
    if not employee.branch_id:
        return None

    return (
        Employee.objects.select_related("job_title", "branch", "department", "section", "company")
        .filter(branch_id=employee.branch_id, is_active=True, job_title__isnull=False)
        .filter(job_title__name__iregex=r"(?i)supervisor")
        .exclude(pk=employee.pk)
        .order_by("full_name")
        .first()
    )




def get_short_structure_label(value):
    if value in [None, ""]:
        return "—"

    if hasattr(value, "name"):
        display_value = (getattr(value, "name", "") or "").strip()
    else:
        display_value = str(value).strip()

    if not display_value:
        return "—"

    parts = [part.strip() for part in re.split(r"\s*-\s*", display_value) if part.strip()]
    if len(parts) > 1:
        return parts[-1]

    if hasattr(value, "department") and getattr(value, "department", None):
        department_name = (getattr(value.department, "name", "") or "").strip()
        if department_name:
            lowered_display = display_value.lower()
            lowered_department = department_name.lower()

            if lowered_display.startswith(lowered_department + " "):
                shortened_value = display_value[len(department_name):].strip()
                if shortened_value:
                    return shortened_value

            if lowered_display.startswith(lowered_department + "-"):
                shortened_value = display_value[len(department_name):].lstrip("-").strip()
                if shortened_value:
                    return shortened_value

    return display_value

def get_employee_team_leader(employee):
    if not employee or not employee.branch_id:
        return None

    leadership_queryset = Employee.objects.select_related(
        "job_title",
        "branch",
        "department",
        "section",
        "company",
    ).filter(
        branch_id=employee.branch_id,
        is_active=True,
        job_title__isnull=False,
    ).filter(
        Q(job_title__name__iregex=r"(?i)team\s*leader")
        | Q(job_title__name__iregex=r"(?i)leader")
    )

    if employee.section_id:
        same_section_team_leader = (
            leadership_queryset
            .filter(section_id=employee.section_id)
            .order_by(
                Case(
                    When(pk=employee.pk, then=0),
                    default=1,
                    output_field=IntegerField(),
                ),
                "full_name",
            )
            .first()
        )
        if same_section_team_leader:
            return same_section_team_leader

    return leadership_queryset.order_by(
        Case(
            When(pk=employee.pk, then=0),
            default=1,
            output_field=IntegerField(),
        ),
        "full_name",
    ).first()


def build_branch_team_structure(employee):
    if not employee.branch_id:
        return {"branch_team_members": [], "branch_team_groups": [], "branch_team_total": 0}

    branch_team_queryset = (
        Employee.objects.select_related("job_title", "branch", "department", "section", "company")
        .filter(branch_id=employee.branch_id, is_active=True)
        .order_by("full_name")
    )
    branch_team_members = list(branch_team_queryset)

    leadership_patterns = [
        ("Supervisor", [r"(?i)supervisor"]),
        ("Team Leader", [r"(?i)team\s*leader", r"(?i)leader"]),
        ("Team Members", []),
    ]

    grouped = []
    used_ids = set()

    for label, patterns in leadership_patterns:
        if patterns:
            members = []
            for member in branch_team_members:
                job_title_name = (member.job_title.name if member.job_title else "") or ""
                if any(re.search(pattern, job_title_name) for pattern in patterns):
                    members.append(member)
                    used_ids.add(member.pk)
        else:
            members = [member for member in branch_team_members if member.pk not in used_ids]

        if members:
            grouped.append({"label": label, "members": members})

    return {
        "branch_team_members": branch_team_members,
        "branch_team_groups": grouped,
        "branch_team_total": len(branch_team_members),
    }


FREE_SCHEDULE_GRID_COLUMN_COUNT = 10
FREE_SCHEDULE_GRID_DAY_THEMES = {
    1: "sunday",
    2: "monday",
    3: "tuesday",
    4: "wednesday",
    5: "thursday",
    6: "friday",
    7: "saturday",
    8: "notes",
    9: "orders",
    10: "followup",
}
FREE_SCHEDULE_SHIFT_OPTIONS = [
    "",
    "9 am to 5 pm",
    "2 pm to 10 pm",
    "3 pm to 11 pm",
    "Off",
    "Extra Off",
    "Morning",
    "Evening",
    "Split Shift",
]


def build_branch_schedule_free_grid(branch):
    if not branch:
        return {
            "free_grid_columns": [],
            "free_grid_headers": [],
            "free_grid_rows": [],
            "free_grid_filled_cells": 0,
        }

    team_members = list(Employee.objects.select_related("job_title").filter(branch=branch, is_active=True).order_by("full_name", "employee_id"))
    employee_options = [
        {
            "id": member.id,
            "label": member.full_name,
            "job_title": getattr(getattr(member, "job_title", None), "name", "") or "",
        }
        for member in team_members
    ]
    employee_map = {member.id: member for member in team_members}
    row_total = len(team_members)
    existing_rows = {
        row.row_index: row
        for row in BranchScheduleGridRow.objects.select_related("employee", "employee__job_title").filter(
            branch=branch,
            row_index__lte=max(row_total, 1),
        )
    }
    existing_headers = {
        header.column_index: header.label
        for header in BranchScheduleGridHeader.objects.filter(branch=branch)
    }
    existing_cells = {
        (cell.row_index, cell.column_index): cell.value
        for cell in BranchScheduleGridCell.objects.filter(branch=branch, row_index__lte=max(row_total, 1))
    }
    columns = [
        {
            "index": index,
            "label": f"Column {index}",
            "theme": FREE_SCHEDULE_GRID_DAY_THEMES.get(index, "generic"),
        }
        for index in range(1, FREE_SCHEDULE_GRID_COLUMN_COUNT + 1)
    ]
    free_grid_headers = [
        {
            "column_index": 0,
            "input_name": "header_0",
            "value": existing_headers.get(0, FREE_SCHEDULE_GRID_DEFAULT_HEADERS[0]),
            "default_label": FREE_SCHEDULE_GRID_DEFAULT_HEADERS[0],
        },
        {
            "column_index": 1,
            "input_name": "header_1",
            "value": existing_headers.get(1, FREE_SCHEDULE_GRID_DEFAULT_HEADERS[1]),
            "default_label": FREE_SCHEDULE_GRID_DEFAULT_HEADERS[1],
        },
    ] + [
        {
            "column_index": column["index"] + 1,
            "input_name": f"header_{column['index'] + 1}",
            "value": existing_headers.get(
                column["index"] + 1,
                FREE_SCHEDULE_GRID_DEFAULT_HEADERS.get(column["index"] + 1, column["label"]),
            ),
            "default_label": FREE_SCHEDULE_GRID_DEFAULT_HEADERS.get(column["index"] + 1, column["label"]),
            "theme": column["theme"],
        }
        for column in columns
    ]
    roster_day_columns = [column for column in columns if column["theme"] in {
        "sunday",
        "monday",
        "tuesday",
        "wednesday",
        "thursday",
        "friday",
        "saturday",
    }]
    roster_extra_columns = [column for column in columns if column["theme"] not in {
        "sunday",
        "monday",
        "tuesday",
        "wednesday",
        "thursday",
        "friday",
        "saturday",
    }]
    rows = []
    filled_cells = 0

    for row_index in range(1, row_total + 1):
        assigned_row = existing_rows.get(row_index)
        assigned_employee = getattr(assigned_row, "employee", None)
        cells = []
        for column in columns:
            cell_value = existing_cells.get((row_index, column["index"]), "")
            if cell_value:
                filled_cells += 1
            cells.append(
                {
                    "column_index": column["index"],
                    "value": cell_value,
                    "input_name": f"grid_{row_index}_{column['index']}",
                    "row_index": row_index,
                    "theme": column["theme"],
                }
            )

        rows.append(
            {
                "row_index": row_index,
                "employee": assigned_employee,
                "employee_select_name": f"row_employee_{row_index}",
                "employee_job_title": getattr(getattr(assigned_employee, "job_title", None), "name", "") or "",
                "employee_options": employee_options,
                "cells": cells,
            }
        )

    return {
        "free_grid_columns": columns,
        "free_grid_headers": free_grid_headers,
        "free_grid_rows": rows,
        "free_grid_filled_cells": filled_cells,
        "free_grid_row_total": row_total,
        "free_grid_employee_map": employee_map,
        "roster_day_columns": roster_day_columns,
        "roster_extra_columns": roster_extra_columns,
        "free_grid_shift_options": FREE_SCHEDULE_SHIFT_OPTIONS,
    }


def build_schedule_week_days(week_start):
    if not week_start:
        return []
    return [week_start + timedelta(days=index) for index in range(7)]


def get_pending_off_days_for_week(employee, week_start, week_end, pending_off_map=None):
    if not employee or not week_start or not week_end:
        return 0

    if pending_off_map and employee.id in pending_off_map:
        return pending_off_map[employee.id]

    pending_leave_records = employee.leave_records.filter(
        status=EmployeeLeave.STATUS_PENDING,
        start_date__lte=week_end,
        end_date__gte=week_start,
    )
    total_pending_days = 0

    for leave_record in pending_leave_records:
        overlap_start = max(week_start, leave_record.start_date)
        overlap_end = min(week_end, leave_record.end_date)
        if overlap_start <= overlap_end:
            total_pending_days += count_policy_working_days(overlap_start, overlap_end)

    return total_pending_days


def build_branch_weekly_schedule_summary(branch, week_start):
    if not branch or not week_start:
        return {
            "team_members": [],
            "team_schedule_rows": [],
            "schedule_entries": [],
            "week_days": [],
            "schedule_total": 0,
            "completed_total": 0,
            "in_progress_total": 0,
            "planned_total": 0,
            "on_hold_total": 0,
        }

    week_end = week_start + timedelta(days=6)
    team_members = list(
        Employee.objects.select_related("job_title", "section")
        .filter(branch=branch, is_active=True)
        .order_by("full_name", "employee_id")
    )
    row_order_map = {
        row.employee_id: row.row_index
        for row in BranchScheduleGridRow.objects.filter(branch=branch, employee__isnull=False).select_related("employee")
    }
    team_members.sort(
        key=lambda member: (
            row_order_map.get(member.id, 9999),
            member.full_name.lower(),
            member.employee_id.lower(),
        )
    )
    week_days = build_schedule_week_days(week_start)
    schedule_entries = list(
        BranchWeeklyScheduleEntry.objects.select_related(
            "employee",
            "employee__job_title",
            "employee__section",
            "duty_option",
        )
        .filter(branch=branch, week_start=week_start)
        .order_by("schedule_date", "employee__full_name", "id")
    )
    pending_off_map = {
        record.employee_id: record.pending_off_count
        for record in BranchWeeklyPendingOff.objects.filter(branch=branch, week_start=week_start)
    }

    entries_by_employee_and_date = {}
    for entry in schedule_entries:
        entries_by_employee_and_date[(entry.employee_id, entry.schedule_date)] = entry

    team_schedule_rows = []
    for member in team_members:
        row_cells = []
        member_entries = []
        for current_date in week_days:
            current_entry = entries_by_employee_and_date.get((member.id, current_date))
            row_cells.append(
                {
                    "date": current_date,
                    "entry": current_entry,
                    "has_entry": current_entry is not None,
                    "edit_url": f"{reverse('employees:self_service_weekly_schedule')}?week={week_start.isoformat()}&employee={member.id}&day={current_date.isoformat()}",
                }
            )
            if current_entry is not None:
                member_entries.append(current_entry)

        team_schedule_rows.append(
            {
                "employee": member,
                "entries": member_entries,
                "cells": row_cells,
                "entry_total": len(member_entries),
                "pending_off_total": get_pending_off_days_for_week(
                    member,
                    week_start,
                    week_end,
                    pending_off_map=pending_off_map,
                ),
                "completed_total": sum(
                    1 for entry in member_entries if entry.status == BranchWeeklyScheduleEntry.STATUS_COMPLETED
                ),
                "pending_total": sum(
                    1
                    for entry in member_entries
                    if entry.status in {
                        BranchWeeklyScheduleEntry.STATUS_PLANNED,
                        BranchWeeklyScheduleEntry.STATUS_IN_PROGRESS,
                        BranchWeeklyScheduleEntry.STATUS_ON_HOLD,
                    }
                ),
            }
        )

    return {
        "week_start": week_start,
        "week_end": week_end,
        "team_members": team_members,
        "team_schedule_rows": team_schedule_rows,
        "schedule_entries": schedule_entries,
        "week_days": week_days,
        "schedule_total": len(schedule_entries),
        "completed_total": sum(
            1 for entry in schedule_entries if entry.status == BranchWeeklyScheduleEntry.STATUS_COMPLETED
        ),
        "in_progress_total": sum(
            1 for entry in schedule_entries if entry.status == BranchWeeklyScheduleEntry.STATUS_IN_PROGRESS
        ),
        "planned_total": sum(
            1 for entry in schedule_entries if entry.status == BranchWeeklyScheduleEntry.STATUS_PLANNED
        ),
        "on_hold_total": sum(
            1 for entry in schedule_entries if entry.status == BranchWeeklyScheduleEntry.STATUS_ON_HOLD
        ),
    }


SCHEDULE_IMPORT_COLUMNS = [
    "employee_id",
    "employee_name",
    "schedule_date",
    "duty_label",
    "custom_label",
    "shift_label",
    "start_time",
    "end_time",
    "status",
    "order_note",
]

SCHEDULE_IMPORT_HEADER_ALIASES = {
    "employee code": "employee_id",
    "employee_code": "employee_id",
    "employee id": "employee_id",
    "employee": "employee_name",
    "employee name": "employee_name",
    "date": "schedule_date",
    "duty": "duty_label",
    "duty option": "duty_label",
    "duty_option": "duty_label",
    "shift": "shift_label",
    "shift label": "shift_label",
    "shift_label": "shift_label",
    "custom duty": "custom_label",
    "custom_label": "custom_label",
    "start": "start_time",
    "start time": "start_time",
    "start_time": "start_time",
    "end": "end_time",
    "end time": "end_time",
    "end_time": "end_time",
    "note": "order_note",
    "notes": "order_note",
    "order": "order_note",
    "order note": "order_note",
    "order_note": "order_note",
    "pending off": "pending_off",
    "pending_off": "pending_off",
}

SCHEDULE_WEEKDAY_NAMES = [
    "sunday",
    "monday",
    "tuesday",
    "wednesday",
    "thursday",
    "friday",
    "saturday",
]


def normalize_schedule_import_header(value):
    cleaned = ((value or "").strip().lower()).replace("-", " ").replace("/", " ")
    cleaned = re.sub(r"\s+", " ", cleaned)
    return SCHEDULE_IMPORT_HEADER_ALIASES.get(cleaned, cleaned.replace(" ", "_"))


def parse_schedule_import_date(value):
    if value in [None, ""]:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    formats = ("%Y-%m-%d", "%d-%b-%Y", "%d-%B-%Y", "%d/%m/%Y", "%m/%d/%Y")
    for pattern in formats:
        try:
            return timezone.datetime.strptime(text, pattern).date()
        except ValueError:
            continue

    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def parse_schedule_import_time(value):
    if value in [None, ""]:
        return None
    if hasattr(value, "hour") and hasattr(value, "minute"):
        return value

    text = str(value).strip()
    if not text:
        return None

    for pattern in ("%H:%M", "%I:%M %p", "%I %p"):
        try:
            return timezone.datetime.strptime(text, pattern).time()
        except ValueError:
            continue
    return None


def infer_shift_times_from_label(label):
    text = (label or "").strip()
    if not text:
        return None, None

    range_match = re.match(r"^\s*(.+?)\s+to\s+(.+?)\s*$", text, flags=re.IGNORECASE)
    if not range_match:
        return None, None

    start_time = parse_schedule_import_time(range_match.group(1))
    end_time = parse_schedule_import_time(range_match.group(2))
    return start_time, end_time


def get_schedule_import_raw_rows(uploaded_file):
    filename = (uploaded_file.name or "").lower()

    if filename.endswith(".csv"):
        decoded = uploaded_file.read().decode("utf-8-sig")
        return list(csv.reader(io.StringIO(decoded)))

    workbook = load_workbook(uploaded_file, data_only=True)
    worksheet = workbook.active
    return list(worksheet.iter_rows(values_only=True))


def is_branch_schedule_matrix_format(raw_rows):
    if len(raw_rows) < 2:
        return False

    header_row = [normalize_schedule_import_header(value) for value in raw_rows[0]]
    weekday_hits = sum(1 for value in header_row if value in SCHEDULE_WEEKDAY_NAMES)
    return "employee_id" in header_row and weekday_hits >= 5


def build_schedule_import_rows_from_matrix(raw_rows):
    header_row = list(raw_rows[0])
    date_row = list(raw_rows[1]) if len(raw_rows) > 1 else []
    normalized_headers = [normalize_schedule_import_header(value) for value in header_row]

    employee_id_index = normalized_headers.index("employee_id")
    pending_off_index = normalized_headers.index("pending_off") if "pending_off" in normalized_headers else -1
    employee_name_index = -1
    weekday_column_map = {}

    for index, header in enumerate(normalized_headers):
        if header in SCHEDULE_WEEKDAY_NAMES:
            weekday_column_map[index] = header

    for index, header in enumerate(normalized_headers):
        if index in weekday_column_map or index == employee_id_index or index == pending_off_index:
            continue
        header_text = str(header_row[index] or "").strip()
        if header_text:
            employee_name_index = index
            break

    rows = []
    for raw_row in raw_rows[2:]:
        if not raw_row:
            continue

        employee_id_value = str(raw_row[employee_id_index] or "").strip() if employee_id_index < len(raw_row) else ""
        employee_name_value = str(raw_row[employee_name_index] or "").strip() if employee_name_index >= 0 and employee_name_index < len(raw_row) else ""
        pending_off_value = str(raw_row[pending_off_index] or "").strip() if pending_off_index >= 0 and pending_off_index < len(raw_row) else ""

        if not employee_id_value and not employee_name_value:
            continue

        for column_index, weekday_name in weekday_column_map.items():
            schedule_date = parse_schedule_import_date(date_row[column_index] if column_index < len(date_row) else None)
            duty_value = str(raw_row[column_index] or "").strip() if column_index < len(raw_row) else ""
            if not schedule_date and not duty_value:
                continue

            rows.append(
                {
                    "employee_id": employee_id_value,
                    "employee_name": employee_name_value,
                    "schedule_date": schedule_date,
                    "duty_label": duty_value,
                    "custom_label": "",
                    "shift_label": duty_value,
                    "start_time": "",
                    "end_time": "",
                    "status": BranchWeeklyScheduleEntry.STATUS_PLANNED,
                    "order_note": "",
                    "pending_off": pending_off_value,
                    "weekday_name": weekday_name,
                }
            )

    return rows


def get_schedule_import_rows(uploaded_file):
    raw_rows = get_schedule_import_raw_rows(uploaded_file)
    if not raw_rows:
        return []

    if is_branch_schedule_matrix_format(raw_rows):
        return build_schedule_import_rows_from_matrix(raw_rows)

    headers = [normalize_schedule_import_header(value) for value in raw_rows[0]]
    rows = []
    for raw_row in raw_rows[1:]:
        row = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            row[header] = raw_row[index] if index < len(raw_row) else ""
        rows.append(row)
    return rows


def get_or_create_branch_duty_option_for_import(branch, row, duty_option_map):
    duty_label = str(row.get("duty_label") or row.get("shift_label") or row.get("custom_label") or "").strip()
    if not duty_label:
        return None

    key = duty_label.lower()
    existing_option = duty_option_map.get(key)
    if existing_option:
        return existing_option

    start_time = parse_schedule_import_time(row.get("start_time"))
    end_time = parse_schedule_import_time(row.get("end_time"))
    if not (start_time and end_time):
        inferred_start_time, inferred_end_time = infer_shift_times_from_label(duty_label)
        start_time = start_time or inferred_start_time
        end_time = end_time or inferred_end_time
    lowered = duty_label.lower()

    if lowered in {"off", "day off"}:
        duty_type = BranchWeeklyScheduleEntry.DUTY_TYPE_OFF
        start_time = None
        end_time = None
    elif lowered in {"extra off", "extra_off"}:
        duty_type = BranchWeeklyScheduleEntry.DUTY_TYPE_EXTRA_OFF
        start_time = None
        end_time = None
    elif start_time and end_time:
        duty_type = BranchWeeklyScheduleEntry.DUTY_TYPE_SHIFT
    else:
        duty_type = BranchWeeklyScheduleEntry.DUTY_TYPE_CUSTOM
        start_time = None
        end_time = None

    created_option = BranchWeeklyDutyOption.objects.create(
        branch=branch,
        label=duty_label,
        duty_type=duty_type,
        default_start_time=start_time,
        default_end_time=end_time,
        display_order=BranchWeeklyDutyOption.objects.filter(branch=branch).count() + 1,
        is_active=True,
    )
    duty_option_map[key] = created_option
    return created_option


def import_branch_weekly_schedule_file(*, branch, week_start, uploaded_file, actor_label="", replace_existing=False):
    week_end = week_start + timedelta(days=6)
    team_members = list(Employee.objects.filter(branch=branch, is_active=True).order_by("full_name", "employee_id"))
    employee_map = {member.employee_id.strip().lower(): member for member in team_members if member.employee_id}
    duty_option_map = {
        option.label.strip().lower(): option
        for option in BranchWeeklyDutyOption.objects.filter(branch=branch)
    }
    pending_off_updates = {}

    rows = get_schedule_import_rows(uploaded_file)
    imported_count = 0
    errors = []
    skipped_empty_cells = 0

    if replace_existing:
        BranchWeeklyScheduleEntry.objects.filter(branch=branch, week_start=week_start).delete()
        BranchWeeklyPendingOff.objects.filter(branch=branch, week_start=week_start).delete()

    for row_number, row in enumerate(rows, start=2):
        employee_id_value = str(row.get("employee_id") or "").strip().lower()
        schedule_date = parse_schedule_import_date(row.get("schedule_date"))

        if not employee_id_value and not schedule_date:
            continue

        employee = employee_map.get(employee_id_value)
        if not employee:
            errors.append(f"Row {row_number}: employee_id '{row.get('employee_id')}' was not found in this branch.")
            continue

        pending_off_value = str(row.get("pending_off") or "").strip()
        if pending_off_value.isdigit():
            pending_off_updates[employee.id] = int(pending_off_value)

        if not schedule_date:
            errors.append(f"Row {row_number}: schedule_date is missing or invalid.")
            continue

        if schedule_date < week_start or schedule_date > week_end:
            errors.append(f"Row {row_number}: schedule_date {schedule_date} is outside the selected week.")
            continue

        duty_label = str(row.get("duty_label") or row.get("shift_label") or row.get("custom_label") or "").strip()
        if not duty_label:
            skipped_empty_cells += 1
            continue

        duty_option = get_or_create_branch_duty_option_for_import(branch, row, duty_option_map)
        if not duty_option:
            errors.append(f"Row {row_number}: duty_label or shift_label is required.")
            continue

        status_value = str(row.get("status") or "").strip().lower() or BranchWeeklyScheduleEntry.STATUS_PLANNED
        valid_statuses = {choice[0] for choice in BranchWeeklyScheduleEntry.STATUS_CHOICES}
        if status_value not in valid_statuses:
            status_value = BranchWeeklyScheduleEntry.STATUS_PLANNED

        BranchWeeklyScheduleEntry.objects.update_or_create(
            branch=branch,
            employee=employee,
            schedule_date=schedule_date,
            defaults={
                "week_start": week_start,
                "duty_option": duty_option,
                "title": str(row.get("custom_label") or "").strip(),
                "order_note": str(row.get("order_note") or "").strip(),
                "status": status_value,
                "created_by": actor_label,
                "updated_by": actor_label,
            },
        )
        imported_count += 1

    for employee_id, pending_off_count in pending_off_updates.items():
        BranchWeeklyPendingOff.objects.update_or_create(
            branch=branch,
            employee_id=employee_id,
            week_start=week_start,
            defaults={
                "pending_off_count": pending_off_count,
                "created_by": actor_label,
                "updated_by": actor_label,
            },
        )

    return {
        "imported_count": imported_count,
        "errors": errors,
        "parsed_row_count": len(rows),
        "skipped_empty_cells": skipped_empty_cells,
        "replace_existing": replace_existing,
    }


def build_branch_schedule_export_workbook(branch, week_start, *, include_existing_entries=True):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Branch Schedule"

    team_members = list(
        Employee.objects.select_related("job_title")
        .filter(branch=branch, is_active=True)
        .order_by("full_name", "employee_id")
    )
    entry_map = {}
    if include_existing_entries:
        entry_map = {
            (entry.employee_id, entry.schedule_date): entry
            for entry in BranchWeeklyScheduleEntry.objects.select_related("duty_option").filter(
                branch=branch,
                week_start=week_start,
            )
        }
    pending_off_map = {
        record.employee_id: record.pending_off_count
        for record in BranchWeeklyPendingOff.objects.filter(branch=branch, week_start=week_start)
    }

    week_days = build_schedule_week_days(week_start)
    header_row = [
        week_start.strftime("%B"),
        "Employee Code",
        branch.name,
        "Pending off",
    ] + [day.strftime("%A") for day in week_days]
    date_row = ["", "", "", ""] + [f"{day.day}-{day.strftime('%b-%Y')}" if hasattr(day, "strftime") else "" for day in week_days]

    worksheet.append(header_row)
    worksheet.append(date_row)

    for member in team_members:
        row_values = [
            "",
            member.employee_id,
            member.full_name,
            pending_off_map.get(member.id, 0),
        ]
        for schedule_date in week_days:
            entry = entry_map.get((member.id, schedule_date))
            if entry:
                row_values.append(entry.sheet_value)
            else:
                row_values.append("")
        worksheet.append(row_values)

    instructions = workbook.create_sheet("Instructions")
    instructions.append(["Section", "Meaning"])
    instructions.append(["Row 1", "Main roster header. Keep weekday columns in the same order."])
    instructions.append(["Row 2", "Date row. Keep dates inside the selected branch week."])
    instructions.append(["Employee Code", "Required. Must match a branch employee code in the app."])
    instructions.append([branch.name, "Employee name column for visual use."])
    instructions.append(["Pending off", "Optional number for pending off days in that week."])
    instructions.append(["Day cells", "Use values like 2 pm to 10 pm, 9 am to 5 pm, off, extra off, or any custom duty label."])
    instructions.append(["Import result", "The app stores imported values in BranchWeeklyScheduleEntry and BranchWeeklyPendingOff."])
    return workbook




DOCUMENT_GROUP_LABELS = {
    "passport": "Passport",
    "civil_id": "Civil ID",
    "leave": "Leave Documents",
    "cv": "CV / Resume",
    "warning": "Warnings / Disciplinary",
    "resignation": "Resignations",
    "clearance": "Clearance",
    "transfer": "Transfers / Placement Change",
    "contract": "Contracts",
    "medical": "Medical",
    "payroll": "Payroll",
    "certificate": "Certificates",
    "other": "Other Documents",
}


def normalize_document_text(document):
    parts = [
        getattr(document, "title", "") or "",
        getattr(document, "description", "") or "",
        getattr(document, "reference_number", "") or "",
        getattr(document, "filename", "") or "",
        getattr(document, "get_document_type_display", lambda: "")() or "",
    ]
    return " ".join(parts).strip().lower()


def classify_employee_document(document):
    text = normalize_document_text(document)

    if "passport" in text:
        return "passport"
    if "civil id" in text or "civilid" in text or "civil-id" in text or re.search(r"\bcivil\b", text):
        return "civil_id"
    if getattr(document, "linked_leave_id", None):
        return "leave"
    if "resume" in text or re.search(r"\bcv\b", text):
        return "cv"
    if "warning" in text or "disciplinary" in text or "memo" in text:
        return "warning"
    if "resignation" in text or "termination" in text:
        return "resignation"
    if "clearance" in text:
        return "clearance"
    if "transfer" in text or "placement" in text or "movement" in text:
        return "transfer"
    if document.document_type == EmployeeDocument.DOCUMENT_TYPE_CONTRACT:
        return "contract"
    if document.document_type == EmployeeDocument.DOCUMENT_TYPE_MEDICAL:
        return "medical"
    if document.document_type == EmployeeDocument.DOCUMENT_TYPE_PAYROLL:
        return "payroll"
    if document.document_type == EmployeeDocument.DOCUMENT_TYPE_CERTIFICATE:
        return "certificate"
    return "other"


def build_identity_document_statuses(employee):
    all_documents = list(employee.documents.all())
    statuses = []
    today = timezone.localdate()
    direct_field_map = {
        "passport": ("passport_issue_date", "passport_expiry_date"),
        "civil_id": ("civil_id_issue_date", "civil_id_expiry_date"),
    }

    for key, label in [("passport", "Passport"), ("civil_id", "Civil ID")]:
        matching_documents = [document for document in all_documents if classify_employee_document(document) == key]
        preferred_documents = sorted(
            matching_documents,
            key=lambda document: (
                0 if document.expiry_date else 1,
                -(document.uploaded_at.timestamp() if getattr(document, "uploaded_at", None) else 0),
                -document.pk,
            ),
        )
        selected_document = preferred_documents[0] if preferred_documents else None

        issue_attr, expiry_attr = direct_field_map[key]
        direct_issue_date = getattr(employee, issue_attr, None)
        direct_expiry_date = getattr(employee, expiry_attr, None)

        reference_attr_map = {
            "passport": "passport_reference_number",
            "civil_id": "civil_id_reference_number",
        }
        direct_reference_number = getattr(employee, reference_attr_map.get(key, ""), "") or ""

        issue_date = direct_issue_date or (selected_document.issue_date if selected_document else None)
        expiry_date = direct_expiry_date or (selected_document.expiry_date if selected_document else None)
        reference_number = direct_reference_number or (selected_document.reference_number if selected_document else "")

        if expiry_date:
            days_until_expiry = (expiry_date - today).days
            if days_until_expiry < 0:
                state_key = "expired"
                status_label = "Expired"
                badge_class = "badge-danger"
                days_display = f"{abs(days_until_expiry)} day(s) overdue"
            elif days_until_expiry <= 30:
                state_key = "expiring_soon"
                status_label = "Expiring Soon"
                badge_class = "badge-primary"
                days_display = f"{days_until_expiry} day(s) remaining"
            else:
                state_key = "valid"
                status_label = "Valid"
                badge_class = "badge-success"
                days_display = f"{days_until_expiry} day(s) remaining"
        elif issue_date:
            state_key = "missing"
            status_label = "No Expiry Date"
            badge_class = "badge"
            days_display = "No expiry date"
        elif selected_document:
            state_key = "missing"
            status_label = selected_document.compliance_status_label
            badge_class = selected_document.compliance_badge_class
            days_display = "No expiry date"
        else:
            state_key = "missing"
            status_label = "Not Recorded"
            badge_class = "badge"
            days_display = "No record"

        statuses.append(
            {
                "key": key,
                "label": label,
                "document": selected_document,
                "reference_number": reference_number,
                "issue_date": issue_date,
                "expiry_date": expiry_date,
                "status_label": status_label,
                "badge_class": badge_class,
                "days_remaining_display": days_display,
                "state_key": state_key,
            }
        )

    return statuses


def build_document_group_cards(documents):
    grouped_documents = {}

    for document in documents:
        group_key = classify_employee_document(document)
        grouped_documents.setdefault(group_key, []).append(document)

    ordered_cards = []
    ordered_keys = [
        "passport",
        "civil_id",
        "leave",
        "cv",
        "warning",
        "resignation",
        "clearance",
        "transfer",
        "contract",
        "medical",
        "payroll",
        "certificate",
        "other",
    ]

    for group_key in ordered_keys:
        group_documents = grouped_documents.get(group_key, [])
        if not group_documents:
            continue

        ordered_cards.append(
            {
                "key": group_key,
                "label": DOCUMENT_GROUP_LABELS.get(group_key, "Documents"),
                "count": len(group_documents),
                "documents": group_documents,
                "has_expired": any(document.is_expired for document in group_documents),
                "has_expiring_soon": any(document.is_expiring_soon for document in group_documents),
            }
        )

    return ordered_cards


def build_management_document_group_cards(documents, latest_limit=3, expanded_group_keys=None):
    grouped_documents = {}
    expanded_group_keys = set(expanded_group_keys or [])

    for document in documents:
        group_key = classify_employee_document(document)
        grouped_documents.setdefault(group_key, []).append(document)

    ordered_cards = []
    ordered_keys = [
        "passport",
        "civil_id",
        "leave",
        "cv",
        "warning",
        "resignation",
        "clearance",
        "transfer",
        "contract",
        "medical",
        "payroll",
        "certificate",
        "other",
    ]

    for group_key in ordered_keys:
        group_documents = grouped_documents.get(group_key, [])
        if not group_documents:
            continue

        is_expanded = group_key in expanded_group_keys
        visible_documents = group_documents if is_expanded else group_documents[:latest_limit]

        ordered_cards.append(
            {
                "key": group_key,
                "label": DOCUMENT_GROUP_LABELS.get(group_key, "Documents"),
                "count": len(group_documents),
                "documents": visible_documents,
                "hidden_count": 0 if is_expanded else max(len(group_documents) - latest_limit, 0),
                "is_expanded": is_expanded,
                "has_expired": any(document.is_expired for document in group_documents),
                "has_expiring_soon": any(document.is_expiring_soon for document in group_documents),
                "latest_document": group_documents[0],
            }
        )

    return ordered_cards


def get_summary_value(summary, key, default=None):
    if summary is None:
        return default
    if isinstance(summary, dict):
        return summary.get(key, default)
    return getattr(summary, key, default)


def build_employee_detail_tab_url(employee, *, tab="overview", modal="", anchor=""):
    base_url = reverse("employees:employee_detail", kwargs={"pk": employee.pk})
    query_bits = []
    if tab:
        query_bits.append(f"tab={tab}")
    if modal:
        query_bits.append(f"modal={modal}")
    query_string = f"?{'&'.join(query_bits)}" if query_bits else ""
    anchor_string = f"#{anchor}" if anchor else ""
    return f"{base_url}{query_string}{anchor_string}"


def build_employee_360_overview_cards(
    employee,
    attendance_summary,
    working_time_summary,
    identity_document_statuses,
    leave_records,
    required_submission_requests,
    employee_document_requests,
    payroll_profile,
    payroll_lines,
    payroll_obligations,
    action_records,
):
    today = timezone.localdate()
    attendance_total = attendance_summary.get("attendance_total") or 0
    present_total = attendance_summary.get("present_attendance_count") or 0
    attendance_rate = int(round((present_total / attendance_total) * 100)) if attendance_total else 0
    compliance_alert_total = sum(
        1 for status in identity_document_statuses if status["state_key"] in {"expired", "expiring_soon", "missing"}
    )
    overdue_submission_total = sum(1 for item in required_submission_requests if item.is_overdue)
    open_action_total = sum(
        1 for action_record in action_records if action_record.status in {EmployeeActionRecord.STATUS_OPEN, EmployeeActionRecord.STATUS_UNDER_REVIEW}
    )
    active_obligations = [
        obligation for obligation in payroll_obligations if obligation.status == obligation.STATUS_ACTIVE
    ]
    outstanding_obligation_balance = sum(
        (obligation.remaining_balance or Decimal("0.00")) for obligation in active_obligations
    )
    approved_leave_days_year = sum(
        leave_record.total_days
        for leave_record in leave_records
        if leave_record.status == EmployeeLeave.STATUS_APPROVED
        and (
            leave_record.start_date.year == today.year
            or leave_record.end_date.year == today.year
        )
    )
    pending_requests_total = sum(
        1
        for item in employee_document_requests
        if item.status in {EmployeeDocumentRequest.STATUS_REQUESTED, EmployeeDocumentRequest.STATUS_APPROVED}
    )

    return [
        {
            "label": "Service Duration",
            "value": get_summary_value(working_time_summary, "service_duration_display") or "Not set",
            "meta": f"Hired {employee.hire_date:%b %d, %Y}" if employee.hire_date else "Hire date not recorded",
            "tone": "neutral",
        },
        {
            "label": "Attendance Reliability",
            "value": f"{attendance_rate}%",
            "meta": f"{present_total} present day(s) across {attendance_total} attendance record(s)",
            "tone": "good" if attendance_rate >= 90 else "warning" if attendance_rate >= 75 else "danger",
        },
        {
            "label": "Compliance Alerts",
            "value": str(compliance_alert_total + overdue_submission_total),
            "meta": f"{compliance_alert_total} ID alert(s), {overdue_submission_total} overdue submission(s)",
            "tone": "good" if (compliance_alert_total + overdue_submission_total) == 0 else "danger",
        },
        {
            "label": "Payroll Status",
            "value": payroll_profile.get_status_display() if payroll_profile else "Setup Needed",
            "meta": (
                f"Latest net pay {payroll_lines[0].net_pay}" if payroll_lines else
                f"Estimated net {payroll_profile.estimated_net_salary}" if payroll_profile else
                "No payroll lines generated yet"
            ),
            "tone": "good" if payroll_profile and payroll_profile.status == payroll_profile.STATUS_ACTIVE else "warning",
        },
        {
            "label": "Leave This Year",
            "value": str(approved_leave_days_year),
            "meta": f"{sum(1 for item in leave_records if item.status == EmployeeLeave.STATUS_PENDING)} pending request(s) now",
            "tone": "neutral",
        },
        {
            "label": "Open Workforce Items",
            "value": str(open_action_total + pending_requests_total),
            "meta": f"{open_action_total} action item(s), {pending_requests_total} document request(s)",
            "tone": "warning" if (open_action_total + pending_requests_total) else "good",
        },
        {
            "label": "Active Deductions",
            "value": str(len(active_obligations)),
            "meta": f"Outstanding balance {outstanding_obligation_balance}",
            "tone": "warning" if active_obligations else "neutral",
        },
        {
            "label": "Available Annual Leave",
            "value": str(get_summary_value(working_time_summary, "annual_leave_available_after_planning_days", 0) or 0),
            "meta": "Balance after taken and future approved leave",
            "tone": "good",
        },
    ]


def build_employee_360_signal_cards(
    attendance_summary,
    working_time_summary,
    identity_document_statuses,
    leave_records,
    payroll_profile,
    payroll_lines,
    payroll_obligations,
):
    attendance_total = attendance_summary.get("attendance_total") or 0
    present_total = attendance_summary.get("present_attendance_count") or 0
    absence_total = attendance_summary.get("absence_attendance_count") or 0
    attendance_rate = int(round((present_total / attendance_total) * 100)) if attendance_total else 0
    compliance_attention = [
        status for status in identity_document_statuses if status["state_key"] in {"expired", "expiring_soon", "missing"}
    ]
    latest_payroll_line = payroll_lines[0] if payroll_lines else None
    active_obligations = [
        obligation for obligation in payroll_obligations if obligation.status == obligation.STATUS_ACTIVE
    ]
    leave_mix = {
        "annual": get_summary_value(working_time_summary, "annual_leave_days", 0) or 0,
        "sick": get_summary_value(working_time_summary, "sick_leave_days", 0) or 0,
        "unpaid": get_summary_value(working_time_summary, "unpaid_leave_days", 0) or 0,
        "emergency": get_summary_value(working_time_summary, "emergency_leave_days", 0) or 0,
        "other": get_summary_value(working_time_summary, "other_leave_days", 0) or 0,
    }
    dominant_leave_key = max(leave_mix, key=leave_mix.get) if any(leave_mix.values()) else ""
    dominant_leave_label_map = {
        "annual": "Annual leave",
        "sick": "Sick leave",
        "unpaid": "Unpaid leave",
        "emergency": "Emergency leave",
        "other": "Other leave",
    }

    return [
        {
            "title": "Attendance Signal",
            "value": f"{attendance_rate}%",
            "description": (
                f"{present_total} present day(s), {absence_total} absence day(s), "
                f"{attendance_summary.get('total_late_minutes') or 0} late minute(s)."
            ),
            "tone": "good" if attendance_rate >= 90 else "warning" if attendance_rate >= 75 else "danger",
        },
        {
            "title": "Leave Trend",
            "value": str(get_summary_value(working_time_summary, "approved_leave_days", 0) or 0),
            "description": (
                f"Approved leave days total. Strongest pattern: "
                f"{dominant_leave_label_map.get(dominant_leave_key, 'No dominant leave pattern yet')}."
            ),
            "tone": "neutral",
        },
        {
            "title": "Compliance Readiness",
            "value": str(len(compliance_attention)),
            "description": (
                "Passport, Civil ID, and requested submissions are under control."
                if not compliance_attention
                else f"{len(compliance_attention)} identity/compliance alert(s) need follow-up."
            ),
            "tone": "good" if not compliance_attention else "danger",
        },
        {
            "title": "Payroll Stability",
            "value": payroll_profile.get_status_display() if payroll_profile else "Pending",
            "description": (
                f"Latest net pay {latest_payroll_line.net_pay}. {len(active_obligations)} active obligation(s)."
                if latest_payroll_line
                else "Payroll profile is visible, but no payroll line has been generated yet."
                if payroll_profile
                else "Payroll profile setup has not been completed yet."
            ),
            "tone": "good" if payroll_profile and payroll_profile.status == payroll_profile.STATUS_ACTIVE else "warning",
        },
    ]


def build_employee_leave_trend_rows(leave_records):
    leave_type_totals = {
        EmployeeLeave.LEAVE_TYPE_ANNUAL: 0,
        EmployeeLeave.LEAVE_TYPE_SICK: 0,
        EmployeeLeave.LEAVE_TYPE_UNPAID: 0,
        EmployeeLeave.LEAVE_TYPE_EMERGENCY: 0,
        EmployeeLeave.LEAVE_TYPE_OTHER: 0,
    }
    leave_type_labels = dict(EmployeeLeave.LEAVE_TYPE_CHOICES)

    approved_total = 0
    pending_total = 0
    rejected_total = 0
    cancelled_total = 0

    for leave_record in leave_records:
        if leave_record.status == EmployeeLeave.STATUS_APPROVED:
            approved_total += leave_record.total_days
            leave_type_totals[leave_record.leave_type] = leave_type_totals.get(leave_record.leave_type, 0) + leave_record.total_days
        elif leave_record.status == EmployeeLeave.STATUS_PENDING:
            pending_total += leave_record.total_days
        elif leave_record.status == EmployeeLeave.STATUS_REJECTED:
            rejected_total += 1
        elif leave_record.status == EmployeeLeave.STATUS_CANCELLED:
            cancelled_total += 1

    rows = []
    for leave_type, total_days in leave_type_totals.items():
        rows.append(
            {
                "label": leave_type_labels.get(leave_type, leave_type.title()),
                "approved_days": total_days,
                "share": int(round((total_days / approved_total) * 100)) if approved_total else 0,
            }
        )

    return {
        "rows": rows,
        "approved_total": approved_total,
        "pending_total": pending_total,
        "rejected_total": rejected_total,
        "cancelled_total": cancelled_total,
    }


def build_employee_compliance_timeline(identity_document_statuses, documents, required_submission_requests):
    items = []

    for status in identity_document_statuses:
        items.append(
            {
                "title": status["label"],
                "subtitle": status["status_label"],
                "date": status["expiry_date"] or status["issue_date"],
                "date_label": (
                    f"Expiry {status['expiry_date']:%b %d, %Y}" if status["expiry_date"] else
                    f"Issued {status['issue_date']:%b %d, %Y}" if status["issue_date"] else
                    "No date recorded"
                ),
                "description": (
                    f"Reference {status['reference_number']}. {status['days_remaining_display']}"
                    if status["reference_number"] else status["days_remaining_display"]
                ),
                "tone": "good" if status["state_key"] == "valid" else "warning" if status["state_key"] == "expiring_soon" else "danger",
            }
        )

    for document in documents[:6]:
        items.append(
            {
                "title": document.title or document.filename,
                "subtitle": document.get_document_type_display(),
                "date": timezone.localtime(document.uploaded_at).date() if document.uploaded_at else None,
                "date_label": (
                    f"Uploaded {timezone.localtime(document.uploaded_at):%b %d, %Y}"
                    if document.uploaded_at else "Upload date unavailable"
                ),
                "description": document.description or "Document uploaded to employee file.",
                "tone": "danger" if document.is_expired else "warning" if document.is_expiring_soon else "neutral",
            }
        )

    for request_item in required_submission_requests[:6]:
        items.append(
            {
                "title": request_item.title,
                "subtitle": request_item.get_status_display(),
                "date": (
                    request_item.due_date
                    or (request_item.submitted_at.date() if request_item.submitted_at else timezone.localtime(request_item.created_at).date())
                ),
                "date_label": (
                    f"Due {request_item.due_date:%b %d, %Y}" if request_item.due_date else
                    f"Updated {timezone.localtime(request_item.updated_at):%b %d, %Y}"
                ),
                "description": request_item.instructions or "Compliance or file request raised from management workflow.",
                "tone": "danger" if request_item.is_overdue else "warning" if request_item.status != request_item.STATUS_COMPLETED else "good",
            }
        )

    items.sort(key=lambda item: item["date"] or date.min, reverse=True)
    return items[:10]


def build_employee_360_timeline_items(
    history_entries,
    leave_records,
    action_records,
    documents,
    employee_document_requests,
    required_submission_requests,
    payroll_lines,
):
    items = []

    for entry in history_entries:
        items.append(
            {
                "kind": "History",
                "title": entry.title,
                "date": entry.event_date or timezone.localtime(entry.created_at).date(),
                "date_label": (
                    f"{entry.event_date:%b %d, %Y}" if entry.event_date else f"{timezone.localtime(entry.created_at):%b %d, %Y}"
                ),
                "description": entry.description or "Profile event recorded in employee history.",
                "meta": entry.created_by or "System",
                "tone": "neutral",
            }
        )

    for leave_record in leave_records[:8]:
        items.append(
            {
                "kind": "Leave",
                "title": leave_record.get_leave_type_display(),
                "date": leave_record.start_date,
                "date_label": f"{leave_record.start_date:%b %d, %Y} to {leave_record.end_date:%b %d, %Y}",
                "description": leave_record.reason or f"{leave_record.total_days} day(s), status {leave_record.get_status_display().lower()}.",
                "meta": leave_record.get_status_display(),
                "tone": "good" if leave_record.status == leave_record.STATUS_APPROVED else "warning" if leave_record.status == leave_record.STATUS_PENDING else "danger",
            }
        )

    for action_record in action_records[:8]:
        items.append(
            {
                "kind": "Action",
                "title": action_record.title,
                "date": action_record.action_date,
                "date_label": f"{action_record.action_date:%b %d, %Y}",
                "description": action_record.description or action_record.get_action_type_display(),
                "meta": f"{action_record.get_status_display()} • {action_record.get_severity_display()}",
                "tone": "danger" if action_record.severity == action_record.SEVERITY_CRITICAL else "warning",
            }
        )

    for document in documents[:8]:
        items.append(
            {
                "kind": "Document",
                "title": document.title or document.filename,
                "date": timezone.localtime(document.uploaded_at).date() if document.uploaded_at else None,
                "date_label": f"{timezone.localtime(document.uploaded_at):%b %d, %Y}" if document.uploaded_at else "No upload date",
                "description": document.description or document.get_document_type_display(),
                "meta": document.get_document_type_display(),
                "tone": "danger" if document.is_expired else "warning" if document.is_expiring_soon else "neutral",
            }
        )

    for payroll_line in payroll_lines[:6]:
        period = payroll_line.payroll_period
        items.append(
            {
                "kind": "Payroll",
                "title": period.title,
                "date": period.pay_date or period.period_end or period.period_start,
                "date_label": (
                    f"Pay date {period.pay_date:%b %d, %Y}" if period.pay_date else
                    f"Period {period.period_start:%b %d, %Y} to {period.period_end:%b %d, %Y}"
                ),
                "description": f"Net pay {payroll_line.net_pay} from base {payroll_line.base_salary}.",
                "meta": period.get_status_display(),
                "tone": "good" if period.status == period.STATUS_PAID else "warning",
            }
        )

    for document_request in employee_document_requests[:6]:
        request_date = document_request.submitted_at.date() if document_request.submitted_at else timezone.localtime(document_request.created_at).date()
        items.append(
            {
                "kind": "Request",
                "title": document_request.title,
                "date": request_date,
                "date_label": f"{request_date:%b %d, %Y}",
                "description": document_request.request_note or document_request.get_request_type_display(),
                "meta": document_request.get_status_display(),
                "tone": "good" if document_request.status == document_request.STATUS_COMPLETED else "warning" if document_request.status in {document_request.STATUS_REQUESTED, document_request.STATUS_APPROVED} else "danger",
            }
        )

    for submission_request in required_submission_requests[:6]:
        items.append(
            {
                "kind": "Compliance",
                "title": submission_request.title,
                "date": submission_request.due_date or timezone.localtime(submission_request.created_at).date(),
                "date_label": (
                    f"Due {submission_request.due_date:%b %d, %Y}" if submission_request.due_date else
                    f"Created {timezone.localtime(submission_request.created_at):%b %d, %Y}"
                ),
                "description": submission_request.instructions or submission_request.get_request_type_display(),
                "meta": submission_request.get_status_display(),
                "tone": "danger" if submission_request.is_overdue else "warning" if submission_request.status != submission_request.STATUS_COMPLETED else "good",
            }
        )

    items.sort(key=lambda item: item["date"] or date.min, reverse=True)
    return items[:18]


def build_employee_profile_section_actions(employee):
    transfer_url = reverse("employees:employee_transfer", kwargs={"pk": employee.pk})

    return {
        "employee_information": {
            "label": "Edit section",
            "url": build_employee_detail_tab_url(
                employee,
                tab="overview",
                modal="employee_information",
                anchor="employee-information-section",
            ),
            "title": "Edit employee information",
            "modal_target": "employee-information-modal",
        },
        "identity_information": {
            "label": "Edit section",
            "url": build_employee_detail_tab_url(
                employee,
                tab="compliance",
                modal="identity_information",
                anchor="employee-information-section",
            ),
            "title": "Edit passport and civil ID details",
            "modal_target": "identity-information-modal",
        },
        "payroll_information": {
            "label": "Edit payroll",
            "url": build_employee_detail_tab_url(
                employee,
                tab="payroll",
                modal="payroll_information",
                anchor="employee-payroll-section",
            ),
            "title": "Edit payroll profile and salary settings",
            "modal_target": "payroll-information-modal",
        },
        "organization_information": {
            "label": "Edit section",
            "url": f"{transfer_url}#organization-information-section",
            "title": "Edit organization placement",
        },
    }


class EmployeeDetailView(LoginRequiredMixin, DetailView):
    model = Employee
    template_name = "employees/employee_detail.html"
    context_object_name = "employee"

    def dispatch(self, request, *args, **kwargs):
        employee = self.get_object()
        if not can_view_employee_profile(request.user, employee):
            return deny_employee_access(
                request,
                "You do not have permission to view this employee profile.",
                employee=employee,
            )
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee = self.object
        current_user = self.request.user

        can_view_directory = can_view_employee_directory(current_user)
        can_manage_employees = can_create_or_edit_employees(current_user)
        can_edit_employee = can_create_or_edit_employees(current_user)
        can_delete_employee_flag = can_delete_employee(current_user)
        can_transfer_employee_flag = can_transfer_employee(current_user)
        can_change_status = can_change_employee_status(current_user)
        can_manage_documents = can_manage_employee_documents(current_user, employee)
        can_request_leave_flag = can_request_leave(current_user, employee)
        can_review_leave_flag = can_review_leave(current_user)
        can_manage_action_records = can_create_action_records(current_user)
        can_manage_attendance_records_flag = can_manage_attendance_records(current_user)
        can_add_history = can_add_manual_history(current_user)
        is_self_profile = is_self_employee(current_user, employee)
        is_self_service_view = is_self_profile and not can_view_management_employee_sections(current_user, employee)
        is_supervisor_own_profile_view = bool(
            is_self_profile
            and is_branch_scoped_supervisor(current_user)
            and can_view_management_employee_sections(current_user, employee)
        )
        is_operations_own_profile_view = False
        is_management_own_profile_view = bool(
            is_self_profile and should_use_management_own_profile(current_user, employee)
        )
        is_branch_scoped_supervisor_view = is_branch_scoped_supervisor(current_user) and not is_self_service_view
        is_self_focused_profile_view = bool(
            is_self_service_view or is_supervisor_own_profile_view
        )

        all_documents = list(employee.documents.select_related("linked_leave").all())

        if is_self_focused_profile_view:
            documents = [document for document in all_documents if document.linked_leave_id]
            leave_form = kwargs.get("leave_form") or EmployeeSelfServiceLeaveRequestForm()
        else:
            documents = all_documents
            leave_form = kwargs.get("leave_form") or EmployeeLeaveForm()

        identity_document_statuses = build_identity_document_statuses(employee)

        required_submission_queryset = employee.required_submissions.select_related(
            "created_by",
            "reviewed_by",
            "fulfilled_document",
        ).order_by("-updated_at", "-created_at", "-id")
        required_submission_requests = list(required_submission_queryset)
        required_submission_create_form = kwargs.get("required_submission_create_form") or EmployeeRequiredSubmissionCreateForm()
        required_submission_review_form = kwargs.get("required_submission_review_form") or EmployeeRequiredSubmissionReviewForm()
        required_submission_response_forms = {}
        for submission_request in required_submission_requests:
            if submission_request.can_employee_submit:
                response_form = EmployeeRequiredSubmissionResponseForm(instance=submission_request)
                required_submission_response_forms[submission_request.pk] = response_form
                submission_request.response_form = response_form
        required_submission_total = len(required_submission_requests)
        required_submission_requested_count = sum(
            1
            for submission_request in required_submission_requests
            if submission_request.status == EmployeeRequiredSubmission.STATUS_REQUESTED
        )
        required_submission_submitted_count = sum(
            1
            for submission_request in required_submission_requests
            if submission_request.status == EmployeeRequiredSubmission.STATUS_SUBMITTED
        )
        required_submission_completed_count = sum(
            1
            for submission_request in required_submission_requests
            if submission_request.status == EmployeeRequiredSubmission.STATUS_COMPLETED
        )
        required_submission_needs_correction_count = sum(
            1
            for submission_request in required_submission_requests
            if submission_request.status == EmployeeRequiredSubmission.STATUS_NEEDS_CORRECTION
        )
        required_submission_overdue_count = sum(
            1 for submission_request in required_submission_requests if submission_request.is_overdue
        )

        employee_document_request_queryset = employee.document_requests.select_related(
            "created_by",
            "reviewed_by",
            "delivered_document",
        ).order_by("-updated_at", "-created_at", "-id")
        employee_document_requests = list(employee_document_request_queryset)
        employee_document_request_create_form = kwargs.get("employee_document_request_create_form") or EmployeeDocumentRequestCreateForm()
        employee_document_request_review_form = kwargs.get("employee_document_request_review_form") or EmployeeDocumentRequestReviewForm()
        employee_document_request_total = len(employee_document_requests)
        employee_document_request_requested_count = sum(
            1
            for document_request in employee_document_requests
            if document_request.status == EmployeeDocumentRequest.STATUS_REQUESTED
        )
        employee_document_request_approved_count = sum(
            1
            for document_request in employee_document_requests
            if document_request.status == EmployeeDocumentRequest.STATUS_APPROVED
        )
        employee_document_request_completed_count = sum(
            1
            for document_request in employee_document_requests
            if document_request.status == EmployeeDocumentRequest.STATUS_COMPLETED
        )
        employee_document_request_rejected_count = sum(
            1
            for document_request in employee_document_requests
            if document_request.status == EmployeeDocumentRequest.STATUS_REJECTED
        )
        employee_document_request_cancelled_count = sum(
            1
            for document_request in employee_document_requests
            if document_request.status == EmployeeDocumentRequest.STATUS_CANCELLED
        )

        document_form = kwargs.get("document_form") or EmployeeDocumentForm()
        document_total = len(documents)
        required_document_count = sum(1 for document in documents if document.is_required)
        expired_document_count = sum(1 for document in documents if document.is_expired)
        expiring_soon_count = sum(1 for document in documents if document.is_expiring_soon)

        leave_records = list(employee.leave_records.all())
        for leave_record in leave_records:
            leave_record.workflow_owner_label = get_leave_current_stage_owner_label(leave_record)
        leave_total = len(leave_records)
        pending_leave_count = sum(
            1 for leave_record in leave_records if leave_record.status == EmployeeLeave.STATUS_PENDING
        )
        approved_leave_count = sum(
            1 for leave_record in leave_records if leave_record.status == EmployeeLeave.STATUS_APPROVED
        )

        action_records = list(
            employee.action_records.all().order_by("-action_date", "-id")
        )
        recent_action_records = action_records[:8]
        action_form = kwargs.get("action_form") or EmployeeActionRecordForm()
        action_record_total = len(action_records)
        open_action_record_count = sum(
            1 for action_record in action_records if action_record.status == EmployeeActionRecord.STATUS_OPEN
        )
        resolved_action_record_count = sum(
            1 for action_record in action_records if action_record.status == EmployeeActionRecord.STATUS_RESOLVED
        )
        critical_action_record_count = sum(
            1 for action_record in action_records if action_record.severity == EmployeeActionRecord.SEVERITY_CRITICAL
        )

        filter_state = build_attendance_filter_state(self.request)
        attendance_queryset = employee.attendance_ledgers.all()

        if filter_state["start_date"]:
            attendance_queryset = attendance_queryset.filter(attendance_date__gte=filter_state["start_date"])
        if filter_state["end_date"]:
            attendance_queryset = attendance_queryset.filter(attendance_date__lte=filter_state["end_date"])

        attendance_ledgers = list(attendance_queryset)
        attendance_form = kwargs.get("attendance_form") or EmployeeAttendanceLedgerForm(employee=employee)
        attendance_summary = build_attendance_summary(attendance_ledgers)

        sick_leave_day_count = sum(
            1
            for attendance_entry in attendance_ledgers
            if attendance_entry.day_status == EmployeeAttendanceLedger.DAY_STATUS_SICK_LEAVE
        )
        leave_day_count = sum(
            1
            for attendance_entry in attendance_ledgers
            if attendance_entry.day_status in {
                EmployeeAttendanceLedger.DAY_STATUS_PAID_LEAVE,
                EmployeeAttendanceLedger.DAY_STATUS_UNPAID_LEAVE,
            }
        )
        off_day_count = (
            attendance_summary["weekly_off_attendance_count"]
            + attendance_summary["holiday_attendance_count"]
        )
        worked_day_count = attendance_summary["present_attendance_count"]
        absence_day_count = attendance_summary["absence_attendance_count"]

        history_queryset = employee.history_entries.all()
        total_history_count = history_queryset.count()
        history_paginator = Paginator(history_queryset, 5)
        history_page_obj = history_paginator.get_page(self.request.GET.get("timeline_page"))
        history_entries = list(history_page_obj.object_list)
        visible_history_count = len(history_entries)
        has_more_history = history_paginator.num_pages > 1
        timeline_query = self.request.GET.copy()
        timeline_query.pop("timeline_page", None)
        timeline_base_query = timeline_query.urlencode()
        timeline_base_path = self.request.path
        if timeline_base_query:
            timeline_base_path = f"{timeline_base_path}?{timeline_base_query}"
        timeline_base_url = f"{timeline_base_path}#employee-timeline-section"
        timeline_first_url = ""
        timeline_previous_url = ""
        timeline_next_url = ""
        if history_page_obj.number > 1:
            first_query = self.request.GET.copy()
            first_query["timeline_page"] = 1
            timeline_first_url = f"{self.request.path}?{first_query.urlencode()}#employee-timeline-section"
        if history_page_obj.has_previous():
            previous_query = self.request.GET.copy()
            previous_query["timeline_page"] = history_page_obj.previous_page_number()
            timeline_previous_url = f"{self.request.path}?{previous_query.urlencode()}#employee-timeline-section"
        if history_page_obj.has_next():
            next_query = self.request.GET.copy()
            next_query["timeline_page"] = history_page_obj.next_page_number()
            timeline_next_url = f"{self.request.path}?{next_query.urlencode()}#employee-timeline-section"
        history_form = kwargs.get("history_form") or EmployeeHistoryForm()

        working_time_summary = build_employee_working_time_summary(employee)
        can_view_working_time_summary = (
            is_management_user(current_user)
            or can_supervisor_view_employee(current_user, employee)
            or is_self_employee(current_user, employee)
        )

        context["document_form"] = document_form
        context["documents"] = documents
        context["document_total"] = document_total
        context["required_document_count"] = required_document_count
        context["expired_document_count"] = expired_document_count
        context["expiring_soon_count"] = expiring_soon_count
        context["identity_document_statuses"] = identity_document_statuses

        context["required_submission_create_form"] = required_submission_create_form
        context["required_submission_review_form"] = required_submission_review_form
        context["required_submission_response_forms"] = required_submission_response_forms
        context["required_submission_requests"] = required_submission_requests
        context["required_submission_total"] = required_submission_total
        context["required_submission_requested_count"] = required_submission_requested_count
        context["required_submission_submitted_count"] = required_submission_submitted_count
        context["required_submission_completed_count"] = required_submission_completed_count
        context["required_submission_needs_correction_count"] = required_submission_needs_correction_count
        context["required_submission_overdue_count"] = required_submission_overdue_count

        context["employee_document_request_create_form"] = employee_document_request_create_form
        context["employee_document_request_review_form"] = employee_document_request_review_form
        context["employee_document_requests"] = employee_document_requests
        context["employee_document_request_total"] = employee_document_request_total
        context["employee_document_request_requested_count"] = employee_document_request_requested_count
        context["employee_document_request_approved_count"] = employee_document_request_approved_count
        context["employee_document_request_completed_count"] = employee_document_request_completed_count
        context["employee_document_request_rejected_count"] = employee_document_request_rejected_count
        context["employee_document_request_cancelled_count"] = employee_document_request_cancelled_count

        context["leave_form"] = leave_form
        context["leave_records"] = leave_records
        context["leave_total"] = leave_total
        context["pending_leave_count"] = pending_leave_count
        context["approved_leave_count"] = approved_leave_count

        context["action_form"] = action_form
        context["action_records"] = action_records
        context["recent_action_records"] = recent_action_records
        context["action_record_total"] = action_record_total
        context["open_action_record_count"] = open_action_record_count
        context["resolved_action_record_count"] = resolved_action_record_count
        context["critical_action_record_count"] = critical_action_record_count

        context["attendance_ledgers"] = attendance_ledgers
        context["attendance_form"] = attendance_form
        context["attendance_filter_form"] = filter_state["form"]
        context["attendance_filter_type"] = filter_state["filter_type"]
        context["attendance_filter_start_date"] = filter_state["start_date"]
        context["attendance_filter_end_date"] = filter_state["end_date"]
        context["attendance_period_label"] = filter_state["period_label"]
        context["attendance_filter_applied"] = filter_state["is_applied"]

        context.update(attendance_summary)

        context["history_form"] = history_form
        context["history_entries"] = history_entries
        context["total_history_count"] = total_history_count
        context["visible_history_count"] = visible_history_count
        context["has_more_history"] = has_more_history
        context["history_page_obj"] = history_page_obj
        context["history_paginator"] = history_paginator
        context["timeline_base_url"] = timeline_base_url
        context["timeline_first_url"] = timeline_first_url
        context["timeline_previous_url"] = timeline_previous_url
        context["timeline_next_url"] = timeline_next_url

        context["working_time_summary"] = working_time_summary
        context["can_view_working_time_summary"] = can_view_working_time_summary
        context["worked_day_count"] = worked_day_count
        context["off_day_count"] = off_day_count
        context["leave_day_count"] = leave_day_count
        context["sick_leave_day_count"] = sick_leave_day_count
        context["absence_day_count"] = absence_day_count

        context["same_company_url"] = (
            f"{reverse_lazy('employees:employee_list')}?company={employee.company_id}"
            if employee.company_id and can_view_directory
            else None
        )
        context["same_department_url"] = (
            f"{reverse_lazy('employees:employee_list')}?department={employee.department_id}"
            if employee.department_id and can_view_directory
            else None
        )
        context["same_branch_url"] = (
            f"{reverse_lazy('employees:employee_list')}?branch={employee.branch_id}"
            if employee.branch_id and can_view_directory
            else None
        )
        context["same_section_url"] = (
            f"{reverse_lazy('employees:employee_list')}?section={employee.section_id}"
            if employee.section_id and can_view_directory
            else None
        )
        context["same_job_title_url"] = (
            f"{reverse_lazy('employees:employee_list')}?job_title={employee.job_title_id}"
            if employee.job_title_id and can_view_directory
            else None
        )
        context["similar_name_url"] = (
            f"{reverse_lazy('employees:employee_list')}?search={employee.full_name}"
            if employee.full_name and can_view_directory
            else None
        )

        context["can_view_directory"] = can_view_directory
        context["can_manage_employees"] = can_manage_employees
        context["can_edit_employee"] = can_edit_employee
        context["can_delete_employee"] = can_delete_employee_flag
        context["can_transfer_employee"] = can_transfer_employee_flag
        context["can_change_status"] = can_change_status
        context["employee_status_choices"] = Employee.EMPLOYMENT_STATUS_CHOICES
        context["can_manage_documents"] = can_manage_documents
        context["can_manage_employee_required_submissions"] = can_manage_employee_required_submissions(current_user, employee) and not is_self_profile
        context["can_use_profile_section_edit"] = can_edit_employee
        context["profile_section_actions"] = build_employee_profile_section_actions(employee) if can_edit_employee else {}
        context["employee_information_modal_form"] = kwargs.get("employee_information_modal_form") or EmployeeInformationModalForm(instance=employee)
        context["identity_information_modal_form"] = kwargs.get("identity_information_modal_form") or EmployeeIdentityModalForm(instance=employee)
        context["active_profile_modal"] = kwargs.get("active_profile_modal") or (self.request.GET.get("modal") or "").strip()
        context["can_request_leave"] = can_request_leave_flag
        context["can_review_leave"] = can_review_leave_flag
        context["can_manage_action_records"] = can_manage_action_records
        context["can_manage_attendance_records"] = can_manage_attendance_records_flag
        context["can_add_history"] = can_add_history
        context["is_self_service_view"] = is_self_service_view
        context["is_supervisor_own_profile_view"] = is_supervisor_own_profile_view
        context["is_operations_own_profile_view"] = is_operations_own_profile_view
        context["is_management_own_profile_view"] = is_management_own_profile_view
        context["is_self_focused_profile_view"] = is_self_focused_profile_view
        context["can_cancel_leave"] = True

        supervisor_employee = get_employee_supervisor(employee)
        team_leader_employee = get_employee_team_leader(employee)
        department_manager_display = get_department_manager_display(employee)
        branch_supervisor_display = get_branch_supervisor_display(employee)
        team_leader_display = get_team_leader_display(employee)
        branch_team_context = build_branch_team_structure(employee)

        context["employee_display_company"] = get_short_structure_label(employee.company)
        context["employee_display_department"] = get_short_structure_label(employee.department)
        context["employee_display_branch"] = get_short_structure_label(employee.branch)
        context["employee_display_section"] = get_short_structure_label(employee.section)
        context["employee_display_job_title"] = get_short_structure_label(employee.job_title)

        for group in branch_team_context["branch_team_groups"]:
            for member in group["members"]:
                member.short_job_title_display = get_short_structure_label(member.job_title)
                member.short_section_display = get_short_structure_label(member.section)
                member.short_branch_display = get_short_structure_label(member.branch)
                member.short_department_display = get_short_structure_label(member.department)
                member.short_company_display = get_short_structure_label(member.company)

        context["self_service_supervisor"] = supervisor_employee
        context["self_service_team_leader"] = team_leader_employee
        context["department_manager_display"] = department_manager_display
        context["branch_supervisor_display"] = branch_supervisor_display
        context["team_leader_display"] = team_leader_display
        context["is_branch_scoped_supervisor_view"] = is_branch_scoped_supervisor_view
        context["scoped_branch"] = get_user_scope_branch(current_user)
        context["branch_team_members"] = branch_team_context["branch_team_members"]
        context["branch_team_groups"] = branch_team_context["branch_team_groups"]
        context["branch_team_total"] = branch_team_context["branch_team_total"]
        context["self_service_request_records"] = leave_records
        context["self_service_pending_leave_count"] = pending_leave_count
        context["self_service_approved_leave_count"] = approved_leave_count
        context["self_service_rejected_leave_count"] = sum(
            1 for leave_record in leave_records if leave_record.status == EmployeeLeave.STATUS_REJECTED
        )
        context["self_service_cancelled_leave_count"] = sum(
            1 for leave_record in leave_records if leave_record.status == EmployeeLeave.STATUS_CANCELLED
        )
        context["self_service_document_request_records"] = employee_document_requests if is_self_focused_profile_view else []
        context["self_service_document_request_requested_count"] = employee_document_request_requested_count
        context["self_service_document_request_approved_count"] = employee_document_request_approved_count
        context["self_service_document_request_completed_count"] = employee_document_request_completed_count
        context["self_service_document_request_rejected_count"] = employee_document_request_rejected_count
        context["self_service_document_request_cancelled_count"] = employee_document_request_cancelled_count
        context["can_create_employee_document_request"] = can_create_employee_document_request(current_user, employee)
        context["can_cancel_employee_document_request"] = True

        PayrollProfile = apps.get_model("payroll", "PayrollProfile")
        PayrollLine = apps.get_model("payroll", "PayrollLine")
        PayrollObligation = apps.get_model("payroll", "PayrollObligation")
        payroll_profile = PayrollProfile.objects.select_related("company").filter(employee=employee).first()
        latest_payroll_lines = list(
            PayrollLine.objects.select_related("payroll_period")
            .filter(employee=employee)
            .order_by("-payroll_period__period_start", "-id")[:5]
        )
        payroll_obligations = list(
            PayrollObligation.objects.filter(employee=employee).order_by("-created_at", "-id")[:8]
        )
        employee_360_overview_cards = build_employee_360_overview_cards(
            employee,
            attendance_summary,
            working_time_summary,
            identity_document_statuses,
            leave_records,
            required_submission_requests,
            employee_document_requests,
            payroll_profile,
            latest_payroll_lines,
            payroll_obligations,
            action_records,
        )
        employee_360_signal_cards = build_employee_360_signal_cards(
            attendance_summary,
            working_time_summary,
            identity_document_statuses,
            leave_records,
            payroll_profile,
            latest_payroll_lines,
            payroll_obligations,
        )
        employee_leave_trends = build_employee_leave_trend_rows(leave_records)
        employee_compliance_timeline = build_employee_compliance_timeline(
            identity_document_statuses,
            all_documents,
            required_submission_requests,
        )
        employee_360_timeline_items = build_employee_360_timeline_items(
            history_entries,
            leave_records,
            action_records,
            all_documents,
            employee_document_requests,
            required_submission_requests,
            latest_payroll_lines,
        )
        context["employee_payroll_profile"] = payroll_profile
        context["employee_payroll_lines"] = latest_payroll_lines
        context["employee_payroll_line_count"] = len(latest_payroll_lines)
        context["employee_payroll_obligations"] = payroll_obligations
        context["employee_estimated_net_salary"] = payroll_profile.estimated_net_salary if payroll_profile else None
        context["employee_360_overview_url"] = build_employee_detail_tab_url(employee, tab="overview")
        context["employee_360_payroll_url"] = build_employee_detail_tab_url(
            employee,
            tab="payroll",
            anchor="employee-payroll-section",
        )
        employee_payroll_workspace_anchor = "payroll-profiles-section" if payroll_profile else "employees-missing-payroll-section"
        context["employee_payroll_workspace_url"] = (
            f"{reverse('payroll:home')}?employee={employee.pk}#{employee_payroll_workspace_anchor}"
        )
        context["employee_360_documents_url"] = build_employee_detail_tab_url(employee, tab="documents")
        context["employee_360_leave_url"] = build_employee_detail_tab_url(employee, tab="leave")
        context["employee_360_compliance_url"] = build_employee_detail_tab_url(employee, tab="compliance")
        context["employee_360_performance_url"] = build_employee_detail_tab_url(
            employee,
            tab="performance",
            anchor="employee-timeline-section",
        )
        context["employee_360_action_center_url"] = (
            f"{reverse('employees:employee_admin_action_center')}?employee={employee.pk}"
        )
        context["employee_360_attendance_management_url"] = (
            f"{reverse('employees:attendance_management')}?employee={employee.pk}"
        )
        context["employee_360_overview_cards"] = employee_360_overview_cards
        context["employee_360_signal_cards"] = employee_360_signal_cards
        context["employee_leave_trends"] = employee_leave_trends
        context["employee_compliance_timeline"] = employee_compliance_timeline
        context["employee_360_timeline_items"] = employee_360_timeline_items
        context["employee_payroll_profile_form"] = kwargs.get("employee_payroll_profile_form") or PayrollProfileForm(
            instance=payroll_profile,
            employee=employee,
        )

        return context


@login_required
def self_service_profile_page(request):
    employee = get_user_employee_profile(request.user)
    if employee is None:
        raise PermissionDenied("No employee profile is connected to this account.")
    if not can_view_employee_profile(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to view this employee profile.",
            employee=employee,
        )

    if should_use_management_own_profile(request.user, employee):
        return redirect(get_workspace_profile_url(request.user, employee))

    context = build_self_service_page_context(
        request,
        employee,
        current_section="profile",
    )
    return render(request, "employees/self_service_profile.html", context)


@login_required
def self_service_leave_page(request):
    employee = get_user_employee_profile(request.user)
    if employee is None:
        raise PermissionDenied("No employee profile is connected to this account.")
    if not can_view_employee_profile(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to view this employee profile.",
            employee=employee,
        )

    context = build_self_service_page_context(
        request,
        employee,
        current_section="leave",
    )
    return render(request, "employees/self_service_leave.html", context)


@login_required
def self_service_documents_page(request):
    employee = get_user_employee_profile(request.user)
    if employee is None:
        raise PermissionDenied("No employee profile is connected to this account.")
    if not can_view_employee_profile(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to view this employee profile.",
            employee=employee,
        )

    context = build_self_service_page_context(
        request,
        employee,
        current_section="documents",
    )
    return render(request, "employees/self_service_documents.html", context)


@login_required
def self_service_working_time_page(request):
    employee = get_user_employee_profile(request.user)
    if employee is None:
        raise PermissionDenied("No employee profile is connected to this account.")
    if not can_view_employee_profile(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to view this employee profile.",
            employee=employee,
        )

    context = build_self_service_page_context(
        request,
        employee,
        current_section="working_time",
    )
    return render(request, "employees/self_service_working_time.html", context)


@login_required
def self_service_branch_page(request):
    employee = get_user_employee_profile(request.user)
    if employee is None:
        raise PermissionDenied("No employee profile is connected to this account.")
    if not can_view_employee_profile(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to view this employee profile.",
            employee=employee,
        )
    if not can_view_branch_self_service(employee):
        messages.error(request, "This employee is not linked to any branch yet.")
        return redirect("employees:self_service_profile")

    week_value = (request.POST.get("week") or request.GET.get("week") or "").strip()
    selected_week_start = get_schedule_week_start(timezone.localdate())
    if week_value:
        try:
            selected_week_start = get_schedule_week_start(date.fromisoformat(week_value))
        except ValueError:
            messages.warning(request, "Invalid week selected. Showing the current branch week instead.")

    context = build_self_service_page_context(
        request,
        employee,
        current_section="branch",
    )
    context.update(build_branch_team_structure(employee))
    context.update(build_branch_weekly_schedule_summary(employee.branch, selected_week_start))
    context.update(
        build_branch_workspace_context(
            employee.branch,
            request.user,
            employee=employee,
            week_start=selected_week_start,
        )
    )
    context.update(build_employee_schedule_snapshot(employee))
    context["branch"] = employee.branch
    context["selected_week_start"] = selected_week_start
    context["can_manage_branch_weekly_schedule"] = can_manage_branch_weekly_schedule(request.user, employee.branch)
    context["branch_post_form"] = BranchPostForm(
        branch=employee.branch,
        can_manage=context["can_manage_branch_workspace"],
    )
    context["branch_workspace_detail_url"] = reverse(
        "operations:branch_workspace_detail",
        kwargs={"branch_id": employee.branch_id},
    )
    context["branch_workspace_schedule_url"] = reverse("employees:self_service_weekly_schedule")
    return render(request, "employees/self_service_branch.html", context)


def get_branch_standard_duty_option_seed_data():
    return [
        {"label": "9 am to 5 pm", "duty_type": BranchWeeklyScheduleEntry.DUTY_TYPE_SHIFT, "start": "09:00", "end": "17:00", "bg": "#ef4444", "text": "#f8fafc", "order": 1},
        {"label": "2 pm to 10 pm", "duty_type": BranchWeeklyScheduleEntry.DUTY_TYPE_SHIFT, "start": "14:00", "end": "22:00", "bg": "#2563eb", "text": "#f8fafc", "order": 2},
        {"label": "3 pm to 11 pm", "duty_type": BranchWeeklyScheduleEntry.DUTY_TYPE_SHIFT, "start": "15:00", "end": "23:00", "bg": "#7c3aed", "text": "#f8fafc", "order": 3},
        {"label": "4 pm to 12 am", "duty_type": BranchWeeklyScheduleEntry.DUTY_TYPE_SHIFT, "start": "16:00", "end": "23:59", "bg": "#8b5cf6", "text": "#f8fafc", "order": 4},
        {"label": "1 pm to 9 pm", "duty_type": BranchWeeklyScheduleEntry.DUTY_TYPE_SHIFT, "start": "13:00", "end": "21:00", "bg": "#0ea5e9", "text": "#f8fafc", "order": 5},
        {"label": "12 pm to 8 pm", "duty_type": BranchWeeklyScheduleEntry.DUTY_TYPE_SHIFT, "start": "12:00", "end": "20:00", "bg": "#3b82f6", "text": "#f8fafc", "order": 6},
        {"label": "off", "duty_type": BranchWeeklyScheduleEntry.DUTY_TYPE_OFF, "start": None, "end": None, "bg": "#facc15", "text": "#111827", "order": 7},
        {"label": "extra off", "duty_type": BranchWeeklyScheduleEntry.DUTY_TYPE_EXTRA_OFF, "start": None, "end": None, "bg": "#eab308", "text": "#111827", "order": 8},
        {"label": "sick leave", "duty_type": BranchWeeklyScheduleEntry.DUTY_TYPE_CUSTOM, "start": None, "end": None, "bg": "#22c55e", "text": "#052e16", "order": 9},
        {"label": "emergency leave", "duty_type": BranchWeeklyScheduleEntry.DUTY_TYPE_CUSTOM, "start": None, "end": None, "bg": "#f97316", "text": "#fff7ed", "order": 10},
        {"label": "vacation", "duty_type": BranchWeeklyScheduleEntry.DUTY_TYPE_CUSTOM, "start": None, "end": None, "bg": "#14b8a6", "text": "#f0fdfa", "order": 11},
        {"label": "support", "duty_type": BranchWeeklyScheduleEntry.DUTY_TYPE_CUSTOM, "start": None, "end": None, "bg": "#64748b", "text": "#f8fafc", "order": 12},
        {"label": "Morning shift", "duty_type": BranchWeeklyScheduleEntry.DUTY_TYPE_SHIFT, "start": "09:00", "end": "17:00", "bg": "#f59e0b", "text": "#111827", "order": 13},
        {"label": "Middle shift", "duty_type": BranchWeeklyScheduleEntry.DUTY_TYPE_SHIFT, "start": "13:00", "end": "21:00", "bg": "#06b6d4", "text": "#083344", "order": 14},
        {"label": "Evening shift", "duty_type": BranchWeeklyScheduleEntry.DUTY_TYPE_SHIFT, "start": "15:00", "end": "23:00", "bg": "#8b5cf6", "text": "#f8fafc", "order": 15},
    ]


def _parse_seed_time(value):
    if not value:
        return None
    return datetime.strptime(value, "%H:%M").time()


def seed_branch_standard_duty_options(branch):
    created = 0
    updated = 0
    for row in get_branch_standard_duty_option_seed_data():
        option, was_created = BranchWeeklyDutyOption.objects.get_or_create(
            branch=branch,
            label=row["label"],
            defaults={
                "duty_type": row["duty_type"],
                "default_start_time": _parse_seed_time(row["start"]),
                "default_end_time": _parse_seed_time(row["end"]),
                "background_color": row["bg"],
                "text_color": row["text"],
                "display_order": row["order"],
                "is_active": True,
            },
        )
        if was_created:
            created += 1
            continue

        changed = False
        # Safe live fix:
        # keep custom colors that managers already changed manually.
        # Seed should refresh structure/order/timing only for existing rows.
        for field_name, value in {
            "duty_type": row["duty_type"],
            "default_start_time": _parse_seed_time(row["start"]),
            "default_end_time": _parse_seed_time(row["end"]),
            "display_order": row["order"],
            "is_active": True,
        }.items():
            if getattr(option, field_name) != value:
                setattr(option, field_name, value)
                changed = True

        if changed:
            option.save()
            sync_schedule_entries_for_duty_option(option)
            updated += 1
    return created, updated


def sync_schedule_entries_for_duty_option(duty_option):
    if not duty_option:
        return
    update_kwargs = {
        "duty_type": duty_option.duty_type,
        "shift_label": duty_option.label,
        "updated_by": "Duty Shift Master",
    }
    if duty_option.duty_type == BranchWeeklyScheduleEntry.DUTY_TYPE_SHIFT:
        update_kwargs["start_time"] = duty_option.default_start_time
        update_kwargs["end_time"] = duty_option.default_end_time
    else:
        update_kwargs["start_time"] = None
        update_kwargs["end_time"] = None
    BranchWeeklyScheduleEntry.objects.filter(duty_option=duty_option).update(**update_kwargs)


def build_manual_schedule_builder_rows(*, team_schedule_rows, week_days):
    rows = []
    for row in team_schedule_rows:
        employee = row.get("employee")
        if not employee:
            continue
        cells = []
        for current_date, cell in zip(week_days, row.get("cells", [])):
            entry = cell.get("entry")
            cells.append(
                {
                    "date": current_date,
                    "field_name": f"manual_duty_{employee.id}_{current_date.isoformat()}",
                    "selected_duty_option_id": str(entry.duty_option_id) if entry and entry.duty_option_id else "",
                    "entry": entry,
                }
            )
        rows.append(
            {
                "employee": employee,
                "pending_off_total": row.get("pending_off_total", 0),
                "pending_off_field_name": f"manual_pending_off_{employee.id}",
                "cells": cells,
            }
        )
    return rows


@login_required
def self_service_weekly_schedule_page(request):
    employee = get_user_employee_profile(request.user)
    if employee is None:
        raise PermissionDenied("No employee profile is connected to this account.")
    if not can_view_employee_profile(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to view this employee profile.",
            employee=employee,
        )
    if not can_view_branch_self_service(employee):
        messages.error(request, "This employee is not linked to any branch yet.")
        return redirect("employees:self_service_profile")

    week_value = (request.GET.get("week") or request.POST.get("week") or "").strip()
    selected_week_start = get_schedule_week_start(timezone.localdate())
    if week_value:
        try:
            selected_week_start = get_schedule_week_start(date.fromisoformat(week_value))
        except ValueError:
            messages.warning(request, "Invalid week selected. Showing the current branch week instead.")

    branch = employee.branch
    can_manage_schedule = can_manage_branch_weekly_schedule(request.user, branch)
    import_form = BranchWeeklyScheduleImportForm()

    if request.method == "POST" and can_manage_schedule:
        action = (request.POST.get("schedule_action") or "").strip()
        redirect_url = f"{reverse('employees:self_service_weekly_schedule')}?week={selected_week_start.isoformat()}"

        if action == "import_schedule":
            import_form = BranchWeeklyScheduleImportForm(request.POST, request.FILES)
            if import_form.is_valid():
                import_result = import_branch_weekly_schedule_file(
                    branch=branch,
                    week_start=selected_week_start,
                    uploaded_file=import_form.cleaned_data["import_file"],
                    actor_label=get_actor_label(request.user),
                    replace_existing=import_form.cleaned_data.get("replace_existing", False),
                )
                if import_result["imported_count"]:
                    mode_label = "replaced" if import_result.get("replace_existing") else "merged into"
                    messages.success(request, f"Imported {import_result['imported_count']} schedule row(s) and {mode_label} the selected branch week.")
                elif import_result.get("replace_existing"):
                    messages.warning(request, "The current week was cleared, but no non-empty duty cells were imported from the file.")
                else:
                    messages.warning(request, "No schedule rows were imported. If you want the uploaded file to fully replace the current sheet, keep 'Replace current week schedule before import' checked.")
                if import_result.get("skipped_empty_cells"):
                    messages.info(request, f"Skipped {import_result['skipped_empty_cells']} empty schedule cell(s). Empty cells only clear old values when replacement mode is enabled.")
                if import_result["errors"]:
                    messages.warning(request, "Some rows were skipped during import: " + " | ".join(import_result["errors"][:5]))
                return redirect(redirect_url)
            messages.error(request, "Please upload a valid .xlsx or .csv file for schedule import.")

        if action == "export_schedule":
            workbook = build_branch_schedule_export_workbook(branch, selected_week_start, include_existing_entries=True)
            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = f'attachment; filename="{branch.name.lower().replace(" ", "-")}-schedule-{selected_week_start.isoformat()}.xlsx"'
            workbook.save(response)
            return response

        if action == "export_schedule_template":
            workbook = build_branch_schedule_export_workbook(branch, selected_week_start, include_existing_entries=False)
            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = f'attachment; filename="{branch.name.lower().replace(" ", "-")}-schedule-template-{selected_week_start.isoformat()}.xlsx"'
            workbook.save(response)
            return response

        if action == "seed_standard_duty_options":
            created_count, updated_count = seed_branch_standard_duty_options(branch)
            messages.success(request, f"Loaded the standard duty list. Created {created_count} and refreshed {updated_count} existing duty option(s).")
            return redirect(redirect_url)

        if action == "create_duty_option":
            duty_option_create_form = BranchWeeklyDutyOptionForm(request.POST)
            if duty_option_create_form.is_valid():
                new_option = duty_option_create_form.save(commit=False)
                new_option.branch = branch
                if not new_option.display_order:
                    new_option.display_order = BranchWeeklyDutyOption.objects.filter(branch=branch).count() + 1
                new_option.save()
                messages.success(request, f"Created duty option '{new_option.label}'.")
                return redirect(redirect_url)
            messages.error(request, "Please review the new duty option details.")

        if action == "update_duty_option_master":
            duty_option = get_object_or_404(BranchWeeklyDutyOption, pk=request.POST.get("duty_option_id"), branch=branch)
            master_form = BranchWeeklyDutyOptionForm(request.POST, instance=duty_option)
            if master_form.is_valid():
                updated_option = master_form.save()
                sync_schedule_entries_for_duty_option(updated_option)
                messages.success(request, f"Updated duty option '{updated_option.label}'.")
                return redirect(redirect_url)
            messages.error(request, f"Please review the duty option '{duty_option.label}'.")

        if action == "delete_duty_option":
            duty_option = get_object_or_404(BranchWeeklyDutyOption, pk=request.POST.get("duty_option_id"), branch=branch)
            linked_entries = BranchWeeklyScheduleEntry.objects.filter(duty_option=duty_option)
            linked_entries.update(
                duty_option=None,
                duty_type=duty_option.duty_type,
                shift_label=duty_option.label,
                start_time=duty_option.default_start_time if duty_option.duty_type == BranchWeeklyScheduleEntry.DUTY_TYPE_SHIFT else None,
                end_time=duty_option.default_end_time if duty_option.duty_type == BranchWeeklyScheduleEntry.DUTY_TYPE_SHIFT else None,
                updated_by=get_actor_label(request.user),
            )
            deleted_label = duty_option.label
            duty_option.delete()
            messages.success(request, f"Deleted duty option '{deleted_label}'. Existing schedule rows kept their copied label and timing.")
            return redirect(redirect_url)

        if action == "reset_duty_option_style":
            duty_option = get_object_or_404(BranchWeeklyDutyOption, pk=request.POST.get("duty_option_id"), branch=branch)
            duty_option.background_color = ""
            duty_option.text_color = ""
            duty_option.save(update_fields=["background_color", "text_color", "updated_at"])
            sync_schedule_entries_for_duty_option(duty_option)
            messages.success(request, f"Reset colors for duty option '{duty_option.label}'.")
            return redirect(redirect_url)

        if action == "update_schedule_theme":
            schedule_theme, _created = BranchWeeklyScheduleTheme.objects.get_or_create(branch=branch)
            schedule_theme_form = BranchWeeklyScheduleThemeForm(request.POST, instance=schedule_theme)
            if schedule_theme_form.is_valid():
                schedule_theme_form.save()
                messages.success(request, "Updated schedule table colors.")
                return redirect(redirect_url)
            messages.error(request, "Please review the schedule table colors.")

        if action == "reset_schedule_theme":
            schedule_theme, _created = BranchWeeklyScheduleTheme.objects.get_or_create(branch=branch)
            schedule_theme.employee_column_bg = "#101828"
            schedule_theme.employee_column_text = "#f8fafc"
            schedule_theme.job_title_column_bg = "#111827"
            schedule_theme.job_title_column_text = "#f8fafc"
            schedule_theme.pending_off_column_bg = "#172033"
            schedule_theme.pending_off_column_text = "#f8fafc"
            schedule_theme.day_header_bg = "#1d293d"
            schedule_theme.day_header_text = "#f8fafc"
            schedule_theme.save()
            messages.success(request, "Reset the schedule table theme.")
            return redirect(redirect_url)

        if action == "save_manual_schedule_builder":
            week_days = build_schedule_week_days(selected_week_start)
            active_options = {
                str(option.id): option
                for option in BranchWeeklyDutyOption.objects.filter(branch=branch, is_active=True)
            }
            branch_employees = list(Employee.objects.filter(branch=branch, is_active=True).order_by("full_name", "employee_id"))
            existing_entries = {
                (entry.employee_id, entry.schedule_date): entry
                for entry in BranchWeeklyScheduleEntry.objects.filter(branch=branch, week_start=selected_week_start)
            }
            saved_count = 0
            cleared_count = 0
            pending_updates = 0
            invalid_pending = []

            for member in branch_employees:
                pending_field = f"manual_pending_off_{member.id}"
                pending_raw = (request.POST.get(pending_field) or "").strip()
                if pending_raw == "":
                    BranchWeeklyPendingOff.objects.filter(branch=branch, employee=member, week_start=selected_week_start).delete()
                else:
                    try:
                        pending_value = int(pending_raw)
                        if pending_value < 0:
                            raise ValueError
                        BranchWeeklyPendingOff.objects.update_or_create(
                            branch=branch,
                            employee=member,
                            week_start=selected_week_start,
                            defaults={
                                "pending_off_count": pending_value,
                                "created_by": get_actor_label(request.user),
                                "updated_by": get_actor_label(request.user),
                            },
                        )
                        pending_updates += 1
                    except ValueError:
                        invalid_pending.append(member.full_name)

                for current_date in week_days:
                    field_name = f"manual_duty_{member.id}_{current_date.isoformat()}"
                    selected_option_id = (request.POST.get(field_name) or "").strip()
                    existing_entry = existing_entries.get((member.id, current_date))
                    if not selected_option_id:
                        if existing_entry:
                            existing_entry.delete()
                            cleared_count += 1
                        continue
                    duty_option = active_options.get(selected_option_id)
                    if duty_option is None:
                        continue
                    defaults = {
                        "week_start": selected_week_start,
                        "duty_option": duty_option,
                        "title": existing_entry.title if existing_entry else "",
                        "order_note": existing_entry.order_note if existing_entry else "",
                        "status": existing_entry.status if existing_entry else BranchWeeklyScheduleEntry.STATUS_PLANNED,
                        "created_by": existing_entry.created_by if existing_entry and existing_entry.created_by else get_actor_label(request.user),
                        "updated_by": get_actor_label(request.user),
                    }
                    BranchWeeklyScheduleEntry.objects.update_or_create(
                        branch=branch,
                        employee=member,
                        schedule_date=current_date,
                        defaults=defaults,
                    )
                    saved_count += 1
            if invalid_pending:
                messages.warning(request, "Some pending off values were ignored because they were invalid numbers: " + ", ".join(invalid_pending[:5]))
            messages.success(request, f"Saved manual schedule builder changes. Updated {saved_count} cell(s), cleared {cleared_count} cell(s), and refreshed {pending_updates} pending-off value(s).")
            return redirect(redirect_url)

        if action == "update_employee_order":
            active_employee_ids = [employee_id for employee_id in request.POST.getlist("ordered_employee_ids") if employee_id and employee_id.isdigit()]
            seen_ids = set()
            ordered_ids = []
            for employee_id in active_employee_ids:
                if employee_id not in seen_ids:
                    ordered_ids.append(int(employee_id))
                    seen_ids.add(employee_id)

            branch_employees = {member.id: member for member in Employee.objects.filter(branch=branch, is_active=True)}
            fallback_ids = [member_id for member_id in branch_employees.keys() if member_id not in ordered_ids]
            final_order_ids = ordered_ids + sorted(
                fallback_ids,
                key=lambda member_id: (branch_employees[member_id].full_name.lower(), branch_employees[member_id].employee_id.lower()),
            )

            for index, employee_id in enumerate(final_order_ids, start=1):
                BranchScheduleGridRow.objects.update_or_create(branch=branch, row_index=index, defaults={"employee_id": employee_id})
            messages.success(request, "Updated employee row order for the schedule table.")
            return redirect(redirect_url)

    previous_week_start = selected_week_start - timedelta(days=7)
    next_week_start = selected_week_start + timedelta(days=7)
    selected_week_end = selected_week_start + timedelta(days=6)
    from workcalendar.services import get_holidays_for_range
    selected_week_holidays = get_holidays_for_range(selected_week_start, selected_week_end)

    context = build_self_service_page_context(request, employee, current_section="weekly_schedule")
    context.update(build_branch_weekly_schedule_summary(branch, selected_week_start))
    context["branch"] = branch
    context["selected_week_start"] = selected_week_start
    context["selected_week_holidays"] = selected_week_holidays
    context["previous_week_start"] = previous_week_start
    context["next_week_start"] = next_week_start
    context["today"] = timezone.localdate()
    context["can_manage_branch_weekly_schedule"] = can_manage_schedule
    context.update(build_employee_schedule_snapshot(employee))
    context["schedule_import_form"] = import_form
    schedule_theme, _created = BranchWeeklyScheduleTheme.objects.get_or_create(branch=branch)
    context["schedule_theme"] = schedule_theme
    context["schedule_theme_form"] = BranchWeeklyScheduleThemeForm(instance=schedule_theme)
    duty_options_qs = BranchWeeklyDutyOption.objects.filter(branch=branch).order_by("display_order", "label", "id")
    context["duty_option_create_form"] = BranchWeeklyDutyOptionForm()
    context["manual_duty_options"] = list(duty_options_qs.filter(is_active=True))
    context["duty_option_style_forms"] = [
        {
            "option": duty_option,
            "master_form": BranchWeeklyDutyOptionForm(instance=duty_option),
            "style_form": BranchWeeklyDutyOptionStyleForm(instance=duty_option),
            "timing_form": BranchWeeklyDutyOptionTimingForm(instance=duty_option),
        }
        for duty_option in duty_options_qs
    ]
    context["manual_schedule_rows"] = build_manual_schedule_builder_rows(
        team_schedule_rows=context.get("team_schedule_rows", []),
        week_days=context.get("week_days", []),
    )
    context["employee_order_rows"] = [
        {"employee": row["employee"], "position": forloop_index}
        for forloop_index, row in enumerate(context.get("team_schedule_rows", []), start=1)
        if row.get("employee")
    ]
    return render(request, "employees/self_service_weekly_schedule.html", context)


@login_required
def self_service_my_schedule_page(request):
    employee = get_user_employee_profile(request.user)
    if employee is None:
        raise PermissionDenied("No employee profile is connected to this account.")
    if not can_view_employee_profile(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to view this employee profile.",
            employee=employee,
        )
    if not can_view_branch_self_service(employee):
        messages.error(request, "This employee is not linked to any branch yet.")
        return redirect("employees:self_service_profile")

    context = build_self_service_page_context(
        request,
        employee,
        current_section="my_schedule",
    )
    from workcalendar.services import get_holidays_for_range
    this_week_start = get_schedule_week_start(timezone.localdate())
    next_week_start = this_week_start + timedelta(days=7)
    context["my_schedule_this_week_holidays"] = get_holidays_for_range(this_week_start, this_week_start + timedelta(days=6))
    context["my_schedule_next_week_holidays"] = get_holidays_for_range(next_week_start, next_week_start + timedelta(days=6))
    context.update(build_employee_schedule_snapshot(employee))
    context["branch"] = employee.branch
    return render(request, "employees/self_service_my_schedule.html", context)


@login_required
def self_service_attendance_page(request):
    employee = get_user_employee_profile(request.user)
    if employee is None:
        raise PermissionDenied("No employee profile is connected to this account.")
    if not can_view_employee_profile(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to view this employee profile.",
            employee=employee,
        )

    today = timezone.localdate()
    branch = getattr(employee, "branch", None)
    schedule_snapshot = build_employee_schedule_snapshot(employee, reference_date=today)
    attendance_today_schedule_entry = schedule_snapshot.get("my_schedule_today_entry")
    attendance_today_schedule_label = schedule_snapshot.get("my_schedule_today_label") or "No branch duty assigned"
    attendance_today_schedule_time = (
        attendance_today_schedule_entry.formatted_time_range
        if attendance_today_schedule_entry
        else ""
    )
    attendance_blocked_for_today = bool(
        attendance_today_schedule_entry
        and attendance_today_schedule_entry.duty_type
        in {
            BranchWeeklyScheduleEntry.DUTY_TYPE_OFF,
            BranchWeeklyScheduleEntry.DUTY_TYPE_EXTRA_OFF,
            BranchWeeklyScheduleEntry.DUTY_TYPE_CUSTOM,
        }
    )
    attendance_shift_locked_value = ""
    if attendance_today_schedule_entry and attendance_today_schedule_entry.duty_type == BranchWeeklyScheduleEntry.DUTY_TYPE_SHIFT:
        attendance_shift_locked_value = resolve_attendance_shift_value(
            label=attendance_today_schedule_entry.shift_label or attendance_today_schedule_label,
            start_time=attendance_today_schedule_entry.start_time,
            end_time=attendance_today_schedule_entry.end_time,
        )
    attendance_shift_locked = bool(attendance_shift_locked_value)
    shift_choices = build_self_service_shift_choices(branch)
    branch_has_attendance_location_config = bool(
        branch and getattr(branch, "has_attendance_location_config", False)
    )
    attendance_event = (
        employee.attendance_events.filter(attendance_date=today).select_related("synced_ledger").first()
    )

    if request.method == "POST":
        action = (request.POST.get("attendance_action") or "").strip()
        form_initial = {"shift": attendance_shift_locked_value} if attendance_shift_locked_value else None
        form = EmployeeSelfServiceAttendanceForm(
            request.POST,
            initial=form_initial,
            shift_choices=shift_choices,
            shift_locked=attendance_shift_locked,
        )
        if not branch_has_attendance_location_config:
            form.add_error(
                None,
                "Your branch does not have a fixed attendance point configured yet. Please contact HR or Operations.",
            )
        elif attendance_blocked_for_today:
            form.add_error(
                None,
                f"Attendance is blocked today because your assigned duty is {attendance_today_schedule_label}.",
            )
        elif action not in {"check_in", "check_out"}:
            form.add_error(None, "Unknown attendance action requested.")

        if form.is_valid():
            actor_label = get_actor_label(request.user) or employee.full_name
            validation_result = get_branch_attendance_validation_result(
                employee,
                form.cleaned_data["latitude"],
                form.cleaned_data["longitude"],
            )
            if not validation_result["is_configured"]:
                form.add_error(None, validation_result["error_message"])
            elif not validation_result["is_inside_radius"]:
                form.add_error(
                    None,
                    (
                        f"Attendance denied. You are {validation_result['distance_meters']} m away from "
                        f"{validation_result['branch'].name}. Allowed radius is "
                        f"{validation_result['allowed_radius_meters']} m."
                    ),
                )

        if form.is_valid():
            actor_label = get_actor_label(request.user) or employee.full_name
            validation_result = get_branch_attendance_validation_result(
                employee,
                form.cleaned_data["latitude"],
                form.cleaned_data["longitude"],
            )
            now = timezone.localtime()
            attendance_event, _created = EmployeeAttendanceEvent.objects.get_or_create(
                employee=employee,
                attendance_date=today,
                defaults={
                    "shift": form.cleaned_data["shift"],
                },
            )
            attendance_event.shift = attendance_event.shift or form.cleaned_data["shift"]
            if action == "check_in":
                if attendance_event.check_in_at:
                    messages.warning(request, "Check-in is already registered for today.")
                else:
                    attendance_event.shift = form.cleaned_data["shift"]
                    attendance_event.check_in_at = now
                    attendance_event.check_in_latitude = form.cleaned_data.get("latitude")
                    attendance_event.check_in_longitude = form.cleaned_data.get("longitude")
                    attendance_event.check_in_location_label = validation_result["branch_location_label"]
                    attendance_event.check_in_address = validation_result["validation_summary"]
                    attendance_event.branch_latitude_used = validation_result["branch_latitude"]
                    attendance_event.branch_longitude_used = validation_result["branch_longitude"]
                    attendance_event.attendance_radius_meters_used = validation_result["allowed_radius_meters"]
                    attendance_event.check_in_distance_meters = validation_result["distance_meters"]
                    attendance_event.check_in_location_validation_status = validation_result["validation_status"]
                    attendance_event.notes = form.cleaned_data.get("notes") or ""
                    attendance_event.status = EmployeeAttendanceEvent.STATUS_OPEN
                    attendance_event.save()
                    messages.success(
                        request,
                        (
                            f"Check-in registered successfully. Device location validated at "
                            f"{validation_result['distance_meters']} m from the branch point."
                        ),
                    )
            elif action == "check_out":
                if not attendance_event.check_in_at:
                    messages.error(request, "Please check in first before checking out.")
                elif attendance_event.check_out_at:
                    messages.warning(request, "Check-out is already registered for today.")
                else:
                    attendance_event.check_out_at = now
                    attendance_event.check_out_latitude = form.cleaned_data.get("latitude")
                    attendance_event.check_out_longitude = form.cleaned_data.get("longitude")
                    attendance_event.check_out_location_label = validation_result["branch_location_label"]
                    attendance_event.check_out_address = validation_result["validation_summary"]
                    attendance_event.branch_latitude_used = validation_result["branch_latitude"]
                    attendance_event.branch_longitude_used = validation_result["branch_longitude"]
                    attendance_event.attendance_radius_meters_used = validation_result["allowed_radius_meters"]
                    attendance_event.check_out_distance_meters = validation_result["distance_meters"]
                    attendance_event.check_out_location_validation_status = validation_result["validation_status"]
                    if form.cleaned_data.get("notes"):
                        attendance_event.notes = form.cleaned_data["notes"]
                    attendance_event.status = EmployeeAttendanceEvent.STATUS_COMPLETED
                    attendance_event.save()
                    synced_ledger = sync_attendance_event_to_ledger(attendance_event, actor_label=actor_label)
                    if synced_ledger:
                        create_employee_history(
                            employee=employee,
                            title="Self-service attendance completed",
                            description=(
                                f"Check-in: {timezone.localtime(attendance_event.check_in_at):%I:%M %p}. "
                                f"Check-out: {timezone.localtime(attendance_event.check_out_at):%I:%M %p}. "
                                f"Shift: {synced_ledger.get_shift_display()}. "
                                f"Check-in distance: {attendance_event.check_in_distance_meters or 0} m. "
                                f"Check-out distance: {attendance_event.check_out_distance_meters or 0} m."
                            ),
                            event_type=EmployeeHistory.EVENT_STATUS,
                            created_by=actor_label,
                            is_system_generated=True,
                            event_date=today,
                        )
                    messages.success(
                        request,
                        (
                            f"Check-out registered and synced to attendance management. Device location validated at "
                            f"{validation_result['distance_meters']} m from the branch point."
                        ),
                    )
            return redirect("employees:self_service_attendance")
        messages.error(request, "Please review the attendance details and try again.")
    else:
        initial = {}
        if attendance_shift_locked_value:
            initial["shift"] = attendance_shift_locked_value
        elif attendance_event and attendance_event.shift:
            initial["shift"] = attendance_event.shift
        form = EmployeeSelfServiceAttendanceForm(
            initial=initial,
            shift_choices=shift_choices,
            shift_locked=attendance_shift_locked,
        )

    attendance_history_queryset = employee.attendance_events.select_related("synced_ledger").order_by(
        "-attendance_date",
        "-check_in_at",
        "-id",
    )
    attendance_history_paginator = Paginator(attendance_history_queryset, 8)
    attendance_history_page_obj = attendance_history_paginator.get_page(request.GET.get("page"))
    attendance_history_query_params = request.GET.copy()
    attendance_history_query_params.pop("page", None)

    context = build_self_service_page_context(
        request,
        employee,
        current_section="attendance",
    )
    context["attendance_event_today"] = attendance_event
    context["attendance_self_service_form"] = form
    context["recent_attendance_events"] = list(attendance_history_page_obj.object_list)
    context["recent_attendance_page_obj"] = attendance_history_page_obj
    context["recent_attendance_pagination_items"] = build_attendance_history_pagination(
        attendance_history_page_obj
    )
    context["recent_attendance_querystring"] = attendance_history_query_params.urlencode()
    context["attendance_branch"] = branch
    context["attendance_branch_has_location_config"] = branch_has_attendance_location_config
    context["attendance_branch_latitude"] = getattr(branch, "attendance_latitude", None)
    context["attendance_branch_longitude"] = getattr(branch, "attendance_longitude", None)
    context["attendance_branch_radius_meters"] = getattr(branch, "attendance_radius_meters", None)
    context["attendance_today_schedule_entry"] = attendance_today_schedule_entry
    context["attendance_today_schedule_label"] = attendance_today_schedule_label
    context["attendance_today_schedule_time"] = attendance_today_schedule_time
    context["attendance_shift_locked"] = attendance_shift_locked
    context["attendance_blocked_for_today"] = attendance_blocked_for_today
    context["today"] = today
    return render(request, "employees/self_service_attendance.html", context)


class EmployeeCreateView(LoginRequiredMixin, CreateView):
    model = Employee
    form_class = EmployeeForm
    template_name = "employees/employee_form.html"
    success_url = reverse_lazy("employees:employee_list")

    def dispatch(self, request, *args, **kwargs):
        if not can_create_or_edit_employees(request.user):
            return deny_employee_access(request, "You do not have permission to create employee profiles.")
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        response = super().form_valid(form)

        create_employee_history(
            employee=self.object,
            title="Employee profile created",
            description="Employee record was created in the HR system.",
            event_type=EmployeeHistory.EVENT_PROFILE,
            created_by=get_actor_label(self.request.user),
            is_system_generated=True,
            event_date=self.object.hire_date or timezone.localdate(),
        )

        messages.success(self.request, "Employee created successfully.")
        return response
    def form_invalid(self, form):
        messages.error(self.request, "Employee could not be saved. Please review the form errors and try again.")
        return super().form_invalid(form)


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create Employee"
        context["submit_label"] = "Create Employee"
        return context


class EmployeeUpdateView(LoginRequiredMixin, UpdateView):
    model = Employee
    form_class = EmployeeForm
    template_name = "employees/employee_form.html"
    success_url = reverse_lazy("employees:employee_list")

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if not can_create_or_edit_employees(request.user):
            return deny_employee_access(
                request,
                "You do not have permission to update employee profiles.",
                employee=self.object,
            )
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        original_employee = Employee.objects.get(pk=self.object.pk)
        response = super().form_valid(form)

        create_employee_history(
            employee=self.object,
            title="Employee profile updated",
            description=build_employee_change_summary(original_employee, self.object),
            event_type=EmployeeHistory.EVENT_PROFILE,
            created_by=get_actor_label(self.request.user),
            is_system_generated=True,
            event_date=timezone.localdate(),
        )

        for history_message in getattr(form, "account_history_messages", []):
            create_employee_history(
                employee=self.object,
                title=history_message,
                description=history_message,
                event_type=EmployeeHistory.EVENT_PROFILE,
                created_by=get_actor_label(self.request.user),
                is_system_generated=True,
                event_date=timezone.localdate(),
            )

        messages.success(self.request, "Employee updated successfully.")
        return response
    def form_invalid(self, form):
        messages.error(self.request, "Employee could not be updated. Please review the form errors and try again.")
        return super().form_invalid(form)
    
    def get_success_url(self):
        return reverse("employees:employee_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Edit Employee"
        context["submit_label"] = "Save Changes"
        return context


class EmployeeTransferView(LoginRequiredMixin, UpdateView):
    model = Employee
    form_class = EmployeeTransferForm
    template_name = "employees/employee_transfer.html"

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if not can_transfer_employee(request.user):
            return deny_employee_access(
                request,
                "You do not have permission to transfer employee placements.",
                employee=self.object,
            )
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        original_employee = Employee.objects.get(pk=self.object.pk)
        response = super().form_valid(form)

        transfer_note = form.cleaned_data.get("notes", "")
        create_employee_history(
            employee=self.object,
            title="Employee placement transferred",
            description=build_employee_transfer_summary(original_employee, self.object, transfer_note=transfer_note),
            event_type=EmployeeHistory.EVENT_TRANSFER,
            created_by=get_actor_label(self.request.user),
            is_system_generated=True,
            event_date=timezone.localdate(),
        )

        messages.success(self.request, "Employee placement updated successfully.")
        return response

    def get_success_url(self):
        return reverse("employees:employee_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Transfer Employee"
        context["submit_label"] = "Save Placement Change"
        return context


class EmployeeDeleteView(LoginRequiredMixin, ProtectedDeleteMixin, DeleteView):
    model = Employee
    template_name = "employees/employee_confirm_delete.html"
    success_url = reverse_lazy("employees:employee_list")

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if not can_delete_employee(request.user):
            return deny_employee_access(
                request,
                "You do not have permission to delete employee profiles.",
                employee=self.object,
            )
        return super().dispatch(request, *args, **kwargs)

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        employee_name = self.object.full_name
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f"Employee '{employee_name}' deleted successfully.")
        return response


@login_required
@require_POST
def employee_status_update(request, pk):
    employee = get_object_or_404(Employee, pk=pk)

    if not can_change_employee_status(request.user):
        return deny_employee_access(
            request,
            "You do not have permission to change employee status.",
            employee=employee,
        )

    target_status = request.POST.get("target_status", "").strip()
    valid_statuses = {value for value, _label in Employee.EMPLOYMENT_STATUS_CHOICES}
    if target_status not in valid_statuses:
        messages.error(request, "Invalid employee status action.")
        return redirect("employees:employee_detail", pk=employee.pk)

    employee.employment_status = target_status
    employee.is_active = target_status != Employee.EMPLOYMENT_STATUS_INACTIVE
    employee.save(update_fields=["employment_status", "is_active", "updated_at"])

    create_employee_history(
        employee=employee,
        title="Employee status updated",
        description=f"Employee status changed to {employee.get_employment_status_display()}.",
        event_type=EmployeeHistory.EVENT_STATUS,
        created_by=get_actor_label(request.user),
        is_system_generated=True,
        event_date=timezone.localdate(),
    )

    messages.success(request, "Employee status updated successfully.")
    return redirect("employees:employee_detail", pk=employee.pk)


@login_required
def employee_document_view(request, employee_pk, document_pk):
    employee = get_object_or_404(Employee, pk=employee_pk)
    document = get_object_or_404(EmployeeDocument, pk=document_pk, employee=employee)

    if not can_view_employee_profile(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to view this employee document.",
            employee=employee,
        )

    return build_browser_file_response(document.file, force_download=False)


@login_required
def employee_document_download(request, employee_pk, document_pk):
    employee = get_object_or_404(Employee, pk=employee_pk)
    document = get_object_or_404(EmployeeDocument, pk=document_pk, employee=employee)

    if not can_view_employee_profile(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to download this employee document.",
            employee=employee,
        )

    return build_browser_file_response(document.file, force_download=True)


@login_required
def employee_required_submission_response_view(request, request_pk):
    submission_request = get_object_or_404(
        EmployeeRequiredSubmission.objects.select_related("employee"),
        pk=request_pk,
    )
    employee = submission_request.employee

    if not can_view_employee_profile(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to view this submitted file.",
            employee=employee,
        )

    return build_browser_file_response(submission_request.response_file, force_download=False)


@login_required
def employee_required_submission_response_download(request, request_pk):
    submission_request = get_object_or_404(
        EmployeeRequiredSubmission.objects.select_related("employee"),
        pk=request_pk,
    )
    employee = submission_request.employee

    if not can_view_employee_profile(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to download this submitted file.",
            employee=employee,
        )

    return build_browser_file_response(submission_request.response_file, force_download=True)


@login_required
def employee_document_request_response_view(request, request_pk):
    document_request = get_object_or_404(
        EmployeeDocumentRequest.objects.select_related("employee"),
        pk=request_pk,
    )
    employee = document_request.employee

    if not can_view_employee_profile(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to view this reply file.",
            employee=employee,
        )

    return build_browser_file_response(document_request.response_file, force_download=False)


@login_required
def employee_document_request_response_download(request, request_pk):
    document_request = get_object_or_404(
        EmployeeDocumentRequest.objects.select_related("employee"),
        pk=request_pk,
    )
    employee = document_request.employee

    if not can_view_employee_profile(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to download this reply file.",
            employee=employee,
        )

    return build_browser_file_response(document_request.response_file, force_download=True)



@login_required
def employee_document_update(request, employee_pk, document_pk):
    employee = get_object_or_404(Employee, pk=employee_pk)
    document = get_object_or_404(EmployeeDocument, pk=document_pk, employee=employee)

    if not can_manage_employee_documents(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to update employee documents.",
            employee=employee,
        )

    if request.method == "POST":
        form = EmployeeDocumentForm(request.POST, request.FILES, instance=document)
        if form.is_valid():
            document = form.save()

            create_employee_history(
                employee=employee,
                title=f"Document updated: {document.title or document.filename}",
                description=build_document_summary(document),
                event_type=EmployeeHistory.EVENT_DOCUMENT,
                created_by=get_actor_label(request.user),
                is_system_generated=True,
                event_date=document.issue_date or timezone.localdate(),
            )
            messages.success(request, "Document updated successfully.")
    return redirect("employees:employee_detail", pk=employee.pk)


@login_required
@require_POST
def employee_document_create(request, pk):
    employee = get_object_or_404(Employee, pk=pk)

    if not can_manage_employee_documents(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to upload employee documents.",
            employee=employee,
        )

    form = EmployeeDocumentForm(request.POST, request.FILES)

    if form.is_valid():
        document = form.save(commit=False)
        document.employee = employee
        document.save()

        create_employee_history(
            employee=employee,
            title=f"Document uploaded: {document.title or document.filename}",
            description=build_document_summary(document),
            event_type=EmployeeHistory.EVENT_DOCUMENT,
            created_by=get_actor_label(request.user),
            is_system_generated=True,
            event_date=document.issue_date or document.uploaded_at.date(),
        )

        messages.success(request, "Document uploaded successfully.")
    else:
        messages.error(request, "Please review the document form and try again.")

    return redirect("employees:employee_detail", pk=employee.pk)


@login_required
@require_POST
def employee_document_delete(request, employee_pk, document_pk):
    employee = get_object_or_404(Employee, pk=employee_pk)
    document = get_object_or_404(EmployeeDocument, pk=document_pk, employee=employee)

    if not can_manage_employee_documents(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to delete employee documents.",
            employee=employee,
        )

    document_label = document.title or getattr(document, "filename", "") or str(document)
    document.delete()

    create_employee_history(
        employee=employee,
        title=f"Document deleted: {document_label}",
        description="Employee document was deleted from the employee profile.",
        event_type=EmployeeHistory.EVENT_DOCUMENT,
        created_by=get_actor_label(request.user),
        is_system_generated=True,
        event_date=timezone.localdate(),
    )

    messages.success(request, "Document deleted successfully.")
    return redirect("employees:employee_detail", pk=employee.pk)



@login_required
@require_POST
def employee_required_submission_create(request, pk):
    employee = get_object_or_404(Employee, pk=pk)

    if not can_manage_employee_required_submissions(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to create employee required submission requests.",
            employee=employee,
        )

    form = EmployeeRequiredSubmissionCreateForm(request.POST)
    if form.is_valid():
        submission_request = form.save(commit=False)
        submission_request.employee = employee
        submission_request.created_by = request.user
        submission_request.status = EmployeeRequiredSubmission.STATUS_REQUESTED
        submission_request.reviewed_by = None
        submission_request.review_note = ""
        submission_request.reviewed_at = None
        submission_request.submitted_at = None
        submission_request.save()

        create_employee_history(
            employee=employee,
            title=f"Required employee submission requested: {submission_request.title}",
            description=(
                f"Request Type: {submission_request.get_request_type_display()}. "
                f"Priority: {submission_request.get_priority_display()}. "
                + (
                    f"Due Date: {submission_request.due_date.strftime('%B %d, %Y')}. "
                    if submission_request.due_date else ""
                )
                + (f"Instructions: {submission_request.instructions}" if submission_request.instructions else "")
            ).strip(),
            event_type=EmployeeHistory.EVENT_DOCUMENT,
            created_by=get_actor_label(request.user),
            is_system_generated=True,
            event_date=timezone.localdate(),
        )
        messages.success(request, "Required employee submission request created successfully.")
    else:
        first_error = "Please review the required submission request form and try again."
        if form.errors:
            first_field_errors = next(iter(form.errors.values()))
            if first_field_errors:
                first_error = first_field_errors[0]
        messages.error(request, first_error)

    return redirect(
        get_safe_next_url(
            request,
            reverse("employees:employee_detail", kwargs={"pk": employee.pk}),
        )
    )


@login_required
@require_POST
def employee_required_submission_submit(request, request_pk):
    submission_request = get_object_or_404(
        EmployeeRequiredSubmission.objects.select_related('employee', 'created_by', 'reviewed_by'),
        pk=request_pk,
    )
    employee = submission_request.employee

    if not can_submit_employee_required_submission(request.user, submission_request):
        return deny_employee_access(
            request,
            "You do not have permission to submit this required employee request.",
            employee=employee,
        )

    form = EmployeeRequiredSubmissionResponseForm(
        request.POST,
        request.FILES,
        instance=submission_request,
    )
    if form.is_valid():
        submission_request = form.save(commit=False)
        submission_request.status = EmployeeRequiredSubmission.STATUS_SUBMITTED
        submission_request.submitted_at = timezone.now()
        submission_request.reviewed_by = None
        submission_request.reviewed_at = None
        submission_request.review_note = ""
        submission_request.save()

        request_type_document_map = {
            EmployeeRequiredSubmission.REQUEST_TYPE_PASSPORT_COPY: EmployeeDocument.DOCUMENT_TYPE_ID,
            EmployeeRequiredSubmission.REQUEST_TYPE_CIVIL_ID_COPY: EmployeeDocument.DOCUMENT_TYPE_ID,
            EmployeeRequiredSubmission.REQUEST_TYPE_CONTRACT_COPY: EmployeeDocument.DOCUMENT_TYPE_CONTRACT,
            EmployeeRequiredSubmission.REQUEST_TYPE_MEDICAL_DOCUMENT: EmployeeDocument.DOCUMENT_TYPE_MEDICAL,
            EmployeeRequiredSubmission.REQUEST_TYPE_CERTIFICATE: EmployeeDocument.DOCUMENT_TYPE_CERTIFICATE,
            EmployeeRequiredSubmission.REQUEST_TYPE_GENERAL_DOCUMENT: EmployeeDocument.DOCUMENT_TYPE_GENERAL,
            EmployeeRequiredSubmission.REQUEST_TYPE_OTHER: EmployeeDocument.DOCUMENT_TYPE_OTHER,
        }

        fulfilled_document = EmployeeDocument.objects.create(
            employee=employee,
            title=submission_request.title,
            document_type=request_type_document_map.get(
                submission_request.request_type,
                EmployeeDocument.DOCUMENT_TYPE_GENERAL,
            ),
            reference_number=submission_request.response_reference_number or "",
            issue_date=submission_request.response_issue_date,
            expiry_date=submission_request.response_expiry_date,
            is_required=True,
            file=submission_request.response_file,
            description=(submission_request.employee_note or submission_request.instructions or "").strip(),
        )
        submission_request.fulfilled_document = fulfilled_document
        submission_request.save(update_fields=['fulfilled_document', 'updated_at'])

        create_employee_history(
            employee=employee,
            title=f"Employee submitted requested file: {submission_request.title}",
            description=(
                f"Request Type: {submission_request.get_request_type_display()}. "
                f"Status: {submission_request.get_status_display()}. "
                + (
                    f"Reference Number: {submission_request.response_reference_number}. "
                    if submission_request.response_reference_number else ""
                )
                + (f"Employee Note: {submission_request.employee_note}" if submission_request.employee_note else "")
            ).strip(),
            event_type=EmployeeHistory.EVENT_DOCUMENT,
            created_by=get_actor_label(request.user),
            is_system_generated=True,
            event_date=timezone.localdate(),
        )
        messages.success(request, "Requested file submitted successfully.")
    else:
        first_error = "Please review the requested file submission form and try again."
        if form.errors:
            first_field_errors = next(iter(form.errors.values()))
            if first_field_errors:
                first_error = first_field_errors[0]
        messages.error(request, first_error)

    return redirect(
        get_safe_next_url(
            request,
            reverse("employees:employee_detail", kwargs={"pk": employee.pk}),
        )
    )


@login_required
@require_POST
def employee_required_submission_review(request, request_pk):
    submission_request = get_object_or_404(
        EmployeeRequiredSubmission.objects.select_related('employee', 'created_by', 'reviewed_by'),
        pk=request_pk,
    )
    employee = submission_request.employee

    if not can_review_employee_required_submission(request.user, submission_request):
        return deny_employee_access(
            request,
            "You do not have permission to review this required employee request.",
            employee=employee,
        )

    form = EmployeeRequiredSubmissionReviewForm(request.POST, instance=submission_request)
    if form.is_valid():
        updated_request = form.save(commit=False)
        updated_request.reviewed_by = request.user
        updated_request.reviewed_at = timezone.now()

        if updated_request.status == EmployeeRequiredSubmission.STATUS_COMPLETED and not updated_request.submitted_at:
            updated_request.submitted_at = timezone.now()

        updated_request.save()

        create_employee_history(
            employee=employee,
            title=f"Required employee submission reviewed: {updated_request.title}",
            description=(
                f"Review Status: {updated_request.get_status_display()}. "
                + (f"Review Note: {updated_request.review_note}" if updated_request.review_note else "")
            ).strip(),
            event_type=EmployeeHistory.EVENT_DOCUMENT,
            created_by=get_actor_label(request.user),
            is_system_generated=True,
            event_date=timezone.localdate(),
        )
        messages.success(request, "Required employee submission reviewed successfully.")
    else:
        first_error = "Please review the submission review form and try again."
        if form.errors:
            first_field_errors = next(iter(form.errors.values()))
            if first_field_errors:
                first_error = first_field_errors[0]
        messages.error(request, first_error)

    return redirect('employees:employee_detail', pk=employee.pk)

@login_required
@require_POST
def employee_document_request_create(request, pk):
    employee = get_object_or_404(Employee, pk=pk)

    if not can_create_employee_document_request(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to request management documents for this employee profile.",
            employee=employee,
        )

    form = EmployeeDocumentRequestCreateForm(request.POST)
    if form.is_valid():
        document_request = form.save(commit=False)
        document_request.employee = employee
        document_request.created_by = request.user
        document_request.status = EmployeeDocumentRequest.STATUS_REQUESTED
        document_request.submitted_at = timezone.now()
        document_request.save()

        create_employee_history(
            employee=employee,
            title=f"Document requested: {document_request.title}",
            description=build_employee_document_request_summary(document_request),
            event_type=EmployeeHistory.EVENT_NOTE,
            created_by=get_actor_label(request.user),
            is_system_generated=True,
            event_date=timezone.localdate(),
        )

        messages.success(request, "Document request submitted successfully.")
    else:
        first_error = "Please review the document request form and try again."
        if form.errors:
            first_field_errors = next(iter(form.errors.values()))
            if first_field_errors:
                first_error = first_field_errors[0]
        messages.error(request, first_error)

    return redirect(
        get_safe_next_url(
            request,
            reverse("employees:employee_detail", kwargs={"pk": employee.pk}),
        )
    )


@login_required
@require_POST
def employee_document_request_review(request, request_pk):
    document_request = get_object_or_404(
        EmployeeDocumentRequest.objects.select_related("employee", "reviewed_by", "created_by", "delivered_document"),
        pk=request_pk,
    )
    employee = document_request.employee

    if not can_review_employee_document_request(request.user, document_request):
        return deny_employee_access(
            request,
            "You do not have permission to review this employee document request.",
            employee=employee,
        )

    form = EmployeeDocumentRequestReviewForm(request.POST, request.FILES, instance=document_request)
    if form.is_valid():
        previous_status = document_request.get_status_display()
        updated_request = form.save(commit=False)
        updated_request.reviewed_by = request.user
        updated_request.reviewed_at = timezone.now()

        if updated_request.status == EmployeeDocumentRequest.STATUS_COMPLETED:
            if not updated_request.completed_at:
                updated_request.completed_at = timezone.now()
        else:
            updated_request.completed_at = None

        updated_request.save()

        if updated_request.response_file and updated_request.status in {
            EmployeeDocumentRequest.STATUS_APPROVED,
            EmployeeDocumentRequest.STATUS_COMPLETED,
        }:
            delivered_document = updated_request.delivered_document
            if delivered_document is None:
                delivered_document = EmployeeDocument(
                    employee=employee,
                    title=updated_request.default_document_title,
                    document_type=updated_request.mapped_document_type,
                    file=updated_request.response_file,
                    description=updated_request.management_note or f"Delivered from employee document request: {updated_request.get_request_type_display()}.",
                )
            else:
                delivered_document.employee = employee
                delivered_document.title = updated_request.default_document_title
                delivered_document.document_type = updated_request.mapped_document_type
                delivered_document.file = updated_request.response_file
                delivered_document.description = updated_request.management_note or delivered_document.description

            delivered_document.save()

            if updated_request.delivered_document_id != delivered_document.pk:
                updated_request.delivered_document = delivered_document
                updated_request.save(update_fields=["delivered_document", "updated_at"])

        create_employee_history(
            employee=employee,
            title=f"Document request reviewed: {updated_request.title}",
            description=build_employee_document_request_review_summary(
                document_request=updated_request,
                previous_status=previous_status,
                new_status=updated_request.get_status_display(),
                management_note=updated_request.management_note,
            ),
            event_type=EmployeeHistory.EVENT_NOTE,
            created_by=get_actor_label(request.user),
            is_system_generated=True,
            event_date=timezone.localdate(),
        )

        messages.success(request, "Employee document request updated successfully.")
    else:
        first_error = "Please review the employee document request review form and try again."
        if form.errors:
            first_field_errors = next(iter(form.errors.values()))
            if first_field_errors:
                first_error = first_field_errors[0]
        messages.error(request, first_error)

    return redirect("employees:employee_requests_overview")


@login_required
@require_POST
def employee_document_request_cancel(request, request_pk):
    document_request = get_object_or_404(EmployeeDocumentRequest.objects.select_related("employee"), pk=request_pk)
    employee = document_request.employee

    if not can_cancel_employee_document_request(request.user, document_request):
        return deny_employee_access(
            request,
            "You do not have permission to cancel this document request.",
            employee=employee,
        )

    previous_status = document_request.get_status_display()
    document_request.status = EmployeeDocumentRequest.STATUS_CANCELLED
    document_request.reviewed_at = timezone.now()
    document_request.reviewed_by = request.user
    document_request.save(update_fields=["status", "reviewed_at", "reviewed_by", "updated_at"])

    create_employee_history(
        employee=employee,
        title=f"Document request cancelled: {document_request.title}",
        description=build_employee_document_request_review_summary(
            document_request=document_request,
            previous_status=previous_status,
            new_status=document_request.get_status_display(),
            management_note="Cancelled by employee.",
        ),
        event_type=EmployeeHistory.EVENT_NOTE,
        created_by=get_actor_label(request.user),
        is_system_generated=True,
        event_date=timezone.localdate(),
    )

    messages.success(request, "Document request cancelled successfully.")
    return redirect(
        get_safe_next_url(
            request,
            reverse("employees:employee_detail", kwargs={"pk": employee.pk}),
        )
    )


@login_required
@require_POST
def employee_leave_create(request, pk):
    employee = get_object_or_404(Employee, pk=pk)

    if not can_request_leave(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to create leave requests for this employee.",
            employee=employee,
        )

    if is_self_employee(request.user, employee):
        form = EmployeeSelfServiceLeaveRequestForm(request.POST, request.FILES)
    else:
        form = EmployeeLeaveForm(request.POST)

    if form.is_valid():
        leave_record = form.save(commit=False)
        leave_record.employee = employee
        leave_record.requested_by = request.user
        leave_record.created_by = get_actor_label(request.user)
        leave_record.updated_by = get_actor_label(request.user)
        leave_record.status = EmployeeLeave.STATUS_PENDING
        leave_record.current_stage = EmployeeLeave.STAGE_SUPERVISOR_REVIEW
        leave_record.save()

        supporting_document = None
        attachment_file = form.cleaned_data.get("attachment_file") if hasattr(form, "cleaned_data") else None

        if attachment_file:
            attachment_title = form.cleaned_data.get("attachment_title") or (
                f"{leave_record.get_leave_type_display()} Supporting Document"
            )
            supporting_document = EmployeeDocument.objects.create(
                employee=employee,
                linked_leave=leave_record,
                title=attachment_title,
                document_type=form.cleaned_data.get("attachment_document_type") or EmployeeDocument.DOCUMENT_TYPE_OTHER,
                reference_number=form.cleaned_data.get("attachment_reference_number", ""),
                issue_date=form.cleaned_data.get("attachment_issue_date"),
                expiry_date=form.cleaned_data.get("attachment_expiry_date"),
                file=attachment_file,
                description=form.cleaned_data.get("attachment_description", ""),
            )

        leave_history_description = build_leave_request_summary(leave_record)
        if supporting_document:
            leave_history_description += f" Supporting document uploaded: {supporting_document.title or supporting_document.filename}."

        create_employee_history(
            employee=employee,
            title=f"Leave requested: {leave_record.get_leave_type_display()}",
            description=leave_history_description,
            event_type=EmployeeHistory.EVENT_STATUS,
            created_by=get_actor_label(request.user),
            is_system_generated=True,
            event_date=leave_record.start_date,
        )

        if supporting_document:
            create_employee_history(
                employee=employee,
                title=f"Document uploaded: {supporting_document.title or supporting_document.filename}",
                description=build_document_summary(supporting_document),
                event_type=EmployeeHistory.EVENT_DOCUMENT,
                created_by=get_actor_label(request.user),
                is_system_generated=True,
                event_date=supporting_document.issue_date or timezone.localdate(),
            )

        messages.success(request, "Leave request submitted successfully.")
    else:
        first_error = "Please review the leave request form and try again."
        if form.errors:
            first_field_errors = next(iter(form.errors.values()))
            if first_field_errors:
                first_error = first_field_errors[0]
        messages.error(request, first_error)

    return redirect("employees:employee_detail", pk=employee.pk)


@login_required
@require_POST
def employee_leave_approve(request, employee_pk, leave_pk):
    employee = get_object_or_404(Employee, pk=employee_pk)
    leave_record = get_object_or_404(EmployeeLeave, pk=leave_pk, employee=employee)

    if not can_review_leave(request.user):
        return deny_employee_access(
            request,
            "You do not have permission to approve leave requests.",
            employee=employee,
        )

    if is_branch_scoped_supervisor(request.user) and not can_supervisor_view_employee(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to review leave requests outside your branch.",
            employee=employee,
        )

    if leave_record.status != EmployeeLeave.STATUS_PENDING:
        messages.error(request, "Only pending leave requests can be approved from this workflow action.")
        return redirect("employees:employee_detail", pk=employee.pk)

    if not can_user_review_leave_stage(request.user, leave_record):
        messages.error(
            request,
            f"This leave request is currently assigned to {get_leave_current_stage_owner_label(leave_record)} and cannot be approved from your workflow step.",
        )
        return redirect("employees:employee_detail", pk=employee.pk)

    previous_status = leave_record.get_status_display()
    approval_note = (request.POST.get("approval_note") or "").strip()
    actor_label = get_actor_label(request.user)
    current_time = timezone.now()

    leave_record.reviewed_by = request.user
    leave_record.rejected_by = None
    leave_record.cancelled_by = None
    leave_record.approval_note = approval_note
    leave_record.updated_by = actor_label

    history_title = f"Leave approved: {leave_record.get_leave_type_display()}"
    success_message = "Leave request approved successfully and the workflow was updated."

    if is_branch_scoped_supervisor(request.user):
        leave_record.status = EmployeeLeave.STATUS_PENDING
        leave_record.current_stage = EmployeeLeave.STAGE_OPERATIONS_REVIEW
        leave_record.supervisor_reviewed_by = request.user
        leave_record.supervisor_reviewed_at = current_time
        leave_record.supervisor_review_note = approval_note
        history_title = f"Leave moved to operations review: {leave_record.get_leave_type_display()}"
        success_message = "Leave request reviewed and moved to Operations for the next stage."
    elif is_operations_manager_user(request.user):
        leave_record.status = EmployeeLeave.STATUS_PENDING
        leave_record.current_stage = EmployeeLeave.STAGE_HR_REVIEW
        leave_record.operations_reviewed_by = request.user
        leave_record.operations_reviewed_at = current_time
        leave_record.operations_review_note = approval_note
        history_title = f"Leave moved to HR review: {leave_record.get_leave_type_display()}"
        success_message = "Leave request reviewed and moved to HR for final review."
    else:
        leave_record.status = EmployeeLeave.STATUS_APPROVED
        leave_record.current_stage = EmployeeLeave.STAGE_FINAL_APPROVED
        leave_record.approved_by = request.user
        leave_record.hr_reviewed_by = request.user
        leave_record.hr_reviewed_at = current_time
        leave_record.hr_review_note = approval_note
        leave_record.finalized_at = current_time

    leave_record.save()

    create_employee_history(
        employee=employee,
        title=history_title,
        description=build_leave_status_summary(
            leave_record=leave_record,
            previous_status=previous_status,
            new_status=leave_record.get_status_display(),
            approval_note=approval_note,
        ),
        event_type=EmployeeHistory.EVENT_STATUS,
        created_by=actor_label,
        is_system_generated=True,
        event_date=leave_record.start_date,
    )

    messages.success(request, success_message)
    return redirect("employees:employee_detail", pk=employee.pk)


@login_required
@require_POST
def employee_leave_reject(request, employee_pk, leave_pk):
    employee = get_object_or_404(Employee, pk=employee_pk)
    leave_record = get_object_or_404(EmployeeLeave, pk=leave_pk, employee=employee)

    if not can_review_leave(request.user):
        return deny_employee_access(
            request,
            "You do not have permission to reject leave requests.",
            employee=employee,
        )

    if is_branch_scoped_supervisor(request.user) and not can_supervisor_view_employee(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to review leave requests outside your branch.",
            employee=employee,
        )

    if leave_record.status != EmployeeLeave.STATUS_PENDING:
        messages.error(request, "Only pending leave requests can be rejected from this workflow action.")
        return redirect("employees:employee_detail", pk=employee.pk)

    if not can_user_review_leave_stage(request.user, leave_record):
        messages.error(
            request,
            f"This leave request is currently assigned to {get_leave_current_stage_owner_label(leave_record)} and cannot be rejected from your workflow step.",
        )
        return redirect("employees:employee_detail", pk=employee.pk)

    previous_status = leave_record.get_status_display()
    approval_note = (request.POST.get("approval_note") or "").strip()
    actor_label = get_actor_label(request.user)
    current_time = timezone.now()

    leave_record.status = EmployeeLeave.STATUS_REJECTED
    leave_record.current_stage = EmployeeLeave.STAGE_FINAL_REJECTED
    leave_record.reviewed_by = request.user
    leave_record.rejected_by = request.user
    leave_record.approved_by = None
    leave_record.cancelled_by = None
    leave_record.approval_note = approval_note
    leave_record.updated_by = actor_label
    leave_record.finalized_at = current_time

    if is_branch_scoped_supervisor(request.user):
        leave_record.supervisor_reviewed_by = request.user
        leave_record.supervisor_reviewed_at = current_time
        leave_record.supervisor_review_note = approval_note
    elif is_operations_manager_user(request.user):
        leave_record.operations_reviewed_by = request.user
        leave_record.operations_reviewed_at = current_time
        leave_record.operations_review_note = approval_note
    else:
        leave_record.hr_reviewed_by = request.user
        leave_record.hr_reviewed_at = current_time
        leave_record.hr_review_note = approval_note

    leave_record.save()

    create_employee_history(
        employee=employee,
        title=f"Leave rejected: {leave_record.get_leave_type_display()}",
        description=build_leave_status_summary(
            leave_record=leave_record,
            previous_status=previous_status,
            new_status=leave_record.get_status_display(),
            approval_note=approval_note,
        ),
        event_type=EmployeeHistory.EVENT_STATUS,
        created_by=actor_label,
        is_system_generated=True,
        event_date=leave_record.start_date,
    )

    messages.success(request, "Leave request rejected and closed as a final workflow outcome.")
    return redirect("employees:employee_detail", pk=employee.pk)


@login_required
@require_POST
def employee_leave_cancel(request, employee_pk, leave_pk):
    employee = get_object_or_404(Employee, pk=employee_pk)
    leave_record = get_object_or_404(EmployeeLeave, pk=leave_pk, employee=employee)

    if not can_cancel_leave(request.user, leave_record):
        return deny_employee_access(
            request,
            "You do not have permission to cancel this leave request.",
            employee=employee,
        )

    if is_branch_scoped_supervisor(request.user) and not can_supervisor_view_employee(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to review leave requests outside your branch.",
            employee=employee,
        )

    previous_status = leave_record.get_status_display()
    approval_note = (request.POST.get("approval_note") or "").strip()
    actor_label = get_actor_label(request.user)
    current_time = timezone.now()

    leave_record.status = EmployeeLeave.STATUS_CANCELLED
    leave_record.current_stage = EmployeeLeave.STAGE_CANCELLED
    leave_record.reviewed_by = request.user if can_review_leave(request.user) else leave_record.reviewed_by
    leave_record.approved_by = None
    leave_record.rejected_by = None
    leave_record.cancelled_by = request.user
    leave_record.approval_note = approval_note
    leave_record.updated_by = actor_label
    leave_record.finalized_at = current_time
    leave_record.save()

    create_employee_history(
        employee=employee,
        title=f"Leave cancelled: {leave_record.get_leave_type_display()}",
        description=build_leave_status_summary(
            leave_record=leave_record,
            previous_status=previous_status,
            new_status=leave_record.get_status_display(),
            approval_note=approval_note,
        ),
        event_type=EmployeeHistory.EVENT_STATUS,
        created_by=actor_label,
        is_system_generated=True,
        event_date=leave_record.start_date,
    )

    messages.success(request, "Leave request cancelled successfully.")
    return redirect("employees:employee_detail", pk=employee.pk)


@login_required
@require_POST
def employee_action_record_create(request, pk):
    employee = get_object_or_404(Employee, pk=pk)

    if not can_create_action_records(request.user):
        return deny_employee_access(
            request,
            "You do not have permission to create attendance / incident / discipline records.",
            employee=employee,
        )

    if is_branch_scoped_supervisor(request.user) and not can_supervisor_view_employee(request.user, employee):
        return deny_employee_access(
            request,
            "You do not have permission to create action records outside your branch scope.",
            employee=employee,
        )

    form = EmployeeActionRecordForm(request.POST)

    if form.is_valid():
        action_record = form.save(commit=False)
        action_record.employee = employee
        actor_label = get_actor_label(request.user)
        action_record.created_by = actor_label
        action_record.updated_by = actor_label
        action_record.save()

        create_employee_history(
            employee=employee,
            title=f"Action record added: {action_record.title}",
            description=build_action_record_summary(action_record),
            event_type=EmployeeHistory.EVENT_STATUS,
            created_by=actor_label,
            is_system_generated=True,
            event_date=action_record.action_date,
        )

        messages.success(request, "Attendance / incident record added successfully.")
    else:
        first_error = "Please review the attendance / incident form and try again."
        if form.errors:
            first_field_errors = next(iter(form.errors.values()))
            if first_field_errors:
                first_error = first_field_errors[0]
        messages.error(request, first_error)

    return redirect("employees:employee_detail", pk=employee.pk)


@login_required
@require_POST
def employee_attendance_create(request, pk):
    employee = get_object_or_404(Employee, pk=pk)

    if not can_manage_attendance_records(request.user):
        return deny_employee_access(
            request,
            "You do not have permission to create attendance ledger entries.",
            employee=employee,
        )

    form = EmployeeAttendanceLedgerForm(request.POST, employee=employee)

    if form.is_valid():
        attendance_entry = form.save(commit=False)
        attendance_entry.employee = employee
        attendance_entry.source = EmployeeAttendanceLedger.SOURCE_MANUAL
        actor_label = get_actor_label(request.user)
        attendance_entry.created_by = actor_label
        attendance_entry.updated_by = actor_label
        attendance_entry.save()

        create_employee_history(
            employee=employee,
            title=f"Attendance ledger entry added: {attendance_entry.attendance_date}",
            description=build_attendance_ledger_summary(attendance_entry),
            event_type=EmployeeHistory.EVENT_STATUS,
            created_by=actor_label,
            is_system_generated=True,
            event_date=attendance_entry.attendance_date,
        )

        messages.success(request, "Attendance ledger entry added successfully.")
    else:
        first_error = "Please review the attendance ledger form and try again."
        if form.errors:
            first_field_errors = next(iter(form.errors.values()))
            if first_field_errors:
                first_error = first_field_errors[0]
        messages.error(request, first_error)

    return redirect("employees:employee_detail", pk=employee.pk)


@login_required
@require_POST
def employee_history_create(request, pk):
    employee = get_object_or_404(Employee, pk=pk)

    if not can_add_manual_history(request.user):
        return deny_employee_access(
            request,
            "You do not have permission to add timeline history entries.",
            employee=employee,
        )

    form = EmployeeHistoryForm(request.POST)

    if form.is_valid():
        history_entry = form.save(commit=False)
        history_entry.employee = employee
        history_entry.created_by = get_actor_label(request.user)
        history_entry.save()
        messages.success(request, "Timeline entry added successfully.")
    else:
        messages.error(request, "Please review the timeline entry form and try again.")

    return redirect("employees:employee_detail", pk=employee.pk)


@login_required
@login_required
@require_POST
def employee_attendance_correction_create(request, attendance_pk):
    attendance_entry = get_object_or_404(
        EmployeeAttendanceLedger.objects.select_related("employee"),
        pk=attendance_pk,
    )
    employee = attendance_entry.employee

    if not can_manage_attendance_records(request.user):
        return deny_employee_access(
            request,
            "You do not have permission to request attendance corrections.",
            employee=employee,
        )

    next_url = (request.POST.get("next") or reverse("employees:attendance_management")).strip()
    form = EmployeeAttendanceCorrectionForm(request.POST, attendance_entry=attendance_entry)

    if form.is_valid():
        correction = form.save(commit=False)
        actor_label = get_actor_label(request.user)
        correction.linked_attendance = attendance_entry
        correction.employee = employee
        correction.requested_by = request.user
        correction.created_by = actor_label
        correction.updated_by = actor_label
        correction.save()

        create_employee_history(
            employee=employee,
            title=f"Attendance correction requested: {attendance_entry.attendance_date}",
            description=build_attendance_correction_summary(correction),
            event_type=EmployeeHistory.EVENT_STATUS,
            created_by=actor_label,
            is_system_generated=True,
            event_date=attendance_entry.attendance_date,
        )
        messages.success(request, "Attendance correction request created successfully.")
    else:
        first_error = "Please review the attendance correction form and try again."
        if form.errors:
            first_field_errors = next(iter(form.errors.values()))
            if first_field_errors:
                first_error = first_field_errors[0]
        messages.error(request, first_error)
        separator = "&" if "?" in next_url else "?"
        next_url = f"{next_url}{separator}correct={attendance_entry.pk}"

    return redirect(next_url)


@login_required
@require_POST
def employee_attendance_correction_apply(request, correction_pk):
    correction = get_object_or_404(
        EmployeeAttendanceCorrection.objects.select_related("employee", "linked_attendance"),
        pk=correction_pk,
    )
    employee = correction.employee

    if not can_manage_attendance_records(request.user):
        return deny_employee_access(
            request,
            "You do not have permission to apply attendance corrections.",
            employee=employee,
        )

    next_url = (request.POST.get("next") or reverse("employees:attendance_management")).strip()

    if correction.status != EmployeeAttendanceCorrection.STATUS_PENDING:
        messages.error(request, "Only pending attendance corrections can be applied.")
        return redirect(next_url)

    attendance_entry = correction.linked_attendance
    actor_label = get_actor_label(request.user)
    review_notes = (request.POST.get("review_notes") or "").strip()

    attendance_entry.day_status = correction.requested_day_status
    attendance_entry.clock_in_time = correction.requested_clock_in_time
    attendance_entry.clock_out_time = correction.requested_clock_out_time
    attendance_entry.scheduled_hours = correction.requested_scheduled_hours
    attendance_entry.late_minutes = correction.requested_late_minutes
    attendance_entry.early_departure_minutes = correction.requested_early_departure_minutes
    attendance_entry.overtime_minutes = correction.requested_overtime_minutes
    attendance_entry.notes = correction.requested_notes
    attendance_entry.source = EmployeeAttendanceLedger.SOURCE_MANUAL
    attendance_entry.updated_by = actor_label
    attendance_entry.save()

    correction.status = EmployeeAttendanceCorrection.STATUS_APPLIED
    correction.review_notes = review_notes
    correction.reviewed_by = request.user
    correction.applied_at = timezone.now()
    correction.updated_by = actor_label
    correction.save()

    create_employee_history(
        employee=employee,
        title=f"Attendance correction applied: {attendance_entry.attendance_date}",
        description=build_attendance_correction_summary(correction),
        event_type=EmployeeHistory.EVENT_STATUS,
        created_by=actor_label,
        is_system_generated=True,
        event_date=attendance_entry.attendance_date,
    )

    messages.success(request, "Attendance correction applied successfully.")
    return redirect(next_url)


@login_required
@require_POST
def employee_attendance_correction_reject(request, correction_pk):
    correction = get_object_or_404(
        EmployeeAttendanceCorrection.objects.select_related("employee", "linked_attendance"),
        pk=correction_pk,
    )
    employee = correction.employee

    if not can_manage_attendance_records(request.user):
        return deny_employee_access(
            request,
            "You do not have permission to reject attendance corrections.",
            employee=employee,
        )

    next_url = (request.POST.get("next") or reverse("employees:attendance_management")).strip()

    if correction.status != EmployeeAttendanceCorrection.STATUS_PENDING:
        messages.error(request, "Only pending attendance corrections can be rejected.")
        return redirect(next_url)

    correction.status = EmployeeAttendanceCorrection.STATUS_REJECTED
    correction.review_notes = (request.POST.get("review_notes") or "").strip()
    correction.reviewed_by = request.user
    correction.updated_by = get_actor_label(request.user)
    correction.save()

    create_employee_history(
        employee=employee,
        title=f"Attendance correction rejected: {correction.linked_attendance.attendance_date}",
        description=build_attendance_correction_summary(correction),
        event_type=EmployeeHistory.EVENT_STATUS,
        created_by=get_actor_label(request.user),
        is_system_generated=True,
        event_date=correction.linked_attendance.attendance_date,
    )

    messages.success(request, "Attendance correction rejected successfully.")
    return redirect(next_url)


def build_attendance_history_management_context(request, *, supervisor_history_only=False):
    filter_state = build_attendance_management_filter_state(request, user=request.user)
    scoped_employee_queryset = get_employee_directory_queryset_for_user(
        request.user,
        Employee.objects.select_related(
            "company",
            "branch",
            "department",
            "section",
            "job_title",
        ).all(),
    )

    attendance_queryset = EmployeeAttendanceLedger.objects.select_related(
        "employee",
        "employee__company",
        "employee__branch",
        "employee__department",
        "employee__section",
        "employee__job_title",
        "linked_leave",
        "linked_action_record",
    ).filter(employee__in=scoped_employee_queryset)

    if filter_state["search_value"]:
        search_value = filter_state["search_value"]
        attendance_queryset = attendance_queryset.filter(
            Q(employee__full_name__icontains=search_value)
            | Q(employee__employee_id__icontains=search_value)
            | Q(employee__email__icontains=search_value)
            | Q(notes__icontains=search_value)
            | Q(created_by__icontains=search_value)
            | Q(updated_by__icontains=search_value)
        )

    if filter_state["employee"]:
        attendance_queryset = attendance_queryset.filter(employee=filter_state["employee"])
    if filter_state["company"]:
        attendance_queryset = attendance_queryset.filter(employee__company=filter_state["company"])
    if filter_state["branch"]:
        attendance_queryset = attendance_queryset.filter(employee__branch=filter_state["branch"])
    if filter_state["department"]:
        attendance_queryset = attendance_queryset.filter(employee__department=filter_state["department"])
    if filter_state["section"]:
        attendance_queryset = attendance_queryset.filter(employee__section=filter_state["section"])
    if filter_state["day_status"]:
        attendance_queryset = attendance_queryset.filter(day_status=filter_state["day_status"])
    if filter_state["start_date"]:
        attendance_queryset = attendance_queryset.filter(attendance_date__gte=filter_state["start_date"])
    if filter_state["end_date"]:
        attendance_queryset = attendance_queryset.filter(attendance_date__lte=filter_state["end_date"])

    attendance_queryset = attendance_queryset.order_by("-attendance_date", "employee__full_name", "-id")
    attendance_entries = list(attendance_queryset)
    attendance_summary = build_attendance_summary(attendance_entries)
    pending_event_queryset = EmployeeAttendanceEvent.objects.select_related(
        "employee",
        "employee__company",
        "employee__branch",
        "employee__department",
        "employee__section",
        "employee__job_title",
    ).filter(
        employee__in=scoped_employee_queryset,
        synced_ledger__isnull=True,
    )

    if filter_state["search_value"]:
        search_value = filter_state["search_value"]
        pending_event_queryset = pending_event_queryset.filter(
            Q(employee__full_name__icontains=search_value)
            | Q(employee__employee_id__icontains=search_value)
            | Q(employee__email__icontains=search_value)
            | Q(notes__icontains=search_value)
        )

    if filter_state["employee"]:
        pending_event_queryset = pending_event_queryset.filter(employee=filter_state["employee"])
    if filter_state["company"]:
        pending_event_queryset = pending_event_queryset.filter(employee__company=filter_state["company"])
    if filter_state["branch"]:
        pending_event_queryset = pending_event_queryset.filter(employee__branch=filter_state["branch"])
    if filter_state["department"]:
        pending_event_queryset = pending_event_queryset.filter(employee__department=filter_state["department"])
    if filter_state["section"]:
        pending_event_queryset = pending_event_queryset.filter(employee__section=filter_state["section"])
    if filter_state["day_status"] and filter_state["day_status"] != EmployeeAttendanceLedger.DAY_STATUS_PRESENT:
        pending_event_queryset = pending_event_queryset.none()
    if filter_state["start_date"]:
        pending_event_queryset = pending_event_queryset.filter(attendance_date__gte=filter_state["start_date"])
    if filter_state["end_date"]:
        pending_event_queryset = pending_event_queryset.filter(attendance_date__lte=filter_state["end_date"])

    pending_event_entries = list(
        pending_event_queryset.order_by("-attendance_date", "employee__full_name", "-id")
    )
    attendance_display_records = sorted(
        [*attendance_entries, *pending_event_entries],
        key=lambda entry: (
            -(entry.attendance_date.toordinal() if entry.attendance_date else 0),
            (entry.employee.full_name or "").lower(),
            -(entry.pk or 0),
        ),
    )

    snapshot_date = filter_state["end_date"] or filter_state["start_date"] or timezone.localdate()
    snapshot_is_single_day = bool(
        filter_state["start_date"]
        and filter_state["end_date"]
        and filter_state["start_date"] == filter_state["end_date"]
    )
    if snapshot_is_single_day:
        attendance_snapshot_note = "Daily attendance snapshot for the selected day."
    elif filter_state["start_date"] or filter_state["end_date"]:
        attendance_snapshot_note = "Daily attendance snapshot based on the end date of the current filtered period."
    else:
        attendance_snapshot_note = "Daily attendance snapshot for today when no fixed date range is selected."

    snapshot_employee_queryset = get_employee_directory_queryset_for_user(
        request.user,
        Employee.objects.select_related(
            "company",
            "branch",
            "department",
            "section",
            "job_title",
        ).all(),
    )

    if filter_state["search_value"]:
        snapshot_search_value = filter_state["search_value"]
        snapshot_employee_queryset = snapshot_employee_queryset.filter(
            Q(full_name__icontains=snapshot_search_value)
            | Q(employee_id__icontains=snapshot_search_value)
            | Q(email__icontains=snapshot_search_value)
            | Q(phone__icontains=snapshot_search_value)
        )

    if filter_state["employee"]:
        snapshot_employee_queryset = snapshot_employee_queryset.filter(pk=filter_state["employee"].pk)
    if filter_state["company"]:
        snapshot_employee_queryset = snapshot_employee_queryset.filter(company=filter_state["company"])
    if filter_state["branch"]:
        snapshot_employee_queryset = snapshot_employee_queryset.filter(branch=filter_state["branch"])
    if filter_state["department"]:
        snapshot_employee_queryset = snapshot_employee_queryset.filter(department=filter_state["department"])
    if filter_state["section"]:
        snapshot_employee_queryset = snapshot_employee_queryset.filter(section=filter_state["section"])

    snapshot_employee_queryset = snapshot_employee_queryset.filter(is_active=True).filter(
        Q(hire_date__isnull=True) | Q(hire_date__lte=snapshot_date)
    )

    snapshot_scope_employees = list(snapshot_employee_queryset.order_by("full_name", "employee_id"))
    snapshot_scope_employee_ids = [employee.pk for employee in snapshot_scope_employees]

    snapshot_recorded_queryset = EmployeeAttendanceLedger.objects.select_related(
        "employee",
        "employee__branch",
        "employee__department",
        "employee__section",
        "employee__job_title",
    ).filter(
        employee_id__in=snapshot_scope_employee_ids,
        attendance_date=snapshot_date,
    )
    snapshot_recorded_entries = list(
        snapshot_recorded_queryset.order_by("employee__full_name", "employee__employee_id", "-id")
    )
    snapshot_recorded_employee_ids = {entry.employee_id for entry in snapshot_recorded_entries}
    snapshot_pending_event_employee_ids = set(
        EmployeeAttendanceEvent.objects.filter(
            employee_id__in=snapshot_scope_employee_ids,
            attendance_date=snapshot_date,
            synced_ledger__isnull=True,
        ).values_list("employee_id", flat=True)
    )
    snapshot_recorded_employee_ids.update(snapshot_pending_event_employee_ids)

    snapshot_unrecorded_employees = [
        employee for employee in snapshot_scope_employees if employee.pk not in snapshot_recorded_employee_ids
    ]

    approved_leave_ids = set(
        EmployeeLeave.objects.filter(
            employee_id__in=[employee.pk for employee in snapshot_unrecorded_employees],
            status=EmployeeLeave.STATUS_APPROVED,
            start_date__lte=snapshot_date,
            end_date__gte=snapshot_date,
        ).values_list("employee_id", flat=True)
    )

    policy_weekly_off_ids = set()
    policy_holiday_ids = set()
    if is_policy_holiday(snapshot_date):
        policy_holiday_ids = {employee.pk for employee in snapshot_unrecorded_employees}
    elif is_policy_weekly_off_day(snapshot_date):
        policy_weekly_off_ids = {employee.pk for employee in snapshot_unrecorded_employees}

    attendance_snapshot_missing_employees = [
        employee
        for employee in snapshot_unrecorded_employees
        if employee.pk not in approved_leave_ids
        and employee.pk not in policy_weekly_off_ids
        and employee.pk not in policy_holiday_ids
    ]
    attendance_snapshot_leave_covered_employees = [
        employee for employee in snapshot_unrecorded_employees if employee.pk in approved_leave_ids
    ]
    attendance_snapshot_weekly_off_employees = [
        employee
        for employee in snapshot_unrecorded_employees
        if employee.pk in policy_weekly_off_ids and employee.pk not in approved_leave_ids
    ]
    attendance_snapshot_holiday_employees = [
        employee
        for employee in snapshot_unrecorded_employees
        if employee.pk in policy_holiday_ids and employee.pk not in approved_leave_ids
    ]

    paginator = Paginator(attendance_display_records, 25)
    page_obj = paginator.get_page(request.GET.get("page"))
    pagination_items = build_attendance_history_pagination(page_obj)

    correction_queryset = EmployeeAttendanceCorrection.objects.select_related(
        "employee",
        "linked_attendance",
        "requested_by",
        "reviewed_by",
    ).filter(employee__in=scoped_employee_queryset)

    if filter_state["search_value"]:
        search_value = filter_state["search_value"]
        correction_queryset = correction_queryset.filter(
            Q(employee__full_name__icontains=search_value)
            | Q(employee__employee_id__icontains=search_value)
            | Q(request_reason__icontains=search_value)
            | Q(requested_notes__icontains=search_value)
            | Q(review_notes__icontains=search_value)
        )

    if filter_state["employee"]:
        correction_queryset = correction_queryset.filter(employee=filter_state["employee"])
    if filter_state["company"]:
        correction_queryset = correction_queryset.filter(employee__company=filter_state["company"])
    if filter_state["branch"]:
        correction_queryset = correction_queryset.filter(employee__branch=filter_state["branch"])
    if filter_state["department"]:
        correction_queryset = correction_queryset.filter(employee__department=filter_state["department"])
    if filter_state["section"]:
        correction_queryset = correction_queryset.filter(employee__section=filter_state["section"])
    if filter_state["day_status"]:
        correction_queryset = correction_queryset.filter(requested_day_status=filter_state["day_status"])
    if filter_state["start_date"]:
        correction_queryset = correction_queryset.filter(linked_attendance__attendance_date__gte=filter_state["start_date"])
    if filter_state["end_date"]:
        correction_queryset = correction_queryset.filter(linked_attendance__attendance_date__lte=filter_state["end_date"])

    correction_queryset = correction_queryset.order_by("-created_at", "-id")
    correction_records = list(correction_queryset[:50]) if not supervisor_history_only else []

    selected_attendance_record = None
    correction_form = None
    correction_target = request.GET.get("correct", "").strip()
    if correction_target and not supervisor_history_only:
        try:
            target_pk = int(correction_target)
        except ValueError:
            target_pk = None
        if target_pk:
            selected_attendance_record = attendance_queryset.filter(pk=target_pk).first()
            if selected_attendance_record:
                correction_form = EmployeeAttendanceCorrectionForm(
                    attendance_entry=selected_attendance_record,
                    initial={
                        "requested_day_status": selected_attendance_record.day_status,
                        "requested_clock_in_time": selected_attendance_record.clock_in_time,
                        "requested_clock_out_time": selected_attendance_record.clock_out_time,
                        "requested_scheduled_hours": selected_attendance_record.scheduled_hours,
                        "requested_late_minutes": selected_attendance_record.late_minutes,
                        "requested_early_departure_minutes": selected_attendance_record.early_departure_minutes,
                        "requested_overtime_minutes": selected_attendance_record.overtime_minutes,
                        "requested_notes": selected_attendance_record.notes,
                    },
                )

    querystring_data = request.GET.copy()
    querystring_data.pop("page", None)
    action_querystring_data = querystring_data.copy()
    action_querystring_data.pop("correct", None)
    attendance_management_querystring = querystring_data.urlencode()
    attendance_management_base_querystring = action_querystring_data.urlencode()
    attendance_route_name = (
        "employees:supervisor_attendance_history"
        if supervisor_history_only
        else "employees:attendance_management"
    )
    attendance_management_base_url = reverse(attendance_route_name)
    if attendance_management_base_querystring:
        attendance_management_base_url = (
            f"{attendance_management_base_url}?{attendance_management_base_querystring}"
        )

    scoped_branch = get_user_scope_branch(request.user)
    page_title = "Attendance History"
    page_subtitle = (
        "Management attendance history for filtering, auditing, and reviewing all employee attendance records inside your existing management scope."
    )
    ledger_subtitle = "Filtered company-wide attendance records with quick access back to each employee profile."
    empty_message = "Adjust the filters or start creating attendance ledger entries from employee profiles."
    back_button_label = "Back to Directory"

    if supervisor_history_only and scoped_branch:
        page_title = "Team Attendance History"
        page_subtitle = (
            f"Supervisor attendance history for {scoped_branch.name}. Only team members inside your current supervisor scope appear here."
        )
        ledger_subtitle = "Branch-scoped attendance history with click-only detail sections, reduced initial load, and no management-wide controls."
        empty_message = "No attendance history matched the current team filters."
        back_button_label = "Back to Team Directory"

    context = {
        "attendance_page_obj": page_obj,
        "attendance_records": page_obj.object_list,
        "attendance_pagination_items": pagination_items,
        "attendance_filter_form": filter_state["form"],
        "attendance_period_label": filter_state["period_label"],
        "attendance_filter_applied": filter_state["is_applied"],
        "attendance_management_querystring": attendance_management_querystring,
        "attendance_management_base_querystring": attendance_management_base_querystring,
        "attendance_management_base_url": attendance_management_base_url,
        "attendance_employee_count": len(
            {
                *[entry.employee_id for entry in attendance_entries],
                *[entry.employee_id for entry in pending_event_entries],
            }
        ),
        "attendance_day_status_choices": EmployeeAttendanceLedger.DAY_STATUS_CHOICES,
        "selected_attendance_record": selected_attendance_record,
        "attendance_correction_form": correction_form,
        "attendance_correction_records": correction_records,
        "attendance_correction_pending_count": correction_queryset.filter(status=EmployeeAttendanceCorrection.STATUS_PENDING).count() if not supervisor_history_only else 0,
        "attendance_correction_applied_count": correction_queryset.filter(status=EmployeeAttendanceCorrection.STATUS_APPLIED).count() if not supervisor_history_only else 0,
        "attendance_correction_rejected_count": correction_queryset.filter(status=EmployeeAttendanceCorrection.STATUS_REJECTED).count() if not supervisor_history_only else 0,
        "attendance_snapshot_date": snapshot_date,
        "attendance_snapshot_note": attendance_snapshot_note,
        "attendance_snapshot_scope_count": len(snapshot_scope_employees),
        "attendance_snapshot_recorded_count": len(snapshot_recorded_employee_ids),
        "attendance_snapshot_missing_count": len(attendance_snapshot_missing_employees),
        "attendance_snapshot_leave_covered_count": len(attendance_snapshot_leave_covered_employees),
        "attendance_snapshot_weekly_off_count": len(attendance_snapshot_weekly_off_employees),
        "attendance_snapshot_holiday_count": len(attendance_snapshot_holiday_employees),
        "attendance_snapshot_missing_employees": attendance_snapshot_missing_employees[:12],
        "attendance_snapshot_missing_more_count": max(len(attendance_snapshot_missing_employees) - 12, 0),
        "attendance_snapshot_is_single_day": snapshot_is_single_day,
        "attendance_live_open_count": len(pending_event_entries),
        "half_day_attendance_count": sum(1 for entry in attendance_entries if entry.day_status == EmployeeAttendanceLedger.DAY_STATUS_PRESENT and entry.worked_hours and entry.worked_hours < entry.scheduled_hours),
        "early_exit_flag_count": sum(1 for entry in attendance_entries if (entry.early_departure_minutes or 0) > 0),
        "overtime_ready_count": sum(1 for entry in attendance_entries if (entry.overtime_minutes or 0) > 0),
        "attendance_history_is_supervisor_scope": supervisor_history_only,
        "attendance_history_page_title": page_title,
        "attendance_history_page_subtitle": page_subtitle,
        "attendance_history_snapshot_title": "Team Snapshot" if supervisor_history_only else "Daily Attendance Snapshot",
        "attendance_history_ledger_subtitle": ledger_subtitle,
        "attendance_history_back_button_label": back_button_label,
        "attendance_history_empty_message": empty_message,
        "attendance_history_can_request_correction": not supervisor_history_only,
        "attendance_history_can_view_corrections": not supervisor_history_only,
        "attendance_history_route_name": attendance_route_name,
        "attendance_history_show_compact_summary": supervisor_history_only,
    }
    context.update(attendance_summary)

    return context


@login_required
def attendance_management(request):
    if not can_view_attendance_management(request.user):
        return deny_employee_access(
            request,
            "You do not have permission to access attendance management.",
        )

    context = build_attendance_history_management_context(request, supervisor_history_only=False)
    return render(request, "employees/attendance_management.html", context)


@login_required
def supervisor_attendance_history(request):
    if not is_branch_scoped_supervisor(request.user):
        return deny_employee_access(
            request,
            "You do not have permission to access team attendance history.",
        )

    context = build_attendance_history_management_context(request, supervisor_history_only=True)
    return render(request, "employees/attendance_management.html", context)


class ActionCenterEmployeeProfileForm(forms.ModelForm):
    class Meta:
        model = Employee
        fields = [
            "full_name",
            "photo",
            "email",
            "phone",
            "birth_date",
            "marital_status",
            "nationality",
            "hire_date",
            "passport_reference_number",
            "passport_issue_date",
            "passport_expiry_date",
            "civil_id_reference_number",
            "civil_id_issue_date",
            "civil_id_expiry_date",
            "salary",
            "notes",
        ]
        widgets = {
            "birth_date": forms.DateInput(format="%Y-%m-%d", attrs={"type": "date"}),
            "hire_date": forms.DateInput(format="%Y-%m-%d", attrs={"type": "date"}),
            "passport_issue_date": forms.DateInput(format="%Y-%m-%d", attrs={"type": "date"}),
            "passport_expiry_date": forms.DateInput(format="%Y-%m-%d", attrs={"type": "date"}),
            "civil_id_issue_date": forms.DateInput(format="%Y-%m-%d", attrs={"type": "date"}),
            "civil_id_expiry_date": forms.DateInput(format="%Y-%m-%d", attrs={"type": "date"}),
            "notes": forms.Textarea(attrs={"rows": 4}),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for date_field_name in [
            "birth_date",
            "hire_date",
            "passport_issue_date",
            "passport_expiry_date",
            "civil_id_issue_date",
            "civil_id_expiry_date",
        ]:
            if date_field_name in self.fields:
                self.fields[date_field_name].input_formats = ["%Y-%m-%d"]

        for field_name, field in self.fields.items():
            widget = field.widget
            if isinstance(widget, forms.CheckboxInput):
                widget.attrs["class"] = "form-check-input"
            elif isinstance(widget, forms.FileInput):
                widget.attrs["class"] = "form-control"
            else:
                existing = widget.attrs.get("class", "")
                widget.attrs["class"] = f"{existing} form-control".strip()

    def clean(self):
        cleaned_data = super().clean()

        birth_date = cleaned_data.get("birth_date")
        passport_issue_date = cleaned_data.get("passport_issue_date")
        passport_expiry_date = cleaned_data.get("passport_expiry_date")
        civil_id_issue_date = cleaned_data.get("civil_id_issue_date")
        civil_id_expiry_date = cleaned_data.get("civil_id_expiry_date")

        if birth_date and birth_date > timezone.localdate():
            self.add_error("birth_date", "Birth date cannot be in the future.")

        if passport_issue_date and passport_expiry_date and passport_issue_date > passport_expiry_date:
            self.add_error(
                "passport_expiry_date",
                "Passport expiry date must be on or after the passport issue date.",
            )

        if civil_id_issue_date and civil_id_expiry_date and civil_id_issue_date > civil_id_expiry_date:
            self.add_error(
                "civil_id_expiry_date",
                "Civil ID expiry date must be on or after the Civil ID issue date.",
            )

        return cleaned_data



class EmployeeInformationModalForm(ActionCenterEmployeeProfileForm):
    class Meta(ActionCenterEmployeeProfileForm.Meta):
        fields = [
            "full_name",
            "photo",
            "email",
            "phone",
            "birth_date",
            "marital_status",
            "nationality",
            "hire_date",
            "salary",
        ]


class EmployeeIdentityModalForm(ActionCenterEmployeeProfileForm):
    class Meta(ActionCenterEmployeeProfileForm.Meta):
        fields = [
            "passport_reference_number",
            "passport_issue_date",
            "passport_expiry_date",
            "civil_id_reference_number",
            "civil_id_issue_date",
            "civil_id_expiry_date",
        ]


def build_employee_payroll_modal_summary(old_profile, new_profile):
    changes = []
    tracked_fields = [
        ("company", "Payroll company"),
        ("base_salary", "Base salary"),
        ("housing_allowance", "Housing allowance"),
        ("transport_allowance", "Transport allowance"),
        ("fixed_deduction", "Fixed deduction"),
        ("bank_name", "Bank name"),
        ("iban", "IBAN"),
        ("status", "Payroll status"),
    ]

    for field_name, label in tracked_fields:
        old_value = getattr(old_profile, field_name, None) if old_profile else None
        new_value = getattr(new_profile, field_name, None)
        if old_value != new_value:
            changes.append(
                f"{label} changed from {format_history_value(old_value)} to {format_history_value(new_value)}."
            )

    return " ".join(changes) if changes else "Payroll profile details were updated from the employee profile."


def build_employee_information_modal_summary(old_employee, new_employee):
    changes = []
    tracked_fields = [
        ("full_name", "Full name"),
        ("email", "Email"),
        ("phone", "Phone"),
        ("birth_date", "Birth date"),
        ("marital_status", "Marital status"),
        ("nationality", "Nationality"),
        ("hire_date", "Hire date"),
        ("salary", "Salary"),
    ]

    for field_name, label in tracked_fields:
        old_value = getattr(old_employee, field_name)
        new_value = getattr(new_employee, field_name)
        if old_value != new_value:
            changes.append(
                f"{label} changed from {format_history_value(old_value)} to {format_history_value(new_value)}."
            )

    if getattr(old_employee, "photo", None) != getattr(new_employee, "photo", None):
        changes.append("Profile photo was updated.")

    return " ".join(changes) if changes else "Employee information was updated from the profile modal."


def build_employee_identity_modal_summary(old_employee, new_employee):
    changes = []
    tracked_fields = [
        ("passport_reference_number", "Passport reference number"),
        ("passport_issue_date", "Passport issue date"),
        ("passport_expiry_date", "Passport expiry date"),
        ("civil_id_reference_number", "Civil ID reference number"),
        ("civil_id_issue_date", "Civil ID issue date"),
        ("civil_id_expiry_date", "Civil ID expiry date"),
    ]

    for field_name, label in tracked_fields:
        old_value = getattr(old_employee, field_name)
        new_value = getattr(new_employee, field_name)
        if old_value != new_value:
            changes.append(
                f"{label} changed from {format_history_value(old_value)} to {format_history_value(new_value)}."
            )

    return " ".join(changes) if changes else "Passport and Civil ID details were updated from the profile modal."


def render_employee_detail_with_modal_forms(request, employee, **kwargs):
    detail_view = EmployeeDetailView()
    detail_view.request = request
    detail_view.object = employee
    detail_view.kwargs = {"pk": employee.pk}
    detail_view.args = ()
    context = detail_view.get_context_data(**kwargs)
    return render(request, detail_view.template_name, context)



@login_required
@require_POST
def employee_profile_payroll_information_update(request, pk):
    employee = get_object_or_404(Employee, pk=pk)

    if not can_create_or_edit_employees(request.user):
        return deny_employee_access(
            request,
            "You do not have permission to edit payroll details for this employee.",
            employee=employee,
        )

    PayrollProfile = apps.get_model("payroll", "PayrollProfile")
    existing_profile = PayrollProfile.objects.filter(employee=employee).first()
    original_snapshot = PayrollProfile.objects.get(pk=existing_profile.pk) if existing_profile else None
    form = PayrollProfileForm(request.POST, instance=existing_profile, employee=employee)

    if form.is_valid():
        payroll_profile = form.save(commit=False)
        payroll_profile.employee = employee
        if not payroll_profile.company_id and employee.company_id:
            payroll_profile.company_id = employee.company_id
        payroll_profile.save()
        create_employee_history(
            employee=employee,
            title="Payroll profile updated",
            description=build_employee_payroll_modal_summary(original_snapshot, payroll_profile),
            event_type=EmployeeHistory.EVENT_PROFILE,
            created_by=get_actor_label(request.user),
            is_system_generated=True,
            event_date=timezone.localdate(),
        )
        messages.success(request, "Payroll profile updated successfully.")
        return redirect(f"{reverse('employees:employee_detail', kwargs={'pk': employee.pk})}#employee-payroll-section")

    messages.error(request, "Please correct the payroll profile fields and try again.")
    return render_employee_detail_with_modal_forms(
        request,
        employee,
        employee_payroll_profile_form=form,
        active_profile_modal="payroll_information",
    )


@login_required
@require_POST
def employee_profile_employee_information_update(request, pk):
    employee = get_object_or_404(Employee, pk=pk)

    if not can_create_or_edit_employees(request.user):
        return deny_employee_access(
            request,
            "You do not have permission to edit this employee section.",
            employee=employee,
        )

    original_employee = Employee.objects.get(pk=employee.pk)
    form = EmployeeInformationModalForm(request.POST, request.FILES, instance=employee)

    if form.is_valid():
        form.save()
        create_employee_history(
            employee=employee,
            title="Employee information updated",
            description=build_employee_information_modal_summary(original_employee, employee),
            event_type=EmployeeHistory.EVENT_PROFILE,
            created_by=get_actor_label(request.user),
            is_system_generated=True,
            event_date=timezone.localdate(),
        )
        messages.success(request, "Employee information updated successfully.")
        return redirect(f"{reverse('employees:employee_detail', kwargs={'pk': employee.pk})}#employee-information-section")

    messages.error(request, "Please correct the employee information fields and try again.")
    return render_employee_detail_with_modal_forms(
        request,
        employee,
        employee_information_modal_form=form,
        active_profile_modal="employee_information",
    )


@login_required
@require_POST
def employee_profile_identity_information_update(request, pk):
    employee = get_object_or_404(Employee, pk=pk)

    if not can_create_or_edit_employees(request.user):
        return deny_employee_access(
            request,
            "You do not have permission to edit this employee section.",
            employee=employee,
        )

    original_employee = Employee.objects.get(pk=employee.pk)
    form = EmployeeIdentityModalForm(request.POST, instance=employee)

    if form.is_valid():
        form.save()
        create_employee_history(
            employee=employee,
            title="Identity information updated",
            description=build_employee_identity_modal_summary(original_employee, employee),
            event_type=EmployeeHistory.EVENT_PROFILE,
            created_by=get_actor_label(request.user),
            is_system_generated=True,
            event_date=timezone.localdate(),
        )
        messages.success(request, "Passport and Civil ID details updated successfully.")
        return redirect(f"{reverse('employees:employee_detail', kwargs={'pk': employee.pk})}#employee-information-section")

    messages.error(request, "Please correct the passport and Civil ID fields and try again.")
    return render_employee_detail_with_modal_forms(
        request,
        employee,
        identity_information_modal_form=form,
        active_profile_modal="identity_information",
    )



def employee_admin_action_center(request):
    if not is_management_user(request.user):
        raise PermissionDenied("You do not have permission to access the employee action center.")

    today = timezone.localdate()
    current_user = request.user

    employee_queryset = get_employee_directory_queryset_for_user(
        current_user,
        Employee.objects.select_related(
            "company",
            "branch",
            "department",
            "section",
            "job_title",
        ).all(),
    )

    search_query = ((request.POST.get("search") if request.method == "POST" else request.GET.get("search")) or "").strip()
    selected_employee = None
    selected_employee_param = ((request.POST.get("employee") if request.method == "POST" else request.GET.get("employee")) or "").strip()
    current_page_param = ((request.POST.get("page") if request.method == "POST" else request.GET.get("page")) or "1").strip()

    employee_picker_queryset = employee_queryset.order_by("full_name", "employee_id")
    if search_query:
        employee_picker_queryset = employee_picker_queryset.filter(
            Q(full_name__icontains=search_query)
            | Q(employee_id__icontains=search_query)
        )

    employee_picker_paginator = Paginator(employee_picker_queryset, 6)
    current_page_number = 1
    if current_page_param.isdigit():
        current_page_number = max(1, int(current_page_param))

    employee_picker_page = employee_picker_paginator.get_page(current_page_number)
    quick_employee_results = list(employee_picker_page.object_list)

    if selected_employee_param.isdigit():
        selected_employee = employee_queryset.filter(pk=int(selected_employee_param)).first()
    elif request.method != "POST" and employee_picker_paginator.count == 1:
        selected_employee = employee_picker_queryset.first()

    action_center_action_form = EmployeeActionRecordForm()
    action_center_required_submission_form = EmployeeRequiredSubmissionCreateForm()
    action_center_leave_form = EmployeeLeaveForm()
    action_center_attendance_form = EmployeeAttendanceLedgerForm(employee=selected_employee) if selected_employee else EmployeeAttendanceLedgerForm()
    action_center_transfer_form = EmployeeTransferForm(instance=selected_employee) if selected_employee else EmployeeTransferForm()
    action_center_profile_form = ActionCenterEmployeeProfileForm(instance=selected_employee) if selected_employee else ActionCenterEmployeeProfileForm()

    def build_action_center_redirect(employee_obj=None, search_value="", page_number=None):
        employee_obj = employee_obj or selected_employee
        params = []
        if search_value:
            params.append(f"search={search_value}")
        if employee_obj:
            params.append(f"employee={employee_obj.pk}")
        resolved_page = page_number or employee_picker_page.number
        if resolved_page:
            params.append(f"page={resolved_page}")
        base_url = reverse("employees:employee_admin_action_center")
        return f"{base_url}?{'&'.join(params)}" if params else base_url

    if request.method == "POST":
        if not selected_employee:
            messages.error(request, "Select an employee first before submitting an Action Center form.")
            return redirect(build_action_center_redirect(search_value=search_query))

        action_center_post = (request.POST.get("action_center_post") or "").strip()

        if action_center_post == "status_action":
            if not can_change_employee_status(current_user):
                return deny_employee_access(
                    request,
                    "You do not have permission to change employee status.",
                    employee=selected_employee,
                )

            target_status = (request.POST.get("target_status") or "").strip()
            valid_statuses = {value for value, _label in Employee.EMPLOYMENT_STATUS_CHOICES}
            if target_status not in valid_statuses:
                messages.error(request, "Invalid employee status action.")
            else:
                selected_employee.employment_status = target_status
                selected_employee.is_active = target_status != Employee.EMPLOYMENT_STATUS_INACTIVE
                selected_employee.save(update_fields=["employment_status", "is_active", "updated_at"])
                create_employee_history(
                    employee=selected_employee,
                    title="Employee status updated",
                    description=f"Employee status changed to {selected_employee.get_employment_status_display()} from the Action Center.",
                    event_type=EmployeeHistory.EVENT_STATUS,
                    created_by=get_actor_label(request.user),
                    is_system_generated=True,
                    event_date=timezone.localdate(),
                )
                messages.success(request, "Employee status updated successfully from the Action Center.")
            return redirect(build_action_center_redirect(selected_employee, search_query))

        if action_center_post == "action_record":
            if not can_create_action_records(current_user):
                return deny_employee_access(
                    request,
                    "You do not have permission to create employee action records.",
                    employee=selected_employee,
                )

            action_center_action_form = EmployeeActionRecordForm(request.POST)
            if action_center_action_form.is_valid():
                action_record = action_center_action_form.save(commit=False)
                action_record.employee = selected_employee
                action_record.created_by = get_actor_label(request.user)
                action_record.updated_by = get_actor_label(request.user)
                action_record.save()
                create_employee_history(
                    employee=selected_employee,
                    title=f"Action record created: {action_record.title}",
                    description=(
                        f"Action Type: {action_record.get_action_type_display()}. "
                        f"Status: {action_record.get_status_display()}. "
                        f"Severity: {action_record.get_severity_display()}. "
                        + (f"Description: {action_record.description}" if action_record.description else "")
                    ).strip(),
                    event_type=EmployeeHistory.EVENT_NOTE,
                    created_by=get_actor_label(request.user),
                    is_system_generated=True,
                    event_date=action_record.action_date or timezone.localdate(),
                )
                messages.success(request, "Employee action record created successfully from the Action Center.")
                return redirect(build_action_center_redirect(selected_employee, search_query))

            messages.error(request, "Please review the action record form and try again.")

        elif action_center_post == "required_submission":
            if not can_manage_employee_required_submissions(current_user, selected_employee):
                return deny_employee_access(
                    request,
                    "You do not have permission to create employee required submission requests.",
                    employee=selected_employee,
                )

            action_center_required_submission_form = EmployeeRequiredSubmissionCreateForm(request.POST)
            if action_center_required_submission_form.is_valid():
                submission_request = action_center_required_submission_form.save(commit=False)
                submission_request.employee = selected_employee
                submission_request.created_by = request.user
                submission_request.status = EmployeeRequiredSubmission.STATUS_REQUESTED
                submission_request.reviewed_by = None
                submission_request.review_note = ""
                submission_request.reviewed_at = None
                submission_request.submitted_at = None
                submission_request.save()
                create_employee_history(
                    employee=selected_employee,
                    title=f"Required employee submission requested: {submission_request.title}",
                    description=(
                        f"Request Type: {submission_request.get_request_type_display()}. "
                        f"Priority: {submission_request.get_priority_display()}. "
                        + (
                            f"Due Date: {submission_request.due_date.strftime('%B %d, %Y')}. "
                            if submission_request.due_date else ""
                        )
                        + (f"Instructions: {submission_request.instructions}" if submission_request.instructions else "")
                    ).strip(),
                    event_type=EmployeeHistory.EVENT_DOCUMENT,
                    created_by=get_actor_label(request.user),
                    is_system_generated=True,
                    event_date=timezone.localdate(),
                )
                messages.success(request, "Required employee submission request created successfully from the Action Center.")
                return redirect(build_action_center_redirect(selected_employee, search_query))

            first_error = "Please review the required submission request form and try again."
            if action_center_required_submission_form.errors:
                first_field_errors = next(iter(action_center_required_submission_form.errors.values()))
                if first_field_errors:
                    first_error = first_field_errors[0]
            messages.error(request, first_error)

        elif action_center_post == "leave_request":
            if not can_request_leave(current_user, selected_employee):
                return deny_employee_access(
                    request,
                    "You do not have permission to create leave requests for this employee.",
                    employee=selected_employee,
                )

            action_center_leave_form = EmployeeLeaveForm(request.POST)
            if action_center_leave_form.is_valid():
                leave_record = action_center_leave_form.save(commit=False)
                leave_record.employee = selected_employee
                leave_record.requested_by = request.user
                leave_record.created_by = get_actor_label(request.user)
                leave_record.updated_by = get_actor_label(request.user)
                leave_record.status = EmployeeLeave.STATUS_PENDING
                leave_record.current_stage = EmployeeLeave.STAGE_SUPERVISOR_REVIEW
                leave_record.save()
                create_employee_history(
                    employee=selected_employee,
                    title=f"Leave requested: {leave_record.get_leave_type_display()}",
                    description=build_leave_request_summary(leave_record),
                    event_type=EmployeeHistory.EVENT_STATUS,
                    created_by=get_actor_label(request.user),
                    is_system_generated=True,
                    event_date=leave_record.start_date,
                )
                messages.success(request, "Leave request created successfully from the Action Center.")
                return redirect(build_action_center_redirect(selected_employee, search_query))

            first_error = "Please review the leave request form and try again."
            if action_center_leave_form.errors:
                first_field_errors = next(iter(action_center_leave_form.errors.values()))
                if first_field_errors:
                    first_error = first_field_errors[0]
            messages.error(request, first_error)

        elif action_center_post == "attendance_action":
            if not can_manage_attendance_records(current_user):
                return deny_employee_access(
                    request,
                    "You do not have permission to create attendance ledger entries.",
                    employee=selected_employee,
                )

            action_center_attendance_form = EmployeeAttendanceLedgerForm(request.POST, employee=selected_employee)
            if action_center_attendance_form.is_valid():
                attendance_entry = action_center_attendance_form.save(commit=False)
                attendance_entry.employee = selected_employee
                attendance_entry.source = EmployeeAttendanceLedger.SOURCE_MANUAL
                actor_label = get_actor_label(request.user)
                attendance_entry.created_by = actor_label
                attendance_entry.updated_by = actor_label
                attendance_entry.save()
                create_employee_history(
                    employee=selected_employee,
                    title=f"Attendance ledger entry added: {attendance_entry.attendance_date}",
                    description=build_attendance_ledger_summary(attendance_entry),
                    event_type=EmployeeHistory.EVENT_STATUS,
                    created_by=actor_label,
                    is_system_generated=True,
                    event_date=attendance_entry.attendance_date,
                )
                messages.success(request, "Attendance entry created successfully from the Action Center.")
                return redirect(build_action_center_redirect(selected_employee, search_query))

            first_error = "Please review the attendance form and try again."
            if action_center_attendance_form.errors:
                first_field_errors = next(iter(action_center_attendance_form.errors.values()))
                if first_field_errors:
                    first_error = first_field_errors[0]
            messages.error(request, first_error)

        elif action_center_post == "transfer_action":
            if not can_transfer_employee(current_user):
                return deny_employee_access(
                    request,
                    "You do not have permission to transfer employee placements.",
                    employee=selected_employee,
                )

            original_employee = Employee.objects.get(pk=selected_employee.pk)
            action_center_transfer_form = EmployeeTransferForm(request.POST, instance=selected_employee)
            if action_center_transfer_form.is_valid():
                transferred_employee = action_center_transfer_form.save()
                transfer_note = action_center_transfer_form.cleaned_data.get("notes", "")
                create_employee_history(
                    employee=transferred_employee,
                    title="Employee placement transferred",
                    description=build_employee_transfer_summary(original_employee, transferred_employee, transfer_note=transfer_note),
                    event_type=EmployeeHistory.EVENT_TRANSFER,
                    created_by=get_actor_label(request.user),
                    is_system_generated=True,
                    event_date=timezone.localdate(),
                )
                messages.success(request, "Employee placement updated successfully from the Action Center.")
                return redirect(build_action_center_redirect(transferred_employee, search_query))

            first_error = "Please review the transfer form and try again."
            if action_center_transfer_form.errors:
                first_field_errors = next(iter(action_center_transfer_form.errors.values()))
                if first_field_errors:
                    first_error = first_field_errors[0]
            messages.error(request, first_error)

        elif action_center_post == "profile_update":
            if not can_create_or_edit_employees(current_user):
                return deny_employee_access(
                    request,
                    "You do not have permission to update employee profile details.",
                    employee=selected_employee,
                )

            original_employee = Employee.objects.get(pk=selected_employee.pk)
            action_center_profile_form = ActionCenterEmployeeProfileForm(
                request.POST,
                request.FILES,
                instance=selected_employee,
            )
            if action_center_profile_form.is_valid():
                updated_employee = action_center_profile_form.save()
                create_employee_history(
                    employee=updated_employee,
                    title="Employee profile updated",
                    description=build_employee_change_summary(original_employee, updated_employee),
                    event_type=EmployeeHistory.EVENT_PROFILE,
                    created_by=get_actor_label(request.user),
                    is_system_generated=True,
                    event_date=timezone.localdate(),
                )
                messages.success(request, "Employee profile updated successfully from the Action Center.")
                return redirect(build_action_center_redirect(updated_employee, search_query))

            first_error = "Please review the profile update form and try again."
            if action_center_profile_form.errors:
                first_field_errors = next(iter(action_center_profile_form.errors.values()))
                if first_field_errors:
                    first_error = first_field_errors[0]
            messages.error(request, first_error)

        elif action_center_post:
            messages.error(request, "Unknown Action Center action requested.")
            return redirect(build_action_center_redirect(selected_employee, search_query))

    pending_leave_queryset = (
        get_leave_queryset_for_user(current_user, EmployeeLeave.objects.select_related("employee", "employee__branch"))
        .filter(status=EmployeeLeave.STATUS_PENDING)
        .order_by("-created_at", "-id")
    )
    supervisor_leave_queue_queryset = pending_leave_queryset.filter(
        current_stage=EmployeeLeave.STAGE_SUPERVISOR_REVIEW
    )
    operations_leave_queue_queryset = pending_leave_queryset.filter(
        current_stage=EmployeeLeave.STAGE_OPERATIONS_REVIEW
    )
    hr_leave_queue_queryset = pending_leave_queryset.filter(
        current_stage=EmployeeLeave.STAGE_HR_REVIEW
    )

    required_submission_queryset = EmployeeRequiredSubmission.objects.select_related(
        "employee",
        "employee__branch",
        "employee__department",
        "employee__company",
        "fulfilled_document",
    ).filter(employee__in=employee_queryset).order_by("-updated_at", "-created_at", "-id")
    outstanding_required_submission_queryset = required_submission_queryset.filter(
        status__in=[
            EmployeeRequiredSubmission.STATUS_REQUESTED,
            EmployeeRequiredSubmission.STATUS_NEEDS_CORRECTION,
        ]
    )
    submitted_required_submission_queryset = required_submission_queryset.filter(
        status=EmployeeRequiredSubmission.STATUS_SUBMITTED
    )

    attendance_queryset_today = EmployeeAttendanceLedger.objects.filter(
        employee__in=employee_queryset,
        attendance_date=today,
    )

    correction_queryset = (
        EmployeeAttendanceCorrection.objects.select_related("employee", "linked_attendance")
        .filter(employee__in=employee_queryset)
        .order_by("-created_at", "-id")
    )

    pending_correction_queryset = correction_queryset.filter(
        status=EmployeeAttendanceCorrection.STATUS_PENDING
    )

    id_attention_limit = today + timedelta(days=30)
    expiring_identity_queryset = employee_queryset.filter(
        is_active=True,
    ).filter(
        Q(passport_expiry_date__isnull=False, passport_expiry_date__lte=id_attention_limit)
        | Q(civil_id_expiry_date__isnull=False, civil_id_expiry_date__lte=id_attention_limit)
    ).order_by("passport_expiry_date", "civil_id_expiry_date", "full_name")

    attendance_exception_statuses = {
        EmployeeAttendanceLedger.DAY_STATUS_ABSENT,
        EmployeeAttendanceLedger.DAY_STATUS_PAID_LEAVE,
        EmployeeAttendanceLedger.DAY_STATUS_UNPAID_LEAVE,
        EmployeeAttendanceLedger.DAY_STATUS_SICK_LEAVE,
        EmployeeAttendanceLedger.DAY_STATUS_OTHER,
    }

    attendance_exception_queryset = (
        attendance_queryset_today.select_related("employee", "employee__branch")
        .filter(day_status__in=attendance_exception_statuses)
        .order_by("employee__full_name", "employee__employee_id", "-id")
    )

    absence_action_queryset_today = (
        EmployeeActionRecord.objects.select_related("employee", "employee__branch")
        .filter(
            employee__in=employee_queryset,
            employee__is_active=True,
            action_type=EmployeeActionRecord.ACTION_TYPE_ABSENCE,
            action_date=today,
        )
        .exclude(employee_id__in=attendance_exception_queryset.values_list("employee_id", flat=True))
        .order_by("employee__full_name", "employee__employee_id", "-id")
    )

    supervisor_stage_pending_count = supervisor_leave_queue_queryset.count()
    operations_stage_pending_count = operations_leave_queue_queryset.count()
    hr_stage_pending_count = hr_leave_queue_queryset.count()

    current_leave_review_stage_label = "No active leave review stage"
    my_leave_review_queryset = pending_leave_queryset.none()
    if is_branch_scoped_supervisor(current_user):
        current_leave_review_stage_label = "Supervisor review queue"
        my_leave_review_queryset = supervisor_leave_queue_queryset
    elif is_operations_manager_user(current_user):
        current_leave_review_stage_label = "Operations review queue"
        my_leave_review_queryset = operations_leave_queue_queryset
    elif is_hr_user(current_user) or is_admin_compatible(current_user):
        current_leave_review_stage_label = "HR final review queue"
        my_leave_review_queryset = hr_leave_queue_queryset

    def build_leave_queue_rows(queryset, limit=6):
        rows = []
        for leave_record in queryset[:limit]:
            rows.append(
                {
                    "title": leave_record.employee.full_name,
                    "subtitle": f"{leave_record.employee.employee_id} • {leave_record.get_leave_type_display()}",
                    "meta": f"{leave_record.start_date:%b %d, %Y} → {leave_record.end_date:%b %d, %Y}",
                    "workflow_owner": get_leave_current_stage_owner_label(leave_record),
                    "stage_label": leave_record.get_current_stage_display(),
                    "url": reverse("employees:employee_detail", kwargs={"pk": leave_record.employee.pk}),
                }
            )
        return rows

    def build_required_submission_queue_rows(queryset):
        rows = []
        for submission_request in queryset[:6]:
            rows.append(
                {
                    "title": submission_request.employee.full_name,
                    "subtitle": (
                        f"{submission_request.employee.employee_id} • "
                        f"{submission_request.get_request_type_display()}"
                    ),
                    "meta": (
                        f"Due: {submission_request.due_date:%b %d, %Y}"
                        if submission_request.due_date
                        else f"Status: {submission_request.get_status_display()}"
                    ),
                    "url": reverse("employees:employee_detail", kwargs={"pk": submission_request.employee.pk}),
                }
            )
        return rows

    def build_correction_queue_rows(queryset):
        rows = []
        for correction in queryset[:6]:
            rows.append(
                {
                    "title": correction.employee.full_name,
                    "subtitle": f"{correction.employee.employee_id} • {correction.linked_attendance.attendance_date:%b %d, %Y}",
                    "meta": f"Requested: {correction.get_requested_day_status_display()}",
                    "url": reverse("employees:attendance_management") + f"?correct={correction.linked_attendance.pk}",
                }
            )
        return rows

    def build_identity_attention_rows(queryset):
        rows = []
        for employee in queryset[:6]:
            expiry_values = []
            if employee.passport_expiry_date:
                expiry_values.append(f"Passport: {employee.passport_expiry_date:%b %d, %Y}")
            if employee.civil_id_expiry_date:
                expiry_values.append(f"Civil ID: {employee.civil_id_expiry_date:%b %d, %Y}")
            rows.append(
                {
                    "title": employee.full_name,
                    "subtitle": f"{employee.employee_id} • {employee.branch.name if employee.branch_id else '—'}",
                    "meta": " • ".join(expiry_values) if expiry_values else "Identity date needs review",
                    "url": reverse("employees:employee_detail", kwargs={"pk": employee.pk}),
                }
            )
        return rows

    def build_attendance_exception_rows(attendance_queryset, absence_action_queryset):
        rows = []

        for attendance_entry in attendance_queryset[:6]:
            employee = attendance_entry.employee
            rows.append(
                {
                    "title": employee.full_name,
                    "subtitle": f"{employee.employee_id} • {employee.branch.name if employee.branch_id else '—'}",
                    "meta": f"{attendance_entry.get_day_status_display()} for {attendance_entry.attendance_date:%b %d, %Y}",
                    "url": reverse("employees:employee_detail", kwargs={"pk": employee.pk}),
                }
            )

        remaining_slots = max(0, 6 - len(rows))
        if remaining_slots:
            for action_record in absence_action_queryset[:remaining_slots]:
                employee = action_record.employee
                rows.append(
                    {
                        "title": employee.full_name,
                        "subtitle": f"{employee.employee_id} • {employee.branch.name if employee.branch_id else '—'}",
                        "meta": f"Absence action recorded for {action_record.action_date:%b %d, %Y}",
                        "url": reverse("employees:employee_detail", kwargs={"pk": employee.pk}),
                    }
                )

        return rows

    total_employees = employee_queryset.count()
    active_employees = employee_queryset.filter(is_active=True).count()
    inactive_employees = employee_queryset.filter(is_active=False).count()
    pending_leave_requests = pending_leave_queryset.count()
    outstanding_required_submission_count = outstanding_required_submission_queryset.count()
    submitted_required_submission_count = submitted_required_submission_queryset.count()
    today_attendance_records = attendance_queryset_today.count()
    pending_correction_count = pending_correction_queryset.count()
    expiring_identity_count = expiring_identity_queryset.count()
    attendance_exception_count = attendance_exception_queryset.count() + absence_action_queryset_today.count()

    context = {
        "search_query": search_query,
        "current_picker_page": employee_picker_page.number,
        "employee_picker_page": employee_picker_page,
        "employee_picker_paginator": employee_picker_paginator,
        "quick_employee_results": quick_employee_results,
        "selected_employee": selected_employee,
        "selected_employee_supervisor_display": (
            get_branch_supervisor_display(selected_employee) if selected_employee else ""
        ),
        "total_employees": total_employees,
        "active_employees": active_employees,
        "inactive_employees": inactive_employees,
        "pending_leave_requests": pending_leave_requests,
        "outstanding_required_submission_count": outstanding_required_submission_count,
        "submitted_required_submission_count": submitted_required_submission_count,
        "today_attendance_records": today_attendance_records,
        "pending_correction_count": pending_correction_count,
        "expiring_identity_count": expiring_identity_count,
        "attendance_exception_count": attendance_exception_count,
        "employment_status_choices": Employee.EMPLOYMENT_STATUS_CHOICES,
        "action_center_action_form": action_center_action_form,
        "action_center_required_submission_form": action_center_required_submission_form,
        "action_center_leave_form": action_center_leave_form,
        "action_center_attendance_form": action_center_attendance_form,
        "action_center_transfer_form": action_center_transfer_form,
        "action_center_profile_form": action_center_profile_form,
        "pending_leave_queue": build_leave_queue_rows(pending_leave_queryset),
        "my_leave_review_queue": build_leave_queue_rows(my_leave_review_queryset),
        "supervisor_leave_queue": build_leave_queue_rows(supervisor_leave_queue_queryset),
        "operations_leave_queue": build_leave_queue_rows(operations_leave_queue_queryset),
        "hr_leave_queue": build_leave_queue_rows(hr_leave_queue_queryset),
        "supervisor_stage_pending_count": supervisor_stage_pending_count,
        "operations_stage_pending_count": operations_stage_pending_count,
        "hr_stage_pending_count": hr_stage_pending_count,
        "my_leave_review_count": my_leave_review_queryset.count(),
        "current_leave_review_stage_label": current_leave_review_stage_label,
        "required_submission_queue": build_required_submission_queue_rows(outstanding_required_submission_queryset),
        "submitted_required_submission_queue": build_required_submission_queue_rows(submitted_required_submission_queryset),
        "pending_correction_queue": build_correction_queue_rows(pending_correction_queryset),
        "identity_attention_queue": build_identity_attention_rows(expiring_identity_queryset),
        "attendance_exception_queue": build_attendance_exception_rows(attendance_exception_queryset, absence_action_queryset_today),
        "today_label": today,
        "can_transfer_selected_employee": bool(selected_employee and can_transfer_employee(current_user)),
        "can_edit_selected_employee": bool(selected_employee and can_create_or_edit_employees(current_user)),
    }
    return render(request, "employees/employee_action_center.html", context)



@login_required
def get_departments_by_company(request):
    if not can_access_employee_form_dependencies(request.user):
        return deny_json_access()

    # Shared company assignment rule:
    # departments remain organization-owned records, but employee forms may
    # reuse the same active departments across different selected companies.
    departments = (
        Department.objects.filter(is_active=True)
        .select_related("company")
        .order_by("company__name", "name")
    )
    results = [
        {"id": department.id, "name": department.name}
        for department in departments
    ]

    return JsonResponse({"results": results})


@login_required
def get_branches_by_company(request):
    if not can_access_employee_form_dependencies(request.user):
        return deny_json_access()

    # Shared company assignment rule:
    # branches remain organization-owned records, but employee forms may
    # reuse the same active branches across different selected companies.
    branches = (
        Branch.objects.filter(is_active=True)
        .select_related("company")
        .order_by("company__name", "name")
    )
    results = [
        {"id": branch.id, "name": branch.name}
        for branch in branches
    ]

    return JsonResponse({"results": results})


@login_required
def get_sections_by_department(request):
    if not can_access_employee_form_dependencies(request.user):
        return deny_json_access()

    department_id = request.GET.get("department_id")
    results = []

    if department_id:
        sections = Section.objects.filter(department_id=department_id, is_active=True).order_by("name")
        results = [{"id": section.id, "name": section.name} for section in sections]

    return JsonResponse({"results": results})


@login_required
def get_job_titles_by_context(request):
    if not can_access_employee_form_dependencies(request.user):
        return deny_json_access()

    department_id = request.GET.get("department_id")
    section_id = request.GET.get("section_id")
    results = []

    if department_id:
        job_titles = JobTitle.objects.filter(
            department_id=department_id,
            is_active=True,
        )

        if section_id:
            job_titles = job_titles.filter(Q(section_id=section_id) | Q(section__isnull=True))
        else:
            job_titles = job_titles.filter(section__isnull=True)

        job_titles = job_titles.order_by("name")
        results = [{"id": job_title.id, "name": job_title.name} for job_title in job_titles]

    return JsonResponse({"results": results})



def get_employee_professional_snapshot(employee):
    return {
        "company": getattr(getattr(employee, "company", None), "name", "") or "",
        "department": getattr(getattr(employee, "department", None), "name", "") or "",
        "branch": getattr(getattr(employee, "branch", None), "name", "") or "",
        "section": getattr(getattr(employee, "section", None), "name", "") or "",
        "job_title": getattr(getattr(employee, "job_title", None), "name", "") or "",
        "hire_date": getattr(employee, "hire_date", None),
        "department_manager": get_department_manager_display(employee),
        "branch_supervisor": get_branch_supervisor_display(employee),
        "team_leader": get_team_leader_display(employee),
    }


def build_employee_request_overview(leave_record):
    supporting_documents = list(
        leave_record.supporting_documents.all().order_by("-uploaded_at", "-id")
    )
    latest_related_date = leave_record.updated_at or leave_record.created_at

    for document in supporting_documents:
        document_date = getattr(document, "updated_at", None) or getattr(document, "created_at", None)
        if document_date and (latest_related_date is None or document_date > latest_related_date):
            latest_related_date = document_date

    employee = leave_record.employee
    submitted_at = timezone.localtime(leave_record.created_at) if leave_record.created_at else None
    submitted_date = submitted_at.date() if submitted_at else leave_record.start_date

    return {
        "leave_record": leave_record,
        "workflow_owner_label": get_leave_current_stage_owner_label(leave_record),
        "employee": employee,
        "professional_snapshot": get_employee_professional_snapshot(employee),
        "supporting_documents": supporting_documents,
        "supporting_documents_count": len(supporting_documents),
        "latest_related_date": latest_related_date,
        "submitted_at": submitted_at,
        "submitted_date": submitted_date,
    }


def get_request_week_start(target_date):
    if not target_date:
        return None

    saturday_weekday = 5
    offset = (target_date.weekday() - saturday_weekday) % 7
    return target_date - timedelta(days=offset)


def build_request_overview_groups(request_cards):
    today = timezone.localdate()
    current_week_start = get_request_week_start(today)
    week_map = {}

    for item in request_cards:
        submitted_date = item.get("submitted_date") or today
        week_start = get_request_week_start(submitted_date) or current_week_start
        week_entry = week_map.setdefault(
            week_start,
            {
                "week_start": week_start,
                "week_end": week_start + timedelta(days=6),
                "day_map": {},
                "request_total": 0,
                "document_total": 0,
            },
        )

        day_entry = week_entry["day_map"].setdefault(
            submitted_date,
            {
                "date": submitted_date,
                "items": [],
                "request_total": 0,
                "document_total": 0,
            },
        )

        day_entry["items"].append(item)
        day_entry["request_total"] += 1
        day_entry["document_total"] += item.get("supporting_documents_count", 0)

        week_entry["request_total"] += 1
        week_entry["document_total"] += item.get("supporting_documents_count", 0)

    grouped_weeks = []
    for week_start in sorted(week_map.keys(), reverse=True):
        week_entry = week_map[week_start]
        ordered_days = []

        for request_date in sorted(week_entry["day_map"].keys(), reverse=True):
            day_entry = week_entry["day_map"][request_date]
            ordered_days.append(
                {
                    "date": request_date,
                    "label": request_date.strftime("%A, %B %d, %Y"),
                    "short_label": request_date.strftime("%b %d"),
                    "request_total": day_entry["request_total"],
                    "document_total": day_entry["document_total"],
                    "items": day_entry["items"],
                    "is_today": request_date == today,
                    "is_open": False,
                }
            )

        grouped_weeks.append(
            {
                "key": week_start.isoformat(),
                "week_start": week_entry["week_start"],
                "week_end": week_entry["week_end"],
                "label": f"{week_entry['week_start'].strftime('%b %d, %Y')} → {week_entry['week_end'].strftime('%b %d, %Y')}",
                "request_total": week_entry["request_total"],
                "document_total": week_entry["document_total"],
                "days": ordered_days,
                "is_current_week": week_start == current_week_start,
                "is_open": False,
            }
        )

    return grouped_weeks


@login_required
def employee_requests_overview(request):
    if not can_view_employee_requests_overview(request.user):
        if is_supervisor_user(request.user):
            messages.error(
                request,
                "Supervisor request review requires linking this login account to an employee profile with an assigned branch.",
            )
            return redirect("dashboard_home")
        linked_employee = get_user_employee_profile(request.user)
        if linked_employee:
            messages.error(request, "You do not have permission to access employee requests overview.")
            return redirect("employees:employee_detail", pk=linked_employee.pk)
        raise PermissionDenied("You do not have permission to access employee requests overview.")

    leave_queryset = get_leave_queryset_for_user(
        request.user,
        EmployeeLeave.objects.select_related(
            "employee",
            "employee__company",
            "employee__department",
            "employee__branch",
            "employee__section",
            "employee__job_title",
            "requested_by",
            "approved_by",
            "rejected_by",
            "cancelled_by",
        )
        .prefetch_related(
            Prefetch(
                "supporting_documents",
                queryset=EmployeeDocument.objects.select_related("employee", "linked_leave").order_by("-uploaded_at", "-id"),
            )
        )
        .order_by("-created_at", "-id")
    )

    search_query = request.GET.get("search", "").strip()
    selected_status = request.GET.get("status", "").strip()
    selected_leave_type = request.GET.get("leave_type", "").strip()
    selected_company = request.GET.get("company", "").strip()
    selected_department = request.GET.get("department", "").strip()
    selected_branch = request.GET.get("branch", "").strip()

    if search_query:
        leave_queryset = leave_queryset.filter(
            Q(employee__full_name__icontains=search_query)
            | Q(employee__employee_id__icontains=search_query)
            | Q(employee__email__icontains=search_query)
            | Q(reason__icontains=search_query)
            | Q(approval_note__icontains=search_query)
            | Q(supporting_documents__title__icontains=search_query)
            | Q(supporting_documents__description__icontains=search_query)
            | Q(supporting_documents__reference_number__icontains=search_query)
            | Q(supporting_documents__original_filename__icontains=search_query)
        ).distinct()

    if selected_status:
        leave_queryset = leave_queryset.filter(status=selected_status)

    if selected_leave_type:
        leave_queryset = leave_queryset.filter(leave_type=selected_leave_type)

    if selected_company:
        leave_queryset = leave_queryset.filter(employee__company_id=selected_company)

    if selected_department:
        leave_queryset = leave_queryset.filter(employee__department_id=selected_department)

    if selected_branch:
        leave_queryset = leave_queryset.filter(employee__branch_id=selected_branch)

    scoped_branch = get_user_scope_branch(request.user)
    is_branch_supervisor_scope = is_branch_scoped_supervisor(request.user)
    if is_branch_supervisor_scope and scoped_branch:
        leave_queryset = leave_queryset.filter(employee__branch_id=scoped_branch.id)
        selected_branch = str(scoped_branch.id)

    leave_records = list(leave_queryset)
    request_cards = [build_employee_request_overview(leave_record) for leave_record in leave_records]
    grouped_request_weeks = build_request_overview_groups(request_cards)

    pending_total = sum(1 for leave_record in leave_records if leave_record.status == EmployeeLeave.STATUS_PENDING)
    approved_total = sum(1 for leave_record in leave_records if leave_record.status == EmployeeLeave.STATUS_APPROVED)
    rejected_total = sum(1 for leave_record in leave_records if leave_record.status == EmployeeLeave.STATUS_REJECTED)
    cancelled_total = sum(1 for leave_record in leave_records if leave_record.status == EmployeeLeave.STATUS_CANCELLED)
    waiting_supervisor_total = sum(
        1
        for leave_record in leave_records
        if leave_record.status == EmployeeLeave.STATUS_PENDING
        and leave_record.current_stage == EmployeeLeave.STAGE_SUPERVISOR_REVIEW
    )
    waiting_operations_total = sum(
        1
        for leave_record in leave_records
        if leave_record.status == EmployeeLeave.STATUS_PENDING
        and leave_record.current_stage == EmployeeLeave.STAGE_OPERATIONS_REVIEW
    )
    waiting_hr_total = sum(
        1
        for leave_record in leave_records
        if leave_record.status == EmployeeLeave.STATUS_PENDING
        and leave_record.current_stage == EmployeeLeave.STAGE_HR_REVIEW
    )
    documents_total = sum(card["supporting_documents_count"] for card in request_cards)

    employee_document_queryset = EmployeeDocument.objects.select_related(
        "employee",
        "employee__company",
        "employee__department",
        "employee__branch",
        "employee__section",
        "employee__job_title",
        "linked_leave",
    ).filter(employee__in=get_employee_directory_queryset_for_user(request.user)).order_by("-uploaded_at", "-id")

    if search_query:
        employee_document_queryset = employee_document_queryset.filter(
            Q(employee__full_name__icontains=search_query)
            | Q(employee__employee_id__icontains=search_query)
            | Q(employee__email__icontains=search_query)
            | Q(title__icontains=search_query)
            | Q(description__icontains=search_query)
            | Q(reference_number__icontains=search_query)
            | Q(original_filename__icontains=search_query)
        )

    if selected_company:
        employee_document_queryset = employee_document_queryset.filter(employee__company_id=selected_company)

    if selected_department:
        employee_document_queryset = employee_document_queryset.filter(employee__department_id=selected_department)

    if selected_branch:
        employee_document_queryset = employee_document_queryset.filter(employee__branch_id=selected_branch)

    employee_documents = list(employee_document_queryset)
    expanded_document_groups = set(request.GET.getlist("document_group"))
    latest_employee_document_groups = build_management_document_group_cards(
        employee_documents,
        latest_limit=3,
        expanded_group_keys=expanded_document_groups,
    )
    employee_documents_total = len(employee_documents)

    for document_group in latest_employee_document_groups:
        group_query = request.GET.copy()
        group_query.pop("page", None)
        current_group_keys = [value for value in group_query.getlist("document_group") if value != document_group["key"]]
        group_query.setlist("document_group", current_group_keys)

        if not document_group["is_expanded"]:
            group_query.setlist("document_group", current_group_keys + [document_group["key"]])

        group_querystring = group_query.urlencode()
        document_group["toggle_url"] = (
            f"{reverse('employees:employee_requests_overview')}?{group_querystring}"
            if group_querystring
            else reverse("employees:employee_requests_overview")
        )

    current_leave_review_stage_label = "No active leave review stage"
    if is_branch_supervisor_scope:
        current_leave_review_stage_label = "Supervisor review queue"
    elif is_operations_manager_user(request.user):
        current_leave_review_stage_label = "Operations review queue"
    elif is_hr_user(request.user) or is_admin_compatible(request.user):
        current_leave_review_stage_label = "HR final review queue"

    my_stage_pending_total = sum(1 for leave_record in leave_records if can_user_review_leave_stage(request.user, leave_record))

    required_submission_queryset = EmployeeRequiredSubmission.objects.select_related(
        "employee",
        "employee__company",
        "employee__department",
        "employee__branch",
        "employee__section",
        "employee__job_title",
        "created_by",
        "reviewed_by",
        "fulfilled_document",
    ).filter(employee__in=get_employee_directory_queryset_for_user(request.user)).order_by("-updated_at", "-created_at", "-id")

    submission_request_total = required_submission_queryset.count()
    submission_requested_total = required_submission_queryset.filter(status=EmployeeRequiredSubmission.STATUS_REQUESTED).count()
    submission_submitted_total = required_submission_queryset.filter(status=EmployeeRequiredSubmission.STATUS_SUBMITTED).count()
    submission_completed_total = required_submission_queryset.filter(status=EmployeeRequiredSubmission.STATUS_COMPLETED).count()
    latest_submission_requests = list(required_submission_queryset[:12])

    employee_document_request_queryset = EmployeeDocumentRequest.objects.select_related(
        "employee",
        "employee__company",
        "employee__department",
        "employee__branch",
        "employee__section",
        "employee__job_title",
        "created_by",
        "reviewed_by",
        "delivered_document",
    ).filter(employee__in=get_employee_directory_queryset_for_user(request.user)).order_by("-updated_at", "-created_at", "-id")

    if search_query:
        employee_document_request_queryset = employee_document_request_queryset.filter(
            Q(employee__full_name__icontains=search_query)
            | Q(employee__employee_id__icontains=search_query)
            | Q(title__icontains=search_query)
            | Q(request_note__icontains=search_query)
            | Q(management_note__icontains=search_query)
        )

    if selected_company:
        employee_document_request_queryset = employee_document_request_queryset.filter(employee__company_id=selected_company)
    if selected_department:
        employee_document_request_queryset = employee_document_request_queryset.filter(employee__department_id=selected_department)
    if selected_branch:
        employee_document_request_queryset = employee_document_request_queryset.filter(employee__branch_id=selected_branch)

    employee_document_requests = list(employee_document_request_queryset)
    for document_request in employee_document_requests:
        if can_review_employee_document_request(request.user, document_request):
            document_request.review_form = EmployeeDocumentRequestReviewForm(instance=document_request)

    employee_document_request_total = len(employee_document_requests)
    employee_document_request_requested_total = sum(1 for document_request in employee_document_requests if document_request.status == EmployeeDocumentRequest.STATUS_REQUESTED)
    employee_document_request_approved_total = sum(1 for document_request in employee_document_requests if document_request.status == EmployeeDocumentRequest.STATUS_APPROVED)
    employee_document_request_completed_total = sum(1 for document_request in employee_document_requests if document_request.status == EmployeeDocumentRequest.STATUS_COMPLETED)
    employee_document_request_rejected_total = sum(1 for document_request in employee_document_requests if document_request.status == EmployeeDocumentRequest.STATUS_REJECTED)
    employee_document_request_cancelled_total = sum(1 for document_request in employee_document_requests if document_request.status == EmployeeDocumentRequest.STATUS_CANCELLED)

    context = {
        "request_cards": request_cards,
        "grouped_request_weeks": grouped_request_weeks,
        "request_total": len(leave_records),
        "pending_total": pending_total,
        "approved_total": approved_total,
        "rejected_total": rejected_total,
        "cancelled_total": cancelled_total,
        "waiting_supervisor_total": waiting_supervisor_total,
        "waiting_operations_total": waiting_operations_total,
        "waiting_hr_total": waiting_hr_total,
        "my_stage_pending_total": my_stage_pending_total,
        "current_leave_review_stage_label": current_leave_review_stage_label,
        "documents_total": documents_total,
        "employee_documents_total": employee_documents_total,
        "latest_employee_document_groups": latest_employee_document_groups,
        "submission_request_total": submission_request_total,
        "submission_requested_total": submission_requested_total,
        "submission_submitted_total": submission_submitted_total,
        "submission_completed_total": submission_completed_total,
        "latest_submission_requests": latest_submission_requests,
        "employee_document_requests": employee_document_requests,
        "employee_document_request_total": employee_document_request_total,
        "employee_document_request_requested_total": employee_document_request_requested_total,
        "employee_document_request_approved_total": employee_document_request_approved_total,
        "employee_document_request_completed_total": employee_document_request_completed_total,
        "employee_document_request_rejected_total": employee_document_request_rejected_total,
        "employee_document_request_cancelled_total": employee_document_request_cancelled_total,
        "search_query": search_query,
        "selected_status": selected_status,
        "selected_leave_type": selected_leave_type,
        "selected_company": selected_company,
        "selected_department": selected_department,
        "selected_branch": selected_branch,
        "status_choices": EmployeeLeave.STATUS_CHOICES,
        "leave_type_choices": EmployeeLeave.LEAVE_TYPE_CHOICES,
        "companies": Company.objects.filter(id=scoped_branch.company_id).order_by("name") if is_branch_supervisor_scope and scoped_branch else Company.objects.order_by("name"),
        "departments": Department.objects.filter(company_id=scoped_branch.company_id).select_related("company").order_by("company__name", "name") if is_branch_supervisor_scope and scoped_branch else Department.objects.select_related("company").order_by("company__name", "name"),
        "branches": Branch.objects.filter(id=scoped_branch.id).select_related("company").order_by("company__name", "name") if is_branch_supervisor_scope and scoped_branch else Branch.objects.select_related("company").order_by("company__name", "name"),
        "can_review_leave": can_review_leave(request.user),
        "is_branch_scoped_supervisor": is_branch_supervisor_scope,
        "scoped_branch": scoped_branch,
    }
    return render(request, "employees/employee_requests_overview.html", context)
